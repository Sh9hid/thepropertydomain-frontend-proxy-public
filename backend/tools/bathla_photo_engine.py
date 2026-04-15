import argparse
import asyncio
import hashlib
import json
import math
import os
import re
import sys
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import requests
from PIL import Image, ImageDraw, ImageEnhance, ImageFilter, ImageFont, ImageOps
from sqlalchemy import text

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from core.database import async_engine


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT_ROOT = ROOT / "TEST_STAGING" / "photo_engine"
SOURCE_XLSX = ROOT / "bathla_reaxml_staging.xlsx"
_PRICE_INDEX: dict[str, str] | None = None
_PRICE_LOT_INDEX: dict[str, str] | None = None


def _get_mapbox_token() -> str:
    return (
        os.getenv("MAPBOX_ACCESS_TOKEN")
        or os.getenv("MAPBOX_TOKEN")
        or os.getenv("MAPBOX_API_TOKEN")
        or ""
    ).strip()


def _load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _safe_text(value: Any) -> str:
    return str(value or "").strip()


def _norm_address(value: str) -> str:
    s = (value or "").strip().lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _money(value: Any) -> str:
    try:
        n = int(float(value))
        if n <= 0:
            return "N/A"
        return f"${n:,.0f}"
    except Exception:
        return "N/A"


def _format_price_text(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return "N/A"
    digits = re.sub(r"[^0-9]", "", raw)
    if digits:
        n = int(digits)
        if n > 0:
            return f"${n:,.0f}"
    if raw.startswith("$"):
        return raw
    return f"${raw}"


def _load_price_index() -> tuple[dict[str, str], dict[str, str]]:
    global _PRICE_INDEX, _PRICE_LOT_INDEX
    if _PRICE_INDEX is not None and _PRICE_LOT_INDEX is not None:
        return _PRICE_INDEX, _PRICE_LOT_INDEX
    by_address: dict[str, str] = {}
    by_lot: dict[str, str] = {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(SOURCE_XLSX, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip().lower() if c is not None else "" for c in next(it, [])]
        idx = {k: i for i, k in enumerate(header)}
        for row in it:
            if not row:
                continue
            address = str(row[idx.get("address", -1)] or "").strip()
            suburb = str(row[idx.get("suburb", -1)] or "").strip()
            lot = str(row[idx.get("lot_number", -1)] or "").strip()
            price = str(row[idx.get("price", -1)] or "").strip()
            if not price:
                continue
            price = _format_price_text(price)
            key = f"{_norm_address(address)}|{_norm_address(suburb)}"
            if address and suburb and key not in by_address:
                by_address[key] = price
            address_only_key = _norm_address(address)
            if address and address_only_key and address_only_key not in by_address:
                by_address[address_only_key] = price
            lot_key = f"{lot.lower()}|{_norm_address(suburb)}"
            if lot and suburb and lot_key not in by_lot:
                by_lot[lot_key] = price
            if lot and lot.lower() not in by_lot:
                by_lot[lot.lower()] = price
    except Exception:
        pass
    _PRICE_INDEX = by_address
    _PRICE_LOT_INDEX = by_lot
    return by_address, by_lot


def _to_json_array(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(v).strip() for v in raw if str(v).strip()]
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return []
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if str(v).strip()]
        except Exception:
            return []
    return []


def _download_image(url: str) -> Optional[Image.Image]:
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        return Image.open(BytesIO(r.content)).convert("RGB")
    except Exception:
        return None


def _resolve_local_image(path_text: str) -> Optional[Image.Image]:
    p = Path(path_text)
    candidates = []
    if p.is_absolute():
        candidates.append(p)
    else:
        candidates.extend([ROOT / path_text, ROOT / "backend" / path_text, ROOT / "frontend" / path_text])
    for c in candidates:
        try:
            if c.exists() and c.is_file():
                return Image.open(c).convert("RGB")
        except Exception:
            continue
    return None


def _fetch_geocoded_map(address_line: str) -> tuple[Optional[Image.Image], dict[str, Any]]:
    meta: dict[str, Any] = {"provider": "none", "lat": None, "lon": None}
    try:
        geo = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address_line, "format": "jsonv2", "limit": 1},
            headers={"User-Agent": "WoononaBathlaPhotoEngine/1.0"},
            timeout=20,
        )
        geo.raise_for_status()
        arr = geo.json() or []
        if not arr:
            return None, meta
        lat = arr[0].get("lat")
        lon = arr[0].get("lon")
        meta.update({"provider": "openstreetmap", "lat": lat, "lon": lon})
        map_resp = requests.get(
            "https://staticmap.openstreetmap.de/staticmap.php",
            params={"center": f"{lat},{lon}", "zoom": 18, "size": "1280x720", "markers": f"{lat},{lon},lightblue1"},
            timeout=20,
        )
        map_resp.raise_for_status()
        return Image.open(BytesIO(map_resp.content)).convert("RGB"), meta
    except Exception:
        return None, meta


def _lonlat_to_tile(lon: float, lat: float, zoom: int) -> tuple[float, float]:
    lat_rad = math.radians(lat)
    n = 2.0 ** zoom
    xtile = (lon + 180.0) / 360.0 * n
    ytile = (1.0 - math.asinh(math.tan(lat_rad)) / math.pi) / 2.0 * n
    return xtile, ytile


def _fetch_esri_satellite(lat: float, lon: float, target_w: int = 1600, target_h: int = 900) -> Optional[Image.Image]:
    zoom = 19
    xt, yt = _lonlat_to_tile(lon, lat, zoom)
    xc, yc = int(xt), int(yt)
    # 4x4 tile fetch for higher detail, then crop center.
    radius = 2
    tile_size = 256
    stitched = Image.new("RGB", ((radius * 2) * tile_size, (radius * 2) * tile_size), (30, 30, 30))
    got_any = False
    for dy in range(-radius, radius):
        for dx in range(-radius, radius):
            x = xc + dx
            y = yc + dy
            url = f"https://services.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{zoom}/{y}/{x}"
            try:
                r = requests.get(url, timeout=15)
                r.raise_for_status()
                tile = Image.open(BytesIO(r.content)).convert("RGB")
                stitched.paste(tile, ((dx + radius) * tile_size, (dy + radius) * tile_size))
                got_any = True
            except Exception:
                continue
    if not got_any:
        return None

    # Center crop around precise fractional tile coordinate.
    px = int((xt - (xc - radius)) * tile_size)
    py = int((yt - (yc - radius)) * tile_size)
    left = max(0, px - target_w // 2)
    top = max(0, py - target_h // 2)
    right = min(stitched.width, left + target_w)
    bottom = min(stitched.height, top + target_h)
    crop = stitched.crop((left, top, right, bottom))
    if crop.width < target_w or crop.height < target_h:
        crop = _fit_cover(crop, target_w, target_h)
    return crop


def _fetch_esri_satellite_with_meta(lat: float, lon: float, target_w: int = 1600, target_h: int = 900) -> tuple[Optional[Image.Image], dict[str, Any]]:
    zoom = 19
    xt, yt = _lonlat_to_tile(lon, lat, zoom)
    xc, yc = int(xt), int(yt)
    radius = 2
    tile_size = 256
    stitched = Image.new("RGB", ((radius * 2) * tile_size, (radius * 2) * tile_size), (30, 30, 30))
    got_any = False
    for dy in range(-radius, radius):
        for dx in range(-radius, radius):
            x = xc + dx
            y = yc + dy
            url = f"https://services.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{zoom}/{y}/{x}"
            try:
                r = requests.get(url, timeout=15)
                r.raise_for_status()
                tile = Image.open(BytesIO(r.content)).convert("RGB")
                stitched.paste(tile, ((dx + radius) * tile_size, (dy + radius) * tile_size))
                got_any = True
            except Exception:
                continue
    if not got_any:
        return None, {}

    px = int((xt - (xc - radius)) * tile_size)
    py = int((yt - (yc - radius)) * tile_size)
    left = max(0, px - target_w // 2)
    top = max(0, py - target_h // 2)
    right = min(stitched.width, left + target_w)
    bottom = min(stitched.height, top + target_h)
    crop = stitched.crop((left, top, right, bottom))
    if crop.width < target_w or crop.height < target_h:
        crop = _fit_cover(crop, target_w, target_h)
        left = 0
        top = 0
    origin_global_px_x = (xc - radius) * tile_size + left
    origin_global_px_y = (yc - radius) * tile_size + top
    meta = {
        "zoom": zoom,
        "origin_global_px_x": origin_global_px_x,
        "origin_global_px_y": origin_global_px_y,
        "width": crop.width,
        "height": crop.height,
    }
    return crop, meta


def _world_px_from_lonlat(lon: float, lat: float, zoom: float, tile_size: float = 512.0) -> tuple[float, float]:
    siny = math.sin(math.radians(lat))
    siny = min(max(siny, -0.9999), 0.9999)
    scale = tile_size * (2 ** zoom)
    x = (lon + 180.0) / 360.0 * scale
    y = (0.5 - math.log((1 + siny) / (1 - siny)) / (4 * math.pi)) * scale
    return x, y


def _fetch_mapbox_satellite_with_meta(
    lat: float,
    lon: float,
    target_w: int = 1600,
    target_h: int = 900,
    token_override: str = "",
) -> tuple[Optional[Image.Image], dict[str, Any]]:
    """Disabled — Mapbox billing removed.

    Always returns (None, provider=disabled) so the Esri satellite fallback
    in _pick_source_image takes over.  Kept as a stub so existing call sites
    don't need to change.
    """
    return None, {"provider": "mapbox", "reason": "disabled_no_billing"}


def _xy_to_lonlat(x: float, y: float, zoom: int) -> tuple[float, float]:
    n = 2.0 ** zoom
    lon = x / n * 360.0 - 180.0
    lat_rad = math.atan(math.sinh(math.pi * (1 - 2 * y / n)))
    lat = math.degrees(lat_rad)
    return lon, lat


def _distance_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(min(1, math.sqrt(a)))


def _fetch_nsw_lot_boundary(lat: float, lon: float, lot_token: str = "") -> tuple[list[tuple[float, float]], dict[str, Any]]:
    query_url = "https://maps.six.nsw.gov.au/arcgis/rest/services/public/NSW_Cadastre/MapServer/9/query"
    best_ring: list[tuple[float, float]] = []
    best_score = float("inf")
    best_meta: dict[str, Any] = {"boundary_match_method": "none"}
    for distance in [20, 35, 50, 80]:
        try:
            params = {
                "f": "pjson",
                "geometry": f"{lon},{lat}",
                "geometryType": "esriGeometryPoint",
                "inSR": 4326,
                "spatialRel": "esriSpatialRelIntersects",
                "distance": distance,
                "units": "esriSRUnit_Meter",
                "returnGeometry": "true",
                "outSR": 4326,
                "outFields": "lotnumber,planlabel,lotidstring,shape_Area",
            }
            if lot_token:
                params["where"] = f"lotnumber='{lot_token}'"
            js = requests.get(query_url, params=params, timeout=20).json()
            feats = js.get("features") or []
            if not feats:
                continue
            for feat in feats:
                attrs = feat.get("attributes") or {}
                rings = (feat.get("geometry") or {}).get("rings") or []
                for ring in rings:
                    if not ring or len(ring) < 4:
                        continue
                    pts = [(float(p[0]), float(p[1])) for p in ring if isinstance(p, (list, tuple)) and len(p) >= 2]
                    if len(pts) < 4:
                        continue
                    cx = sum(p[0] for p in pts) / len(pts)
                    cy = sum(p[1] for p in pts) / len(pts)
                    score = _distance_m(lat, lon, cy, cx)
                    if score < best_score:
                        best_score = score
                        best_ring = pts
                        best_meta = {
                            "boundary_match_method": "lot+distance" if lot_token else "nearest_distance",
                            "lotnumber": attrs.get("lotnumber"),
                            "planlabel": attrs.get("planlabel"),
                            "lotidstring": attrs.get("lotidstring"),
                            "distance_m_from_geocode": round(score, 2),
                        }
            if best_ring:
                break
        except Exception:
            continue
    return best_ring, best_meta


def _apply_boundary_focus(img: Image.Image, ring_lonlat: list[tuple[float, float]], georef: dict[str, Any]) -> Image.Image:
    if not ring_lonlat or not georef:
        return img
    zoom = int(georef.get("zoom", 19))
    tile_size = float(georef.get("tile_size", 256.0))
    ox = float(georef.get("origin_global_px_x", 0))
    oy = float(georef.get("origin_global_px_y", 0))
    w = int(georef.get("width", img.width))
    h = int(georef.get("height", img.height))

    poly: list[tuple[float, float]] = []
    for lon, lat in ring_lonlat:
        gx, gy = _world_px_from_lonlat(lon, lat, float(zoom), tile_size=tile_size)
        px = gx - ox
        py = gy - oy
        poly.append((px, py))
    if len(poly) < 3:
        return img

    outer = ImageEnhance.Contrast(ImageOps.grayscale(img).convert("RGB")).enhance(1.18)
    outer = ImageEnhance.Brightness(outer).enhance(0.55)
    outer = outer.filter(ImageFilter.GaussianBlur(radius=2))
    inner = ImageEnhance.Color(img).enhance(1.08)

    hard_mask = Image.new("L", (w, h), 0)
    ImageDraw.Draw(hard_mask).polygon(poly, fill=255)
    soft_mask = hard_mask.filter(ImageFilter.GaussianBlur(radius=3))
    mixed = Image.composite(inner, outer, soft_mask)

    accent = ImageDraw.Draw(mixed, "RGBA")
    accent.polygon(poly, fill=(207, 161, 68, 22))
    accent.polygon(poly, outline=(207, 161, 68, 90), width=14)
    accent.polygon(poly, outline=(232, 195, 110, 180), width=8)
    accent.polygon(poly, outline=(255, 241, 205, 255), width=3)
    return mixed


def _gradient_placeholder(seed_text: str, size: tuple[int, int]) -> Image.Image:
    digest = hashlib.sha256(seed_text.encode("utf-8")).hexdigest()
    a = int(digest[0:2], 16)
    b = int(digest[2:4], 16)
    c = int(digest[4:6], 16)
    d = int(digest[6:8], 16)
    w, h = size
    img = Image.new("RGB", size, (20, 20, 20))
    draw = ImageDraw.Draw(img)
    for y in range(h):
        t = y / max(1, h - 1)
        r = int((a * (1 - t) + c * t) * 0.7)
        g = int((b * (1 - t) + d * t) * 0.7)
        bl = int((c * (1 - t) + a * t) * 0.7)
        draw.line([(0, y), (w, y)], fill=(r, g, bl))
    return img


def _fit_cover(image: Image.Image, target_w: int, target_h: int) -> Image.Image:
    iw, ih = image.size
    scale = max(target_w / iw, target_h / ih)
    nw, nh = int(iw * scale), int(ih * scale)
    resized = image.resize((nw, nh), Image.Resampling.LANCZOS)
    left = (nw - target_w) // 2
    top = (nh - target_h) // 2
    return resized.crop((left, top, left + target_w, top + target_h))


def _measure_text(font: ImageFont.FreeTypeFont | ImageFont.ImageFont, text: str) -> int:
    bbox = font.getbbox(text or " ")
    return max(0, int(bbox[2] - bbox[0]))


def _wrap_text(
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    max_width: int,
    max_lines: int,
) -> list[str]:
    words = (text or "").split()
    if not words:
        return []
    lines: list[str] = []
    current = ""
    index = 0
    while index < len(words):
        word = words[index]
        trial = word if not current else f"{current} {word}"
        if _measure_text(font, trial) <= max_width or not current:
            current = trial
            index += 1
            continue
        lines.append(current)
        current = ""
        if len(lines) == max_lines - 1:
            break
    if current:
        remaining_words = [current] + words[index:]
    else:
        remaining_words = words[index:]
    final_line = " ".join(remaining_words).strip()
    if not final_line and lines:
        final_line = lines.pop()
    if not final_line:
        return lines[:max_lines]
    if _measure_text(font, final_line) <= max_width:
        lines.append(final_line)
        return lines[:max_lines]
    trimmed = final_line
    while trimmed and _measure_text(font, f"{trimmed}...") > max_width:
        next_trim = trimmed.rsplit(" ", 1)[0]
        if next_trim == trimmed:
            trimmed = trimmed[:-1]
        else:
            trimmed = next_trim
    lines.append(f"{trimmed}..." if trimmed else "...")
    return lines[:max_lines]


def _line_height(font: ImageFont.FreeTypeFont | ImageFont.ImageFont) -> int:
    bbox = font.getbbox("Ag")
    return int(bbox[3] - bbox[1])


def _draw_text_block(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    lines: list[str],
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: tuple[int, int, int, int],
    gap: int,
) -> int:
    cursor_y = y
    step = _line_height(font) + gap
    for line in lines:
        draw.text((x, cursor_y), line, fill=fill, font=font)
        cursor_y += step
    return cursor_y


def _draw_metric_pill(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    w: int,
    h: int,
    label: str,
    value: str,
) -> None:
    fill = (16, 16, 16, 220)
    outline = (207, 161, 68, 255)
    draw.rounded_rectangle([x, y, x + w, y + h], radius=24, fill=fill, outline=outline, width=2)
    label_font = _load_font(18, bold=False)
    value_font = _load_font(28, bold=True)
    draw.text((x + 22, y + 14), label.upper(), fill=(190, 190, 190, 255), font=label_font)
    draw.text((x + 22, y + 40), value or "N/A", fill=(255, 241, 214, 255), font=value_font)


def _draw_gold_frame(
    draw: ImageDraw.ImageDraw,
    left: int,
    top: int,
    right: int,
    bottom: int,
    width: int = 4,
) -> None:
    for inset, alpha in ((14, 34), (8, 56), (3, 96)):
        draw.rounded_rectangle(
            [left - inset, top - inset, right + inset, bottom + inset],
            radius=34,
            outline=(207, 161, 68, alpha),
            width=3,
        )
    draw.rounded_rectangle([left, top, right, bottom], radius=28, outline=(232, 195, 110, 255), width=width)


def _build_subtitle(description: str, listing: Optional[dict[str, Any]] = None) -> str:
    cleaned = re.sub(r"\s+", " ", (description or "").strip())
    if cleaned:
        sentence = re.split(r"(?<=[.!?])\s+", cleaned)[0]
        return sentence[:150].rstrip(" .,") + ("..." if len(sentence) > 150 else "")
    if listing:
        return (
            f"{_safe_text(listing.get('property_type')).title() or 'Land'} parcel in "
            f"{_safe_text(listing.get('suburb')) or 'NSW'}"
        )
    return "Premium land opportunity"


def _format_land_size(value: Any) -> str:
    try:
        size = int(float(value))
        return f"{size:,} sqm" if size > 0 else "N/A"
    except Exception:
        return "N/A"


def _calculate_boundary_confidence(boundary_meta: dict[str, Any]) -> dict[str, Any]:
    method = _safe_text(boundary_meta.get("boundary_match_method")).lower()
    distance_raw = boundary_meta.get("distance_m_from_geocode")
    try:
        distance_m = float(distance_raw) if distance_raw is not None else None
    except Exception:
        distance_m = None

    base_score = {
        "lot+distance": 0.96,
        "nearest_distance": 0.78,
        "none": 0.0,
    }.get(method, 0.62 if boundary_meta.get("nsw_lot_boundary_found") else 0.0)

    if distance_m is not None:
        if distance_m > 5:
            base_score -= min(0.32, (distance_m - 5) / 100.0)
        elif distance_m <= 2:
            base_score += 0.02
    score = max(0.0, min(0.99, round(base_score, 2)))

    if score >= 0.85:
        label = "high"
    elif score >= 0.6:
        label = "medium"
    elif score > 0:
        label = "low"
    else:
        label = "unavailable"

    return {
        "score": score,
        "label": label,
        "method": method or "none",
        "distance_m_from_geocode": round(distance_m, 2) if distance_m is not None else None,
    }


def _save_upload_ready_jpg(image: Image.Image, path: Path) -> None:
    rgb = image.convert("RGB")
    rgb.save(
        path,
        format="JPEG",
        quality=95,
        optimize=True,
        progressive=True,
        subsampling=0,
        dpi=(300, 300),
    )


def _compose_card(base: Image.Image, title: str, subtitle: str, metrics: dict[str, str]) -> Image.Image:
    canvas_w, canvas_h = 1600, 900
    canvas = Image.new("RGB", (canvas_w, canvas_h), (11, 11, 11))

    bg = _fit_cover(base, canvas_w, canvas_h)
    bg = ImageEnhance.Contrast(ImageOps.grayscale(bg).convert("RGB")).enhance(1.15)
    bg = ImageEnhance.Brightness(bg).enhance(0.48)
    bg = bg.filter(ImageFilter.GaussianBlur(radius=5))
    canvas.paste(bg, (0, 0))

    overlay = Image.new("RGBA", (canvas_w, canvas_h), (7, 7, 7, 86))
    canvas = Image.alpha_composite(canvas.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(canvas, "RGBA")

    draw.rounded_rectangle([34, 34, canvas_w - 34, canvas_h - 34], radius=34, outline=(61, 61, 61, 255), width=1)

    frame_left, frame_top = 84, 86
    frame_w, frame_h = 1432, 728
    fg = _fit_cover(base, frame_w, frame_h)
    canvas.paste(fg, (frame_left, frame_top))
    _draw_gold_frame(draw, frame_left, frame_top, frame_left + frame_w, frame_top + frame_h, width=4)

    panel_left = 118
    panel_top = 116
    panel_w = 620
    panel_h = 256
    draw.rounded_rectangle(
        [panel_left, panel_top, panel_left + panel_w, panel_top + panel_h],
        radius=30,
        fill=(8, 8, 8, 205),
        outline=(207, 161, 68, 180),
        width=2,
    )
    draw.text((panel_left + 30, panel_top + 24), "PREMIUM LAND RELEASE", fill=(207, 161, 68, 255), font=_load_font(20))

    title_font = _load_font(54, bold=True)
    subtitle_font = _load_font(28, bold=False)
    title_lines = _wrap_text(title, title_font, panel_w - 60, 2)
    subtitle_lines = _wrap_text(subtitle, subtitle_font, panel_w - 60, 2)
    cursor_y = _draw_text_block(
        draw,
        panel_left + 30,
        panel_top + 62,
        title_lines,
        title_font,
        (255, 247, 233, 255),
        8,
    )
    _draw_text_block(
        draw,
        panel_left + 30,
        cursor_y + 10,
        subtitle_lines,
        subtitle_font,
        (219, 219, 219, 255),
        8,
    )

    pill_y = canvas_h - 174
    _draw_metric_pill(draw, 118, pill_y, 280, 100, "Price", metrics.get("price", "N/A"))
    _draw_metric_pill(draw, 418, pill_y, 240, 100, "Land Size", metrics.get("land_size", "N/A"))
    _draw_metric_pill(draw, 678, pill_y, 236, 100, "Status", metrics.get("status", "N/A"))

    draw.text((canvas_w - 318, canvas_h - 130), "UPLOAD-READY JPG", fill=(207, 207, 207, 255), font=_load_font(24))
    draw.text((canvas_w - 318, canvas_h - 94), "MONO FRAME | COLOR PARCEL", fill=(207, 161, 68, 255), font=_load_font(20))
    return canvas.convert("RGB")


def _compose_detail_sheet(
    base: Image.Image,
    title: str,
    subtitle: str,
    listing: dict[str, Any],
    metrics: dict[str, str],
) -> Image.Image:
    w, h = 1600, 900
    canvas = Image.new("RGB", (w, h), (12, 12, 12))
    bg = _fit_cover(base, w, h)
    bg = ImageEnhance.Contrast(ImageOps.grayscale(bg).convert("RGB")).enhance(1.08)
    bg = ImageEnhance.Brightness(bg).enhance(0.36)
    bg = bg.filter(ImageFilter.GaussianBlur(radius=7))
    canvas.paste(bg, (0, 0))
    canvas = Image.alpha_composite(canvas.convert("RGBA"), Image.new("RGBA", (w, h), (0, 0, 0, 52)))
    draw = ImageDraw.Draw(canvas, "RGBA")

    panel_left, panel_top = 64, 60
    panel_right, panel_bottom = w - 64, h - 60
    draw.rounded_rectangle(
        [panel_left, panel_top, panel_right, panel_bottom],
        radius=34,
        fill=(8, 8, 8, 216),
        outline=(232, 195, 110, 255),
        width=3,
    )

    image_left, image_top = 92, 104
    image_w, image_h = 880, 560
    canvas.paste(_fit_cover(base, image_w, image_h), (image_left, image_top))
    _draw_gold_frame(draw, image_left, image_top, image_left + image_w, image_top + image_h, width=4)

    right_x = 1030
    draw.text((right_x, 112), "PREMIUM LAND LISTING", fill=(207, 161, 68, 255), font=_load_font(22))
    title_lines = _wrap_text(title, _load_font(44, bold=True), 456, 3)
    subtitle_lines = _wrap_text(subtitle, _load_font(24, bold=False), 456, 3)
    cursor_y = _draw_text_block(draw, right_x, 150, title_lines, _load_font(44, bold=True), (255, 246, 230, 255), 8)
    cursor_y = _draw_text_block(draw, right_x, cursor_y + 12, subtitle_lines, _load_font(24, bold=False), (218, 218, 218, 255), 6)

    _draw_metric_pill(draw, right_x, cursor_y + 26, 456, 94, "Price", metrics.get("price", "N/A"))
    _draw_metric_pill(draw, right_x, cursor_y + 136, 220, 94, "Land Size", metrics.get("land_size", "N/A"))
    _draw_metric_pill(draw, right_x + 236, cursor_y + 136, 220, 94, "Status", metrics.get("status", "N/A"))

    fields = [
        ("Address", _safe_text(listing.get("address"))),
        ("Location", " ".join(v for v in [_safe_text(listing.get("suburb")), _safe_text(listing.get("postcode"))] if v)),
        ("Property Type", _safe_text(listing.get("property_type")) or "land"),
    ]
    field_y = cursor_y + 248
    label_font = _load_font(20, bold=False)
    value_font = _load_font(24, bold=True)
    for label, value in fields:
        draw.text((right_x, field_y), label.upper(), fill=(161, 161, 161, 255), font=label_font)
        draw.text((right_x, field_y + 22), value or "N/A", fill=(252, 247, 235, 255), font=value_font)
        field_y += 58

    footer_y = 786
    draw.line([(92, footer_y), (panel_right - 28, footer_y)], fill=(86, 86, 86, 255), width=1)
    draw.text((92, footer_y + 24), "PROFESSIONAL MONOCHROME PRESENTATION WITH PARCEL HIGHLIGHT", fill=(193, 193, 193, 255), font=_load_font(20))
    draw.text((92, footer_y + 56), "OUTPUT: HIGH-QUALITY JPEG FOR LISTING UPLOADS", fill=(207, 161, 68, 255), font=_load_font(20))
    return canvas.convert("RGB")


def _compose_square_thumb(base: Image.Image, title: str, subtitle: str, metrics: dict[str, str]) -> Image.Image:
    size = 1080
    canvas = Image.new("RGB", (size, size), (12, 12, 12))
    bg = _fit_cover(base, size, size)
    bg = ImageEnhance.Contrast(ImageOps.grayscale(bg).convert("RGB")).enhance(1.12)
    bg = ImageEnhance.Brightness(bg).enhance(0.44)
    bg = bg.filter(ImageFilter.GaussianBlur(radius=4))
    canvas.paste(bg, (0, 0))
    canvas = Image.alpha_composite(canvas.convert("RGBA"), Image.new("RGBA", (size, size), (0, 0, 0, 58)))
    draw = ImageDraw.Draw(canvas, "RGBA")

    img_left, img_top = 72, 72
    img_size = 936
    canvas.paste(_fit_cover(base, img_size, img_size), (img_left, img_top))
    _draw_gold_frame(draw, img_left, img_top, img_left + img_size, img_top + img_size, width=5)

    panel_left = 118
    panel_top = 690
    panel_w = 844
    panel_h = 246
    draw.rounded_rectangle(
        [panel_left, panel_top, panel_left + panel_w, panel_top + panel_h],
        radius=30,
        fill=(8, 8, 8, 214),
        outline=(207, 161, 68, 180),
        width=2,
    )

    title_font = _load_font(44, bold=True)
    subtitle_font = _load_font(22, bold=False)
    title_lines = _wrap_text(title, title_font, panel_w - 56, 2)
    subtitle_lines = _wrap_text(subtitle, subtitle_font, panel_w - 56, 2)
    cursor_y = _draw_text_block(draw, panel_left + 28, panel_top + 24, title_lines, title_font, (255, 245, 228, 255), 6)
    _draw_text_block(draw, panel_left + 28, cursor_y + 6, subtitle_lines, subtitle_font, (215, 215, 215, 255), 6)

    _draw_metric_pill(draw, panel_left + 28, panel_top + 138, 250, 82, "Price", metrics.get("price", "N/A"))
    _draw_metric_pill(draw, panel_left + 296, panel_top + 138, 220, 82, "Land Size", metrics.get("land_size", "N/A"))
    _draw_metric_pill(draw, panel_left + 534, panel_top + 138, 184, 82, "Status", metrics.get("status", "N/A"))
    return canvas.convert("RGB")


async def _get_listing(lead_id: Optional[str]) -> dict[str, Any]:
    async with async_engine.begin() as conn:
        if lead_id:
            row = (
                await conn.execute(
                    text(
                        """
                        SELECT id, address, suburb, postcode, listing_headline, listing_description, lot_number,
                               property_type, est_value, land_size_sqm, last_listing_status,
                               main_image,
                               CASE WHEN property_images IS NULL THEN '[]' ELSE CAST(property_images AS text) END AS property_images
                        FROM leads
                        WHERE id = :lead_id
                        LIMIT 1
                        """
                    ),
                    {"lead_id": lead_id},
                )
            ).mappings().first()
            if row:
                return dict(row)

        row = (
            await conn.execute(
                text(
                    """
                    SELECT id, address, suburb, postcode, listing_headline, listing_description, lot_number,
                           property_type, est_value, land_size_sqm, last_listing_status,
                           main_image,
                           CASE WHEN property_images IS NULL THEN '[]' ELSE CAST(property_images AS text) END AS property_images
                    FROM leads
                    WHERE trigger_type = 'bathla_land'
                    ORDER BY updated_at DESC NULLS LAST, created_at DESC NULLS LAST
                    LIMIT 1
                    """
                )
            )
        ).mappings().first()
        if not row:
            raise RuntimeError("No bathla_land listing found.")
        return dict(row)


def _pick_source_image(
    listing: dict[str, Any],
    mapbox_token: str = "",
    strict_mapbox: bool = False,
) -> tuple[Image.Image, dict[str, Any]]:
    main_image = _safe_text(listing.get("main_image"))
    image_list = _to_json_array(listing.get("property_images"))
    candidates = [main_image] + image_list
    for candidate in candidates:
        if not candidate:
            continue
        if candidate.startswith("http://") or candidate.startswith("https://"):
            img = _download_image(candidate)
            if img:
                return img, {"image_source": "listing_url", "image_ref": candidate}
        else:
            img = _resolve_local_image(candidate)
            if img:
                return img, {"image_source": "listing_local", "image_ref": candidate}

    address_line = ", ".join(
        [x for x in [_safe_text(listing.get("address")), _safe_text(listing.get("suburb")), "NSW", "Australia"] if x]
    )

    # Try geocode + Esri satellite imagery first for real land photo context.
    lot_number = _safe_text(listing.get("lot_number"))
    if not lot_number:
        m = re.search(r"\blot\s+(\d+)\b", _safe_text(listing.get("listing_headline")), flags=re.I)
        lot_number = m.group(1) if m else ""
    try:
        geo = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address_line, "format": "jsonv2", "limit": 1},
            headers={"User-Agent": "WoononaBathlaPhotoEngine/1.0"},
            timeout=20,
        )
        geo.raise_for_status()
        arr = geo.json() or []
        if arr:
            lat = float(arr[0].get("lat"))
            lon = float(arr[0].get("lon"))
            # Prefer Mapbox when token exists, then fall back to Esri.
            mapbox_img, mapbox_meta = _fetch_mapbox_satellite_with_meta(lat, lon, token_override=mapbox_token)
            if mapbox_img:
                ring, ring_meta = _fetch_nsw_lot_boundary(lat, lon, lot_number)
                if not ring:
                    ring, ring_meta = _fetch_nsw_lot_boundary(lat, lon, "")
                if ring:
                    mapbox_img = _apply_boundary_focus(mapbox_img, ring, mapbox_meta)
                return mapbox_img, {
                    "image_source": "mapbox_satellite_v9",
                    "image_ref": {
                        "lat": lat,
                        "lon": lon,
                        "provider": "Mapbox",
                        "nsw_lot_boundary_found": bool(ring),
                        "lot_number_input": lot_number or None,
                        **ring_meta,
                    },
                }

            if strict_mapbox:
                fallback = _gradient_placeholder(f"{listing.get('id')}|{address_line}", (1280, 720))
                return fallback, {
                    "image_source": "strict_mapbox_failed",
                    "image_ref": {"lat": lat, "lon": lon, "mapbox_attempt": mapbox_meta},
                }

            sat, sat_meta = _fetch_esri_satellite_with_meta(lat, lon)
            if sat:
                ring, ring_meta = _fetch_nsw_lot_boundary(lat, lon, lot_number)
                if not ring:
                    ring, ring_meta = _fetch_nsw_lot_boundary(lat, lon, "")
                if ring:
                    sat = _apply_boundary_focus(sat, ring, sat_meta)
                return sat, {
                    "image_source": "esri_world_imagery",
                    "image_ref": {
                        "lat": lat,
                        "lon": lon,
                        "provider": "ArcGIS World_Imagery",
                        "nsw_lot_boundary_found": bool(ring),
                        "lot_number_input": lot_number or None,
                        **ring_meta,
                        "mapbox_attempt": mapbox_meta,
                    },
                }
    except Exception:
        pass

    map_img, map_meta = _fetch_geocoded_map(address_line)
    if map_img:
        return map_img, {"image_source": "osm_static_map", "image_ref": map_meta}

    fallback = _gradient_placeholder(f"{listing.get('id')}|{address_line}", (1280, 720))
    return fallback, {"image_source": "generated_gradient", "image_ref": ""}


async def run(
    lead_id: Optional[str],
    out_root: Path,
    title_override: str = "",
    description_override: str = "",
    price_override: str = "",
    land_size_override: str = "",
    mapbox_token: str = "",
    strict_mapbox: bool = False,
) -> Path:
    listing = await _get_listing(lead_id)
    lot_for_title = _safe_text(listing.get("lot_number"))
    if not lot_for_title:
        m = re.search(r"\blot\s+(\d+)\b", _safe_text(listing.get("listing_headline")), flags=re.I)
        lot_for_title = m.group(1) if m else ""
    title = f"Lot {lot_for_title} | {_safe_text(listing.get('address'))}, {_safe_text(listing.get('suburb'))}".strip()
    if title.startswith("Lot  |"):
        title = f"{_safe_text(listing.get('address'))}, {_safe_text(listing.get('suburb'))}"
    if title_override.strip():
        title = title_override.strip()
    desc = _safe_text(listing.get("listing_description"))
    if not desc:
        desc = f"{_safe_text(listing.get('address'))}, {_safe_text(listing.get('suburb'))} {_safe_text(listing.get('postcode'))}. Land listing."
    if description_override.strip():
        desc = description_override.strip()
    subtitle = _build_subtitle(desc, listing)

    base_img, src_meta = _pick_source_image(listing, mapbox_token=mapbox_token, strict_mapbox=strict_mapbox)
    boundary_meta = src_meta.get("image_ref") if isinstance(src_meta.get("image_ref"), dict) else {}
    boundary_confidence = _calculate_boundary_confidence(boundary_meta)
    if isinstance(boundary_meta, dict):
        boundary_meta["boundary_confidence"] = boundary_confidence
    address_norm = _norm_address(_safe_text(listing.get("address")))
    suburb_norm = _norm_address(_safe_text(listing.get("suburb")))
    address_key = f"{address_norm}|{suburb_norm}"
    by_address, by_lot = _load_price_index()
    lot_match = re.search(r"\blot\s+(\d+)\b", title, flags=re.I)
    lot_token = lot_match.group(1).lower() if lot_match else ""
    lot_key = f"{lot_token}|{suburb_norm}" if lot_token else ""
    price_text = _money(listing.get("est_value"))
    if price_text == "N/A":
        if address_key in by_address:
            price_text = by_address[address_key]
        elif address_norm in by_address:
            price_text = by_address[address_norm]
        elif lot_key and lot_key in by_lot:
            price_text = by_lot[lot_key]
        elif lot_token and lot_token in by_lot:
            price_text = by_lot[lot_token]
    if price_override.strip():
        price_text = price_override.strip()
    price_text = _format_price_text(price_text)

    status_token = _safe_text(listing.get("last_listing_status")).lower()
    status_label = "Available" if status_token in {"", "captured", "current", "active"} else status_token.title()
    land_size_text = _format_land_size(listing.get("land_size_sqm"))
    if land_size_override.strip():
        land_size_text = land_size_override.strip()
    metrics = {
        "price": price_text,
        "land_size": land_size_text,
        "status": status_label,
    }
    hero = _compose_card(base_img, title, subtitle, metrics)
    detail = _compose_detail_sheet(base_img, title, subtitle, listing, metrics)
    square = _compose_square_thumb(base_img, title, subtitle, metrics)

    lead_token = _safe_text(listing.get("id")) or "unknown"
    out_dir = out_root / f"bathla_{lead_token}"
    out_dir.mkdir(parents=True, exist_ok=True)

    hero_path = out_dir / "upload_hero_16x9.jpg"
    detail_path = out_dir / "upload_detail_16x9.jpg"
    square_path = out_dir / "upload_square_1x1.jpg"
    meta_path = out_dir / "manifest.json"
    engine_path = out_dir / "engine.json"

    _save_upload_ready_jpg(hero, hero_path)
    _save_upload_ready_jpg(detail, detail_path)
    _save_upload_ready_jpg(square, square_path)

    manifest = {
        "engine": "bathla_photo_engine_v2",
        "repeatable": True,
        "lead_id": lead_token,
        "title": title,
        "subtitle": subtitle,
        "description": desc,
        "address": _safe_text(listing.get("address")),
        "suburb": _safe_text(listing.get("suburb")),
        "postcode": _safe_text(listing.get("postcode")),
        "property_type": _safe_text(listing.get("property_type")),
        "metrics": metrics,
        "boundary_confidence": boundary_confidence,
        "source_image": src_meta,
        "upload_ready": True,
        "jpeg_export": {
            "quality": 95,
            "optimize": True,
            "progressive": True,
            "subsampling": "4:4:4",
            "dpi": [300, 300],
        },
        "output_files": {
            "hero_16x9_jpg": hero_path.name,
            "detail_16x9_jpg": detail_path.name,
            "square_1x1_jpg": square_path.name,
        },
    }
    meta_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    engine_cfg = {
        "name": "bathla_photo_engine_v2",
        "repeatable": True,
        "notes": "Deterministic output folder and composition. Re-run with same lead_id for same result class.",
        "args": {"lead_id": lead_id, "out_root": str(out_root)},
    }
    engine_path.write_text(json.dumps(engine_cfg, indent=2), encoding="utf-8")

    base_prompt = (
        "Generate a polished listing hero image that follows rea guidelines but go as much as dopamine realise hooking retention photo "
        "(go full mr. beast style thumbnail while keeping it professional) using this land satellite reference. "
        f"Text: '{title}' and '{desc}'. "
        f"Display price '{metrics['price']}' and land size '{metrics['land_size']}'. "
        "Keep background monochrome except everything inside highlighted parcel zone, with premium golden border. "
        "Output as clean ad-ready JPG."
    )
    sora_prompt = base_prompt
    nano_prompt = base_prompt
    (out_dir / "sora_prompt.txt").write_text(sora_prompt, encoding="utf-8")
    (out_dir / "nanobanana_prompt.txt").write_text(nano_prompt, encoding="utf-8")

    return out_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="Build repeatable hero-card creative for one Bathla land listing.")
    parser.add_argument("--lead-id", default="", help="Optional lead id. Defaults to latest bathla_land record.")
    parser.add_argument("--out-root", default=str(DEFAULT_OUT_ROOT), help="Output root folder.")
    parser.add_argument("--title", default="", help="Override title text.")
    parser.add_argument("--description", default="", help="Override description text.")
    parser.add_argument("--price", default="", help="Override price text.")
    parser.add_argument("--land-size", default="", help="Override land size text.")
    parser.add_argument("--mapbox-token", default="", help="Mapbox access token override.")
    parser.add_argument("--strict-mapbox", action="store_true", help="Disable non-Mapbox fallbacks.")
    args = parser.parse_args()
    out_dir = asyncio.run(
        run(
            args.lead_id or None,
            Path(args.out_root),
            title_override=args.title,
            description_override=args.description,
            price_override=args.price,
            land_size_override=args.land_size,
            mapbox_token=args.mapbox_token,
            strict_mapbox=bool(args.strict_mapbox),
        )
    )
    print(str(out_dir))


if __name__ == "__main__":
    main()
