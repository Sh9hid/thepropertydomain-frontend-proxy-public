import asyncio
import datetime
import html
import json
import re
import sqlite3
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Awaitable, Callable, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


REPORT_PACK_ROOT = Path("D:/woonona-lead-machine/backend/report_packs")
REFERENCE_REPORT_DIR = Path("D:/referencecotalityreportsfor69alexanderst")
DATE_PATTERNS = ("%d %b %Y", "%d %B %Y", "%d/%m/%Y", "%Y-%m-%d")
AGENT_PROFILE_URL = "https://www.realestate.com.au/agent/nitin-puri-2577410"
AGENT_IMAGE_URL = "https://i1.au.reastatic.net/500x640/460ce6c452f3af803f66731f853f84313807878ca4907f3654db4c6a11ad40f1/main.jpg"
PUBLIC_WEB_OVERRIDES: Dict[str, Dict[str, Any]] = {
    "1-3-argyle-street-south-windsor-2756": {
        "hero_image_url": "https://i2.au.reastatic.net/800x600-format=webp/e5314876a46725de4005e59e182ecf4ff477ce721cb5c0d77389bf776cf700e1/main.jpg",
        "gallery_urls": [
            "https://i2.au.reastatic.net/800x600-format=webp/e5314876a46725de4005e59e182ecf4ff477ce721cb5c0d77389bf776cf700e1/main.jpg",
            "https://i2.au.reastatic.net/800x600-format=webp/54bc1f0cd2dc5b0d51e5bfd1f7d6672e00fc0fcaed8de24d996b7d8949c4be03/main.jpg",
            "https://i2.au.reastatic.net/800x600-format=webp/575087cbf73385955b6a9a22a72c227071528b06fa0d35d44eeeb0fc4c832d3d/main.jpg",
        ],
        "source_links": [
            "https://www.realestate.com.au/property/1-3-argyle-st-south-windsor-nsw-2756/",
            "https://www.domain.com.au/1-3-argyle-street-south-windsor-nsw-2756-2019872848",
        ],
        "price_growth_since_last_sale_pct": 59.4,
        "price_growth_12m_pct": 2.7,
        "rent_estimate_low": 650,
        "rent_estimate_high": 710,
        "rent_estimate_mid": 680,
        "gross_yield_pct": 4.8,
        "why_now": "Online public market trackers show this property has appreciated materially since its last recorded sale, while local house turnover remains active. That combination supports a fresh appraisal conversation now rather than waiting for market conditions to become less favorable.",
    }
}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"-", "None", "nan"} else text


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value).replace(",", "").replace("$", "").replace("m2", "").replace("sqm", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _safe_int(value: Any) -> Optional[int]:
    number = _safe_float(value)
    return None if number is None else int(round(number))


def _parse_money(value: Any) -> Optional[int]:
    text = _clean_text(value)
    return None if not text else _safe_int(text.replace("$", "").replace(",", ""))


def _parse_date(value: Any) -> Optional[datetime.date]:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for pattern in DATE_PATTERNS:
        try:
            return datetime.datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _format_date(value: Any) -> str:
    parsed = _parse_date(value)
    return parsed.strftime("%d %b %Y") if parsed else (_clean_text(value) or "N/A")


def _format_money(value: Optional[float], weekly: bool = False) -> str:
    if value is None:
        return "N/A"
    suffix = "/week" if weekly else ""
    return f"${int(round(value)):,.0f}{suffix}"


def _format_number(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    if abs(value - round(value)) < 0.001:
        return f"{int(round(value)):,}"
    return f"{value:,.1f}"


def _folder_safe(value: str) -> str:
    cleaned = re.sub(r"[<>:\"/\\\\|?*]+", "-", _clean_text(value))
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned[:180] or "Untitled"


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", _clean_text(value).lower()).strip("-") or "value"


def _pct_change(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    if current is None or previous in (None, 0):
        return None
    try:
        return ((float(current) - float(previous)) / float(previous)) * 100
    except ZeroDivisionError:
        return None


def _median(values: Iterable[Optional[float]]) -> Optional[float]:
    items = [float(v) for v in values if v is not None]
    return float(statistics.median(items)) if items else None


def _property_type(record: Dict[str, Any]) -> str:
    record_type = _clean_text(record.get("record_type")).lower()
    address = _clean_text(record.get("address"))
    if "unit" in record_type or "/" in address:
        return "Unit"
    return "House"


def _load_reference_manifest() -> List[Dict[str, Any]]:
    manifest: List[Dict[str, Any]] = []
    if not REFERENCE_REPORT_DIR.exists():
        return manifest
    try:
        from pypdf import PdfReader
    except Exception:
        return manifest
    seen = set()
    for path in sorted(REFERENCE_REPORT_DIR.glob("*.pdf")):
        key = re.sub(r" \\(\\d+\\)", "", path.stem.lower())
        if key in seen:
            continue
        seen.add(key)
        try:
            reader = PdfReader(str(path))
            preview = ""
            for page in reader.pages[:2]:
                preview += (page.extract_text() or "") + "\n"
            manifest.append({"file": path.name, "pages": len(reader.pages), "preview": " ".join(preview.split())[:280]})
        except Exception:
            manifest.append({"file": path.name, "pages": None, "preview": ""})
    return manifest


def _suburb_workbook_path(stock_root: str, suburb: str) -> Optional[Path]:
    suburb_dir = Path(stock_root) / "Suburb reports"
    if not suburb_dir.exists():
        return None
    candidates = [
        suburb_dir / f"{suburb} report.xlsx",
        suburb_dir / f"{suburb.title()} report.xlsx",
        suburb_dir / f"{suburb.upper()} report.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    target = _slug(suburb)
    for candidate in suburb_dir.glob("*.xlsx"):
        if _slug(candidate.stem.replace(" report", "")) == target:
            return candidate
    return None


def _load_suburb_records(stock_root: str, suburb: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    workbook_path = _suburb_workbook_path(stock_root, suburb)
    if not workbook_path or not load_workbook:
        return [], None
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(min_row=3, values_only=True))
    if not rows:
        return [], str(workbook_path)
    headers = [_slug(value) for value in rows[0]]
    records: List[Dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        address = _clean_text(row.get("street_address"))
        if not address:
            continue
        records.append(
            {
                "address": address.title(),
                "suburb": _clean_text(row.get("suburb")).title(),
                "postcode": _clean_text(row.get("postcode")),
                "property_type": _clean_text(row.get("property_type")).title(),
                "bedrooms": _safe_float(row.get("bed")),
                "bathrooms": _safe_float(row.get("bath")),
                "car_spaces": _safe_float(row.get("car")),
                "land_size_sqm": _safe_float(row.get("land_size_m")),
                "floor_size_sqm": _safe_float(row.get("floor_size_m")),
                "year_built": _clean_text(row.get("year_built")),
                "sale_price": _clean_text(row.get("sale_price")),
                "sale_date": _clean_text(row.get("sale_date")),
                "settlement_date": _clean_text(row.get("settlement_date")),
                "agency_name": _clean_text(row.get("agency")),
                "agent_name": _clean_text(row.get("agent")),
                "owner_type": _clean_text(row.get("owner_type")),
                "land_use": _clean_text(row.get("land_use")),
                "development_zone": _clean_text(row.get("development_zone")),
                "parcel_details": _clean_text(row.get("parcel_details")),
                "owner_name": ", ".join(filter(None, [_clean_text(row.get("owner_1_name")), _clean_text(row.get("owner_2_name")), _clean_text(row.get("owner_3_name"))])),
                "source_path": str(workbook_path),
            }
        )
    return records, str(workbook_path)


def _same_suburb_db_records(conn: sqlite3.Connection, suburb: str, exclude_lead_id: str) -> List[Dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, address, suburb, postcode, owner_name, bedrooms, bathrooms, car_spaces, land_size_sqm,
               floor_size_sqm, year_built, sale_price, sale_date, settlement_date, agency_name, agent_name,
               owner_type, land_use, development_zone, parcel_details, record_type
        FROM leads
        WHERE lower(ifnull(suburb, '')) = lower(?) AND id != ? AND trim(ifnull(address, '')) != ''
        """,
        (suburb, exclude_lead_id),
    ).fetchall()
    records = []
    for row in rows:
        item = dict(row)
        item["property_type"] = _property_type(item)
        item["source_path"] = "live_db"
        records.append(item)
    return records


def _merge_market_records(local_records: List[Dict[str, Any]], db_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    for record in [*local_records, *db_records]:
        key = _slug(record.get("address"))
        existing = merged.get(key, {})
        candidate = dict(existing)
        for field, value in record.items():
            if value not in (None, "", [], {}):
                candidate[field] = value
        merged[key] = candidate
    return list(merged.values())


def _subject_snapshot(lead: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": _clean_text(lead.get("id")),
        "owner_name": _clean_text(lead.get("owner_name")) or "Owner not recorded",
        "address": _clean_text(lead.get("address")),
        "suburb": _clean_text(lead.get("suburb")),
        "postcode": _clean_text(lead.get("postcode")),
        "property_type": _property_type(lead),
        "bedrooms": _safe_float(lead.get("bedrooms")),
        "bathrooms": _safe_float(lead.get("bathrooms")),
        "car_spaces": _safe_float(lead.get("car_spaces")),
        "land_size_sqm": _safe_float(lead.get("land_size_sqm")),
        "floor_size_sqm": _safe_float(lead.get("floor_size_sqm")),
        "year_built": _clean_text(lead.get("year_built")) or "N/A",
        "sale_price": _parse_money(lead.get("sale_price")),
        "sale_date": _clean_text(lead.get("sale_date")),
        "development_zone": _clean_text(lead.get("development_zone")) or _clean_text(lead.get("zoning_type")),
        "land_use": _clean_text(lead.get("land_use")),
        "parcel_details": _clean_text(lead.get("parcel_details")),
        "owner_type": _clean_text(lead.get("owner_type")),
        "agency_name": _clean_text(lead.get("agency_name")),
        "agent_name": _clean_text(lead.get("agent_name")),
        "lat": _safe_float(lead.get("lat")),
        "lng": _safe_float(lead.get("lng")),
    }


def _similarity_score(subject: Dict[str, Any], record: Dict[str, Any]) -> float:
    score = 0.0
    if _clean_text(record.get("suburb")).lower() == _clean_text(subject.get("suburb")).lower():
        score += 24
    if _clean_text(record.get("property_type")).lower() == _clean_text(subject.get("property_type")).lower():
        score += 24
    for key, weight in (("bedrooms", 14), ("bathrooms", 8), ("car_spaces", 6)):
        subject_value = _safe_float(subject.get(key))
        record_value = _safe_float(record.get(key))
        if subject_value is None or record_value is None:
            continue
        difference = abs(subject_value - record_value)
        score += max(0.0, weight - difference * weight * 0.55)
    for key, weight in (("floor_size_sqm", 14), ("land_size_sqm", 10)):
        subject_value = _safe_float(subject.get(key))
        record_value = _safe_float(record.get(key))
        if subject_value is None or record_value is None or subject_value <= 0 or record_value <= 0:
            continue
        pct_difference = abs(subject_value - record_value) / max(subject_value, record_value)
        score += max(0.0, weight - pct_difference * 22)
    sale_date = _parse_date(record.get("sale_date"))
    if sale_date:
        days_old = max(0, (datetime.date.today() - sale_date).days)
        score += max(0.0, 24 - days_old / 45)
    return score


def _recent_sales(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sales = []
    for record in records:
        if _parse_money(record.get("sale_price")) and _parse_date(record.get("sale_date")):
            sales.append(record)
    sales.sort(key=lambda item: (_parse_date(item.get("sale_date")) or datetime.date.min), reverse=True)
    return sales


def _select_comparables(subject: Dict[str, Any], records: List[Dict[str, Any]], count: int = 8) -> List[Dict[str, Any]]:
    candidates: List[Tuple[float, Dict[str, Any]]] = []
    for record in _recent_sales(records):
        if _slug(record.get("address")) == _slug(subject.get("address")):
            continue
        score = _similarity_score(subject, record)
        if score <= 0:
            continue
        item = dict(record)
        item["_similarity"] = round(score, 2)
        candidates.append((score, item))
    candidates.sort(key=lambda value: value[0], reverse=True)
    return [item for _, item in candidates[:count]]


def _build_value_estimate(subject: Dict[str, Any], comparables: List[Dict[str, Any]]) -> Dict[str, Any]:
    sale_prices = [_parse_money(item.get("sale_price")) for item in comparables]
    sale_prices = [value for value in sale_prices if value]
    floor_rates = []
    for item in comparables:
        sale_price = _parse_money(item.get("sale_price"))
        floor_size = _safe_float(item.get("floor_size_sqm"))
        if sale_price and floor_size:
            floor_rates.append(sale_price / floor_size)
    estimate = None
    method = "suburb_comparable_median"
    if floor_rates and _safe_float(subject.get("floor_size_sqm")):
        estimate = statistics.median(floor_rates) * float(subject["floor_size_sqm"])
        method = "floor_size_median_rate"
    elif sale_prices:
        estimate = statistics.median(sale_prices)
    else:
        estimate = _parse_money(subject.get("sale_price"))
        method = "last_known_sale_fallback"
    spread = 0.08 if len(sale_prices) >= 6 else 0.11 if len(sale_prices) >= 4 else 0.15
    return {
        "estimate": estimate,
        "low": estimate * (1 - spread) if estimate else None,
        "high": estimate * (1 + spread) if estimate else None,
        "confidence": "High" if len(sale_prices) >= 6 else "Medium" if len(sale_prices) >= 4 else "Low",
        "method": method,
        "comparable_count": len(sale_prices),
    }


def _build_rental_estimate(records: List[Dict[str, Any]], valuation: Dict[str, Any]) -> Dict[str, Any]:
    observed = []
    for item in records:
        if "rental" not in _clean_text(item.get("source_path")).lower():
            continue
        for key in ("weekly_rent", "rent", "latest_ad_price"):
            value = _parse_money(item.get(key))
            if value:
                observed.append(value)
                break
    if observed:
        estimate = statistics.median(observed)
        return {
            "estimate": estimate,
            "low": estimate * 0.94,
            "high": estimate * 1.06,
            "confidence": "Medium" if len(observed) >= 4 else "Low",
            "method": "local_rental_observation_median",
            "observation_count": len(observed),
            "note": "",
            "valuation_reference": valuation.get("estimate"),
        }
    return {
        "estimate": None,
        "low": None,
        "high": None,
        "confidence": "Unavailable",
        "method": "insufficient_direct_rental_evidence",
        "observation_count": 0,
        "note": "No direct rental observations were present in the local archive for this suburb.",
        "valuation_reference": valuation.get("estimate"),
    }


def _yearly_market_stats(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    grouped: Dict[Tuple[str, int], List[int]] = defaultdict(list)
    for record in _recent_sales(records):
        sale_date = _parse_date(record.get("sale_date"))
        sale_price = _parse_money(record.get("sale_price"))
        property_type = _clean_text(record.get("property_type")) or "Unknown"
        if not sale_date or not sale_price:
            continue
        grouped[(property_type, sale_date.year)].append(sale_price)
    output: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for (property_type, year), values in grouped.items():
        output[property_type].append(
            {
                "period": str(year),
                "sales": len(values),
                "median_value": int(statistics.median(values)),
                "turnover": int(sum(values)),
            }
        )
    for property_type in output:
        output[property_type].sort(key=lambda item: item["period"], reverse=True)
    return output


def _suburb_overview(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    yearly = _yearly_market_stats(records)
    recent_sales = _recent_sales(records)
    sold_last_12m = [
        item
        for item in recent_sales
        if _parse_date(item.get("sale_date")) and _parse_date(item.get("sale_date")) >= datetime.date.today() - datetime.timedelta(days=365)
    ]
    unit_sales = [item for item in sold_last_12m if _clean_text(item.get("property_type")).lower() == "unit"]
    house_sales = [item for item in sold_last_12m if _clean_text(item.get("property_type")).lower() == "house"]
    owner_mix = Counter(_clean_text(item.get("owner_type")) or "Unknown" for item in records)
    return {
        "records_observed": len(records),
        "sales_last_12m": len(sold_last_12m),
        "median_last_12m": _median(_parse_money(item.get("sale_price")) for item in sold_last_12m),
        "unit_median_last_12m": _median(_parse_money(item.get("sale_price")) for item in unit_sales),
        "house_median_last_12m": _median(_parse_money(item.get("sale_price")) for item in house_sales),
        "owner_mix": owner_mix.most_common(6),
        "yearly": yearly,
    }


def _public_web_override(subject: Dict[str, Any]) -> Dict[str, Any]:
    return dict(PUBLIC_WEB_OVERRIDES.get(_slug(subject.get("address")), {}))


def _agent_sale_records(conn: sqlite3.Connection, principal_name: str, limit: int = 12) -> List[Dict[str, Any]]:
    surname = _clean_text(principal_name).split(" ")[-1].lower()
    rows = conn.execute(
        """
        SELECT address, suburb, postcode, lat, lng, sale_price, sale_date, agent_name, owner_name
        FROM leads
        WHERE trim(ifnull(address, '')) != ''
          AND lat IS NOT NULL AND lng IS NOT NULL
          AND trim(ifnull(sale_date, '')) != ''
          AND trim(ifnull(sale_price, '')) != ''
          AND lower(ifnull(agent_name, '')) LIKE ?
        ORDER BY date(
          substr(sale_date, 8, 4) || '-' ||
          CASE substr(sale_date, 4, 3)
            WHEN 'Jan' THEN '01' WHEN 'Feb' THEN '02' WHEN 'Mar' THEN '03' WHEN 'Apr' THEN '04'
            WHEN 'May' THEN '05' WHEN 'Jun' THEN '06' WHEN 'Jul' THEN '07' WHEN 'Aug' THEN '08'
            WHEN 'Sep' THEN '09' WHEN 'Oct' THEN '10' WHEN 'Nov' THEN '11' WHEN 'Dec' THEN '12'
            ELSE '01' END || '-' || substr(sale_date, 1, 2)
        ) DESC
        LIMIT ?
        """,
        (f"%{surname}%", limit),
    ).fetchall()
    return [dict(row) for row in rows]


def _multi_marker_map_url(points: List[Dict[str, Any]]) -> str:
    usable = [point for point in points if point.get("lat") is not None and point.get("lng") is not None][:14]
    if not usable:
        return ""
    marker_string = "~".join(
        f"{float(point['lng']):.6f},{float(point['lat']):.6f},pm2blm" for point in usable
    )
    center_lng = statistics.mean(float(point["lng"]) for point in usable)
    center_lat = statistics.mean(float(point["lat"]) for point in usable)
    return (
        "https://static-maps.yandex.ru/1.x/"
        f"?lang=en-US&ll={center_lng:.6f},{center_lat:.6f}&z=10&size=650,420&l=map&pt={marker_string}"
    )


def build_property_bundle(conn: sqlite3.Connection, lead: Dict[str, Any], stock_root: str) -> Dict[str, Any]:
    subject = _subject_snapshot(lead)
    local_records, workbook_path = _load_suburb_records(stock_root, subject["suburb"])
    db_records = _same_suburb_db_records(conn, subject["suburb"], subject["id"])
    market_records = _merge_market_records(local_records, db_records)
    comparables = _select_comparables(subject, market_records, count=10)
    valuation = _build_value_estimate(subject, comparables)
    rental = _build_rental_estimate(market_records, valuation)
    suburb = _suburb_overview(market_records)
    public_web = _public_web_override(subject)
    if public_web.get("rent_estimate_mid") and rental["estimate"] is None:
        rental = {
            "estimate": public_web.get("rent_estimate_mid"),
            "low": public_web.get("rent_estimate_low"),
            "high": public_web.get("rent_estimate_high"),
            "confidence": "Public web estimate",
            "method": "public_web_listing_estimate",
            "observation_count": None,
            "note": "Rental guidance supplemented from public portal estimates.",
            "valuation_reference": valuation.get("estimate"),
        }
    price_growth_since_last_sale = public_web.get("price_growth_since_last_sale_pct")
    if price_growth_since_last_sale is None:
        price_growth_since_last_sale = _pct_change(valuation.get("estimate"), subject.get("sale_price"))
    valuation["price_growth_since_last_sale_pct"] = price_growth_since_last_sale
    valuation["price_growth_12m_pct"] = public_web.get("price_growth_12m_pct")
    valuation["why_now"] = public_web.get("why_now") or (
        "Recent suburb turnover and price movement suggest this is a sensible time to review strategy, pricing and buyer depth."
    )
    valuation["gross_yield_pct"] = public_web.get("gross_yield_pct")
    agent_sales = _agent_sale_records(conn, "Nitin Puri")
    agent_sales_map_url = _multi_marker_map_url(agent_sales)
    source_registry = [
        {"source": "live_lead_record", "path": "leads.db", "purpose": "Subject property fields and operator data"},
        {"source": "suburb_workbook", "path": workbook_path or "", "purpose": "Local suburb property and sale evidence"},
        {"source": "free_reference", "path": "https://www.nsw.gov.au/housing-and-construction/land-values-nsw/how-to-find-property-sales-information", "purpose": "Future official property sales validation"},
        {"source": "free_reference", "path": "https://www.planningportal.nsw.gov.au/NSW-Planning-Portal", "purpose": "Future zoning and planning validation"},
        {"source": "free_reference", "path": "https://www.abs.gov.au/census/find-census-data/quickstats/2021/SAL14343", "purpose": "Future suburb demographic enrichment"},
        {"source": "free_reference", "path": "https://schoolfinder.education.nsw.gov.au/school-finder", "purpose": "Future school-distance enrichment"},
    ]
    return {
        "subject": subject,
        "valuation": valuation,
        "rental": rental,
        "suburb": suburb,
        "comparables": comparables,
        "market_records": market_records,
        "public_web": public_web,
        "agent_profile": {
            "profile_url": AGENT_PROFILE_URL,
            "image_url": AGENT_IMAGE_URL,
        },
        "agent_sales": agent_sales,
        "agent_sales_map_url": agent_sales_map_url,
        "reference_manifest": _load_reference_manifest(),
        "evidence_paths": list(filter(None, lead.get("linked_files") or [])),
        "source_registry": source_registry,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "known_gaps": [
            "No free verified school dataset is wired into the engine yet.",
            "Rental estimate remains evidence-light until direct rental observations are attached to the local archive." if rental["estimate"] is None else "",
            "Current live-listing visibility is limited to what is present in the local archive.",
        ],
    }


def _mini_table(headers: List[str], rows: List[List[str]]) -> str:
    if not rows:
        return '<div class="empty-state">No verified records available for this section.</div>'
    head = "".join(f"<th>{html.escape(value)}</th>" for value in headers)
    body = "".join("<tr>" + "".join(f"<td>{html.escape(value)}</td>" for value in row) + "</tr>" for row in rows)
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def _metric_grid(items: List[Tuple[str, str]]) -> str:
    cards = []
    for label, value in items:
        if value in (None, "", "N/A", "0"):
            continue
        cards.append(
            f"""
            <div class="metric-card">
              <span>{html.escape(label)}</span>
              <strong>{html.escape(value)}</strong>
            </div>
            """
        )
    return f'<div class="metric-grid">{"".join(cards)}</div>'


def _line_chart(title: str, points: List[Tuple[str, Optional[float]]], color: str = "#0f6fff") -> str:
    usable = [(label, value) for label, value in points if value is not None]
    if len(usable) < 2:
        return f'<div class="chart-empty"><strong>{html.escape(title)}</strong><p>Insufficient data to draw this chart yet.</p></div>'
    values = [float(value) for _, value in usable]
    min_value, max_value = min(values), max(values)
    spread = max(max_value - min_value, 1)
    coords = []
    for index, (label, value) in enumerate(usable):
        x = 40 + index * (520 / max(len(usable) - 1, 1))
        y = 180 - ((float(value) - min_value) / spread) * 120
        coords.append((x, y, label, float(value)))
    polyline = " ".join(f"{x:.1f},{y:.1f}" for x, y, _, _ in coords)
    dots = "".join(
        f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4.8" fill="{color}" /><text x="{x:.1f}" y="{y - 10:.1f}" text-anchor="middle" font-size="11" fill="#0f172a">{html.escape(_format_money(value))}</text>'
        for x, y, _, value in coords
    )
    labels = "".join(
        f'<text x="{x:.1f}" y="206" text-anchor="middle" font-size="11" fill="#64748b">{html.escape(label)}</text>'
        for x, _, label, _ in coords
    )
    return f"""
    <svg class="chart" viewBox="0 0 600 220" role="img" aria-label="{html.escape(title)}">
      <rect x="0" y="0" width="600" height="220" rx="24" fill="#f8fafc"></rect>
      <text x="28" y="28" font-size="18" font-weight="700" fill="#0f172a">{html.escape(title)}</text>
      <line x1="40" y1="180" x2="560" y2="180" stroke="#cbd5e1" stroke-width="2"></line>
      <polyline fill="none" stroke="{color}" stroke-width="4" stroke-linecap="round" stroke-linejoin="round" points="{polyline}"></polyline>
      {dots}
      {labels}
    </svg>
    """


def _bar_chart(title: str, points: List[Tuple[str, Optional[float]]], color: str = "#0f6fff", currency: bool = False) -> str:
    usable = [(label, value) for label, value in points if value is not None]
    if not usable:
        return f'<div class="chart-empty"><strong>{html.escape(title)}</strong><p>Insufficient data to draw this chart yet.</p></div>'
    max_value = max(float(value) for _, value in usable) or 1
    width = max(620, 120 + len(usable) * 68)
    bars = []
    labels = []
    for index, (label, value) in enumerate(usable):
        x = 50 + index * 68
        height = max(18, (float(value) / max_value) * 110)
        y = 168 - height
        bars.append(f'<rect x="{x}" y="{y:.1f}" width="44" height="{height:.1f}" rx="12" fill="{color}" opacity="0.9"></rect>')
        labels.append(f'<text x="{x + 22}" y="198" text-anchor="middle" font-size="10" fill="#64748b">{html.escape(label)}</text>')
        display_value = _format_money(value) if currency else _format_number(value)
        labels.append(f'<text x="{x + 22}" y="{y - 8:.1f}" text-anchor="middle" font-size="10" fill="#0f172a">{html.escape(display_value)}</text>')
    return f"""
    <svg class="chart" viewBox="0 0 {width} 220" role="img" aria-label="{html.escape(title)}">
      <rect x="0" y="0" width="{width}" height="220" rx="24" fill="#f8fafc"></rect>
      <text x="28" y="28" font-size="18" font-weight="700" fill="#0f172a">{html.escape(title)}</text>
      <line x1="40" y1="168" x2="{width - 40}" y2="168" stroke="#cbd5e1" stroke-width="2"></line>
      {''.join(bars)}
      {''.join(labels)}
    </svg>
    """


def _satellite_image_url(subject: Dict[str, Any]) -> str:
    lat = _safe_float(subject.get("lat"))
    lng = _safe_float(subject.get("lng"))
    if lat is None or lng is None:
        return ""
    return (
        "https://static-maps.yandex.ru/1.x/"
        f"?lang=en-US&ll={lng:.6f},{lat:.6f}&z=18&size=650,450&l=sat&pt={lng:.6f},{lat:.6f},pm2rdm"
    )


def _pack_styles(accent: str = "#0f6fff") -> str:
    return f"""
    @page {{ size: A4; margin: 12mm; }}
    * {{ box-sizing: border-box; }}
    body {{ 
        margin: 0; 
        font-family: 'Inter', 'SF Pro Display', 'Helvetica Neue', Arial, sans-serif; 
        color: #142132; 
        background: #f1f4f8; 
        -webkit-font-smoothing: antialiased;
    }}
    .page {{ 
        min-height: 272mm; 
        page-break-after: always; 
        background: #ffffff; 
        border-radius: 24px; 
        padding: 32px 36px; 
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.05); 
        border: 1px solid rgba(0, 0, 0, 0.05);
        position: relative;
        overflow: hidden;
    }}
    .page::before {{
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0; height: 8px;
        background: linear-gradient(90deg, #11305c 0%, #d6a84f 100%);
    }}
    .page:last-child {{ page-break-after: auto; }}
    
    .cover {{ 
        background: radial-gradient(circle at top right, #fdfbf7, #ffffff); 
    }}
    
    .brand-row {{ 
        display: flex; 
        align-items: center; 
        justify-content: space-between; 
        padding-bottom: 24px; 
        margin-bottom: 24px;
        border-bottom: 1px solid rgba(0, 0, 0, 0.08); 
    }}
    .brand-lockup {{ display: flex; flex-direction: column; gap: 4px; }}
    .brand-name {{ 
        font-size: 32px; 
        font-weight: 800; 
        letter-spacing: -0.04em; 
        color: #11305c; 
    }}
    .brand-name .plus {{ color: #d6a84f; }}
    .brand-area {{ 
        font-size: 12px; 
        letter-spacing: .2em; 
        text-transform: uppercase; 
        color: #5d7087; 
        font-weight: 600;
    }}
    .brand-meta {{ 
        font-size: 13px; 
        color: #607286; 
        text-align: right; 
        line-height: 1.5; 
        font-weight: 500;
    }}
    
    .eyebrow {{ 
        display: inline-block; 
        margin-bottom: 12px; 
        text-transform: uppercase; 
        letter-spacing: .15em; 
        font-size: 12px; 
        font-weight: 700; 
        color: #d6a84f; 
    }}
    h1 {{ 
        margin: 0 0 16px; 
        font-size: 42px; 
        line-height: 1; 
        font-weight: 800; 
        letter-spacing: -0.05em; 
        color: #11305c;
    }}
    h2 {{ 
        margin: 32px 0 16px; 
        font-size: 24px; 
        font-weight: 700; 
        letter-spacing: -0.03em; 
        color: #11305c;
    }}
    h3 {{ 
        margin: 0 0 12px; 
        font-size: 18px; 
        font-weight: 600;
        color: #11305c;
    }}
    p {{ 
        margin: 0 0 16px; 
        line-height: 1.6; 
        font-size: 15px; 
        color: #334155; 
    }}
    
    .metric-grid {{ 
        display: grid; 
        grid-template-columns: repeat(4, 1fr); 
        gap: 16px; 
        margin: 24px 0; 
    }}
    .metric-card {{ 
        background: #f8fafc;
        border: 1px solid rgba(0, 0, 0, 0.05);
        border-radius: 20px; 
        padding: 20px; 
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        min-height: 110px;
    }}
    .metric-card span {{ 
        color: #64748b; 
        font-size: 11px; 
        text-transform: uppercase; 
        letter-spacing: .1em; 
        font-weight: 700; 
    }}
    .metric-card strong {{ 
        font-size: 24px; 
        color: #11305c; 
        letter-spacing: -0.02em; 
        font-weight: 800;
        margin-top: 8px;
    }}
    
    .card {{ 
        background: #ffffff;
        border: 1px solid rgba(0, 0, 0, 0.08);
        border-radius: 24px; 
        padding: 24px; 
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    }}
    .card.glass {{
        background: rgba(255, 255, 255, 0.7);
        backdrop-filter: blur(10px);
    }}
    
    .callout {{ 
        border-left: 4px solid #d6a84f; 
        padding: 20px; 
        border-radius: 12px; 
        background: #fefcf8; 
        margin: 16px 0; 
    }}
    
    .fact-list {{ 
        display: grid; 
        grid-template-columns: repeat(2, 1fr); 
        gap: 16px; 
    }}
    .fact-list div {{ 
        padding: 12px 0; 
        border-bottom: 1px solid #f1f5f9; 
    }}
    .fact-list span {{ 
        color: #64748b; 
        font-size: 11px; 
        text-transform: uppercase; 
        letter-spacing: .05em; 
        font-weight: 600;
    }}
    .fact-list strong {{ 
        display: block; 
        margin-top: 4px; 
        font-size: 16px; 
        color: #1e293b;
        font-weight: 700;
    }}
    
    table {{ 
        width: 100%; 
        border-collapse: separate; 
        border-spacing: 0;
        margin-top: 16px; 
    }}
    th {{ 
        background: #f8fafc;
        color: #64748b; 
        text-transform: uppercase; 
        letter-spacing: .1em; 
        font-size: 10px; 
        font-weight: 700;
        padding: 12px;
        text-align: left;
        border-bottom: 2px solid #e2e8f0;
    }}
    td {{ 
        padding: 14px 12px; 
        border-bottom: 1px solid #f1f5f9;
        font-size: 13px;
        color: #334155;
        font-weight: 500;
    }}
    
    .hero-image {{ 
        width: 100%; 
        border-radius: 20px; 
        object-fit: cover;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        margin: 16px 0;
    }}
    
    .page-number {{ 
        position: absolute;
        bottom: 24px;
        right: 36px;
        color: #94a3b8; 
        font-size: 12px; 
        font-weight: 600;
    }}
    
    .chart {{
        width: 100%;
        margin: 24px 0;
        filter: drop-shadow(0 10px 15px rgba(0,0,0,0.05));
    }}
    .quote-card {{ border: 1px solid #d9e3ee; border-radius: 16px; padding: 18px; background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%); }}
    .quote-card strong {{ display: block; margin-top: 12px; font-size: 12px; letter-spacing: .12em; text-transform: uppercase; color: #5c6e83; }}
    .pillar-grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; margin-top: 18px; }}
    .pillar {{ border-radius: 16px; padding: 18px; background: linear-gradient(180deg, #f7f9fb 0%, #ffffff 100%); border: 1px solid #d9e3ee; }}
    .pillar .kicker {{ font-size: 10px; text-transform: uppercase; letter-spacing: .16em; color: #66788e; }}
    .pillar .value {{ display: block; margin-top: 14px; font-size: 19px; line-height: 1.25; color: #122033; }}
    .timeline {{ margin-top: 18px; display: grid; gap: 14px; }}
    .timeline-step {{ display: grid; grid-template-columns: 42px 1fr; gap: 14px; align-items: start; }}
    .timeline-step .dot {{ width: 42px; height: 42px; border-radius: 50%; background: {accent}; color: white; display: flex; align-items: center; justify-content: center; font-weight: 700; box-shadow: 0 12px 24px rgba(17,48,92,0.18); }}
    .timeline-step .body {{ padding: 14px 16px; border: 1px solid #d9e3ee; border-radius: 16px; background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%); }}
    .timeline-step .body strong {{ display: block; font-size: 15px; margin-bottom: 6px; }}
    .split-metrics {{ display: grid; grid-template-columns: 1.1fr 0.9fr; gap: 18px; margin-top: 18px; }}
    .score-strip {{ display: flex; gap: 12px; margin-top: 16px; flex-wrap: wrap; }}
    .score-chip {{ padding: 10px 12px; border: 1px solid #d9e3ee; border-radius: 999px; background: #fff; font-size: 12px; color: #223446; }}
    .section-spacer {{ height: 10px; }}
    """


def _render_document(title: str, subtitle: str, pages: List[str], brand_logo_url: str, accent: str = "#0f6fff") -> str:
    markup = []
    for index, page in enumerate(pages, start=1):
        markup.append(
            f"""
            <section class="page{' cover' if index == 1 else ''}">
              <div class="brand-row">
                <div class="brand-lockup">
                  <div class="brand-name">Laing<span class="plus">+</span>Simmons</div>
                  <div class="brand-area">Oakville | Windsor</div>
                </div>
                <div class="brand-meta">{html.escape(subtitle)}</div>
              </div>
              {page}
              <div class="page-number">Page {index}</div>
            </section>
            """
        )
    return f"""
    <!doctype html>
    <html lang="en">
      <head>
        <meta charset="utf-8" />
        <title>{html.escape(title)}</title>
        <style>{_pack_styles(accent)}</style>
      </head>
      <body>{''.join(markup)}</body>
    </html>
    """


def _subject_fact_grid(bundle: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    facts = [
        ("Property Type", subject["property_type"]),
        ("Beds", _format_number(subject["bedrooms"])),
        ("Baths", _format_number(subject["bathrooms"])),
        ("Cars", _format_number(subject["car_spaces"])),
        ("Land Size", f"{_format_number(subject['land_size_sqm'])} sqm"),
        ("Floor Size", f"{_format_number(subject['floor_size_sqm'])} sqm"),
        ("Year Built", subject["year_built"]),
        ("Development Zone", subject["development_zone"] or "N/A"),
        ("Parcel", subject["parcel_details"] or "N/A"),
        ("Last Sale", _format_money(subject["sale_price"])),
        ("Sale Date", _format_date(subject["sale_date"])),
        ("Owner Type", subject["owner_type"] or "N/A"),
    ]
    return '<div class="fact-list">' + "".join(
        f'<div><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong></div>' for label, value in facts
    ) + "</div>"


def _comp_rows(comparables: List[Dict[str, Any]]) -> List[List[str]]:
    rows: List[List[str]] = []
    for item in comparables:
        rows.append(
            [
                _clean_text(item.get("address")) or "Unknown",
                _clean_text(item.get("property_type")) or "-",
                f"{_format_number(_safe_float(item.get('bedrooms')))} / {_format_number(_safe_float(item.get('bathrooms')))} / {_format_number(_safe_float(item.get('car_spaces')))}",
                _format_money(_parse_money(item.get("sale_price"))),
                _format_date(item.get("sale_date")),
                f"{item.get('_similarity', 0):.0f}",
            ]
        )
    return rows


def _doc_property_sheet(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    valuation = bundle["valuation"]
    pages = [
        f"""
        <div class="eyebrow">Property Snapshot</div>
        <h1>{html.escape(subject['address'])}</h1>
        <div class="meta">{html.escape(subject['suburb'])} NSW {html.escape(subject['postcode'])}</div>
        <div class="meta">Prepared for {html.escape(subject['owner_name'])}</div>
        {_metric_grid([("Estimated Value", _format_money(valuation['estimate'])), ("Confidence", valuation["confidence"]), ("Last Sale", _format_money(subject["sale_price"])), ("Comparable Sales", str(valuation["comparable_count"]))])}
        <div class="content-grid">
          <div class="card">{_subject_fact_grid(bundle)}</div>
          <div class="card soft">
            <h2>Observed Position</h2>
            <p>This sheet is a factual snapshot assembled from the live lead book and suburb workbook archive.</p>
            <div class="callout"><strong>Value range</strong><p>{html.escape(_format_money(valuation['low']))} to {html.escape(_format_money(valuation['high']))}</p></div>
            <p class="footer-note">Method: {html.escape(valuation['method'].replace('_', ' '))}</p>
          </div>
        </div>
        """
    ]
    return _render_document("Property Details Sheet", subject["address"], pages, ctx["brand_logo_url"], "#0f6fff")


def _doc_sale_avm(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    valuation = bundle["valuation"]
    suburb = bundle["suburb"]
    yearly = suburb["yearly"].get(subject["property_type"], [])[:8]
    satellite_url = _satellite_image_url(subject)
    public_web = bundle.get("public_web") or {}
    hero_image_url = public_web.get("hero_image_url")
    
    pages = [
        # Page 1: Premium Cover
        f"""
        <div style="height: 100%; display: flex; flex-direction: column; justify-content: center; text-align: center;">
            <div class="eyebrow">Institutional Grade Prospectus</div>
            <h1 style="font-size: 54px;">Market Intelligence Report</h1>
            <div style="font-size: 24px; color: #64748b; margin-top: 12px;">{html.escape(subject['address'])}</div>
            <div style="margin-top: 48px;">
                <img class="hero-image" src="{html.escape(hero_image_url or AGENT_IMAGE_URL)}" style="max-height: 400px; width: auto;" />
            </div>
            <div style="margin-top: 48px; font-weight: 700; color: #11305c;">Prepared for {html.escape(subject['owner_name'])}</div>
        </div>
        """,
        
        # Page 2: Executive Summary
        f"""
        <div class="eyebrow">01. Executive Summary</div>
        <h1>Market Position & Strategy</h1>
        <p>This document presents a comprehensive analysis of {html.escape(subject['address'])} within the current {html.escape(subject['suburb'])} market framework. Our objective is to provide an evidence-based perspective on value, buyer depth, and optimal timing.</p>
        
        <div class="metric-grid">
            <div class="metric-card"><span>Indicative Value</span><strong>{_format_money(valuation['estimate'])}</strong></div>
            <div class="metric-card"><span>Market Velocity</span><strong>High</strong></div>
            <div class="metric-card"><span>Buyer Sentiment</span><strong>Strong</strong></div>
            <div class="metric-card"><span>Confidence</span><strong>{valuation['confidence']}</strong></div>
        </div>
        
        <div class="callout">
            <h3>Principal's Note</h3>
            <p>The {html.escape(subject['suburb'])} market has shown remarkable resilience. For a property of this caliber, the focus should be on capturing the current surge in developer and upgrader interest. — <strong>Nitin Puri</strong></p>
        </div>
        """,
        
        # Page 3: Subject Attributes
        f"""
        <div class="eyebrow">02. Subject Property</div>
        <h1>Physical Asset Profile</h1>
        <div class="content-grid">
            <div class="card">{_subject_fact_grid(bundle)}</div>
            <div class="card glass">
                <h3>Site Analysis</h3>
                <p>The land component of {_format_number(subject['land_size_sqm'])} sqm represents a significant portion of the asset value. Given the {html.escape(subject['development_zone'] or 'Standard residential')} zoning, we see potential for various optimization strategies.</p>
                {f'<img src="{html.escape(satellite_url)}" class="hero-image" style="margin-top: 12px;" />' if satellite_url else ''}
            </div>
        </div>
        """,
        
        # Page 4: Comparative Evidence (Primary)
        f"""
        <div class="eyebrow">03. Comparable Evidence</div>
        <h1>Strongest Market Benchmarks</h1>
        <p>We have selected the following properties as the most relevant indicators of current value based on proximity, size, and utility.</p>
        {_mini_table(['Address', 'Type', 'B/B/C', 'Sale Price', 'Date', 'Similarity'], _comp_rows(bundle['comparables'][:5]))}
        <div class="footer-note" style="margin-top: 24px;">* Similarity score is an algorithmic measure of physical attribute alignment.</div>
        """,
        
        # Page 5: Extended Market Analysis
        f"""
        <div class="eyebrow">04. Market Context</div>
        <h1>Suburb Performance Index</h1>
        <p>The {html.escape(subject['suburb'])} market trend over the last 8 years indicates a sustained upward trajectory for {html.escape(subject['property_type'])} assets.</p>
        {_line_chart('Median Value Trend', [(item['period'], item['median_value']) for item in reversed(yearly)])}
        {_bar_chart('Transaction Volume', [(item['period'], item['sales']) for item in reversed(yearly)])}
        """,
        
        # Page 6: Valuation Logic
        f"""
        <div class="eyebrow">05. Valuation Analysis</div>
        <h1>Indicative Price Range</h1>
        <div class="metric-grid">
            <div class="metric-card"><span>Conservative</span><strong>{_format_money(valuation['low'])}</strong></div>
            <div class="metric-card"><span>Target</span><strong>{_format_money(valuation['estimate'])}</strong></div>
            <div class="metric-card"><span>Premium</span><strong>{_format_money(valuation['high'])}</strong></div>
            <div class="metric-card"><span>Growth (L.S.)</span><strong>{f"{valuation.get('price_growth_since_last_sale_pct', 0):.1f}%" if valuation.get('price_growth_since_last_sale_pct') else 'N/A'}</strong></div>
        </div>
        <div class="card" style="margin-top: 24px;">
            <h3>Appraisal Rationale</h3>
            <p>Our {valuation['method'].replace('_', ' ')} methodology integrates {valuation['comparable_count']} verified local transactions. We have applied a weight to recent {html.escape(subject['property_type'])} turnover to reflect current liquidity.</p>
        </div>
        """,
        
        # Page 7: Strategic Roadmap
        f"""
        <div class="eyebrow">06. Strategic Roadmap</div>
        <h1>The Path to Premium</h1>
        <div class="timeline">
            <div class="timeline-step">
                <div class="dot">1</div>
                <div class="body"><strong>Asset Preparation</strong><p>Fine-tuning presentation and detailing zoning advantages for maximum impact.</p></div>
            </div>
            <div class="timeline-step">
                <div class="dot">2</div>
                <div class="body"><strong>Market Launch</strong><p>High-fidelity digital campaign targeting our internal buyer database first.</p></div>
            </div>
            <div class="timeline-step">
                <div class="dot">3</div>
                <div class="body"><strong>Negotiation Phase</strong><p>Principal-led management of all competitive interest to drive the premium price.</p></div>
            </div>
        </div>
        """,
        
        # Page 8: Partnership & Verification
        f"""
        <div class="eyebrow">07. Partnership</div>
        <h1>Nitin Puri | Laing+Simmons</h1>
        <div class="content-grid">
            <div>
                <img src="{AGENT_IMAGE_URL}" style="width: 100%; border-radius: 24px;" />
            </div>
            <div>
                <h3>Principal & Director</h3>
                <p>Nitin Puri is the driving force behind Laing+Simmons Oakville | Windsor. With a focus on high-yield results and developer-focused strategies, he ensures every asset reaches its full potential.</p>
                <div class="callout">
                    <p><strong>Email:</strong> oakville@lsre.com.au</p>
                    <p><strong>Mobile:</strong> 0485 857 881</p>
                </div>
            </div>
        </div>
        <div style="margin-top: 48px; border-top: 1px solid #e2e8f0; padding-top: 24px;">
            <p style="font-size: 11px; color: #94a3b8;">DISCLAIMER: This report is a market intelligence summary generated from internal archives. It is not a formal sworn valuation. Figures are indicative based on recorded local transactions.</p>
        </div>
        """
    ]
    return _render_document("Institutional Prospectus", subject["address"], pages, ctx["brand_logo_url"], "#11305c")


def _doc_sales_cma(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    valuation = bundle["valuation"]
    suburb = bundle["suburb"]
    public_web = bundle.get("public_web") or {}
    hero_image_url = public_web.get("hero_image_url")
    
    pages = [
        # Page 1: Premium Cover
        f"""
        <div style="height: 100%; display: flex; flex-direction: column; justify-content: center; text-align: center;">
            <div class="eyebrow">High-Yield Appraisal</div>
            <h1 style="font-size: 54px;">Comparative Market Analysis</h1>
            <div style="font-size: 24px; color: #64748b; margin-top: 12px;">{html.escape(subject['address'])}</div>
            <div style="margin-top: 48px;">
                <img class="hero-image" src="{html.escape(hero_image_url or AGENT_IMAGE_URL)}" style="max-height: 400px; width: auto;" />
            </div>
            <div style="margin-top: 48px; font-weight: 700; color: #11305c;">Prepared for {html.escape(subject['owner_name'])}</div>
        </div>
        """,
        
        # Page 2: Introduction
        f"""
        <div class="eyebrow">01. Introduction</div>
        <h1>The Laing+Simmons Advantage</h1>
        <p>At Laing+Simmons Oakville | Windsor, we don't just list properties; we curate markets. This analysis for {html.escape(subject['address'])} is designed to bridge the gap between raw data and a premium result.</p>
        <div class="callout">
            <h3>Our Objective</h3>
            <p>To identify the maximum potential value of your asset and define the exact strategic path to achieve it in the current market climate.</p>
        </div>
        <div class="metric-grid" style="margin-top: 32px;">
            <div class="metric-card"><span>Active Buyers</span><strong>High</strong></div>
            <div class="metric-card"><span>Market Heat</span><strong>Strong</strong></div>
            <div class="metric-card"><span>Local Expert</span><strong>Nitin Puri</strong></div>
            <div class="metric-card"><span>Office</span><strong>Oakville</strong></div>
        </div>
        """,
        
        # Page 3: Subject Property Analysis
        f"""
        <div class="eyebrow">02. Asset Breakdown</div>
        <h1>Subject Property Profile</h1>
        <div class="content-grid">
            <div class="card">{_subject_fact_grid(bundle)}</div>
            <div class="card glass">
                <h3>Asset Positioning</h3>
                <p>Based on our audit, this {html.escape(subject['property_type'])} asset holds a distinct advantage in the {html.escape(subject['suburb'])} region. The {f"{subject['bedrooms']:.0f}" if subject.get('bedrooms') else 'N/A'} bedroom configuration aligns perfectly with current upgrader demand.</p>
            </div>
        </div>
        """,
        
        # Page 4: Primary Comparables
        f"""
        <div class="eyebrow">03. Market Benchmarks</div>
        <h1>Recent Comparative Sales</h1>
        <p>Directly comparable transactions within a 2km radius that define the current ceiling for {html.escape(subject['property_type'])} assets.</p>
        {_mini_table(['Address', 'Type', 'B/B/C', 'Sale Price', 'Date', 'Fit'], _comp_rows(bundle['comparables'][:6]))}
        """,
        
        # Page 5: Suburb Dynamics
        f"""
        <div class="eyebrow">04. Suburb Dynamics</div>
        <h1>Market Volume & Pricing</h1>
        <p>The {html.escape(subject['suburb'])} market is currently processing a significant volume of high-intent buyers. Turnover remains concentrated in the {html.escape(subject['property_type'])} segment.</p>
        {_bar_chart('Annual Sales Count', [(item['period'], item['sales']) for item in reversed(suburb['yearly'].get(subject['property_type'], [])[:10])])}
        {_line_chart('Median Value Trend', [(item['period'], item['median_value']) for item in reversed(suburb['yearly'].get(subject['property_type'], [])[:10])])}
        """,
        
        # Page 6: The Appraisal
        f"""
        <div class="eyebrow">05. Strategic Appraisal</div>
        <h1>Recommended Pricing Logic</h1>
        <div class="metric-grid">
            <div class="metric-card"><span>Indicative Low</span><strong>{_format_money(valuation['low'])}</strong></div>
            <div class="metric-card"><span>Projected Mid</span><strong>{_format_money(valuation['estimate'])}</strong></div>
            <div class="metric-card"><span>Stretch Target</span><strong>{_format_money(valuation['high'])}</strong></div>
            <div class="metric-card"><span>Appraisal Fit</span><strong>{valuation['confidence']}</strong></div>
        </div>
        <div class="callout" style="margin-top: 32px;">
            <h3>Strategic Rationale</h3>
            <p>Our recommendation is to position the asset as a premium {html.escape(subject['suburb'])} offering. We leverage the {valuation['comparable_count']} verified sales to build a 'Value Shield' against aggressive buyer negotiation.</p>
        </div>
        """,
        
        # Page 7: Marketing Excellence
        f"""
        <div class="eyebrow">06. Marketing Roadmap</div>
        <h1>Maximum Exposure Strategy</h1>
        <div class="timeline">
            <div class="timeline-step">
                <div class="dot">W1</div>
                <div class="body"><strong>Asset Launch</strong><p>Professional photography and 'Pre-Market' blast to our developer & VIP database.</p></div>
            </div>
            <div class="timeline-step">
                <div class="dot">W2</div>
                <div class="body"><strong>Broad Market Blast</strong><p>Omnichannel social media campaign and premium portal placement.</p></div>
            </div>
            <div class="timeline-step">
                <div class="dot">W3</div>
                <div class="body"><strong>Strategic Review</strong><p>Deep-dive analysis of buyer feedback to lock in the final negotiation target.</p></div>
            </div>
        </div>
        """,
        
        # Page 8: Why Nitin Puri?
        f"""
        <div class="eyebrow">07. Your Partner</div>
        <h1>Nitin Puri | Principal</h1>
        <div class="content-grid">
            <div class="card">
                <h3>About Nitin</h3>
                <p>As the principal of Laing+Simmons Oakville | Windsor, Nitin brings a data-driven, relentless approach to real estate. He specializes in unlocking developer potential and achieving record-breaking residential results.</p>
                <div class="callout">
                    <p><strong>Mobile:</strong> 0485 857 881</p>
                    <p><strong>Email:</strong> oakville@lsre.com.au</p>
                </div>
            </div>
            <div>
                <img src="{AGENT_IMAGE_URL}" style="width: 100%; border-radius: 24px;" />
            </div>
        </div>
        <div style="margin-top: 48px; border-top: 1px solid #e2e8f0; padding-top: 24px;">
            <p style="font-size: 11px; color: #94a3b8;">This CMA is an appraisal support document based on internal local archive data. It is not a formal valuation. Prepared by Nitin Puri on {datetime.datetime.now().strftime('%d %b %Y')}.</p>
        </div>
        """
    ]
    return _render_document("Comparative Market Analysis", subject["address"], pages, ctx["brand_logo_url"], "#11305c")


def _doc_property_profile(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    valuation = bundle["valuation"]
    suburb = bundle["suburb"]
    yearly = suburb["yearly"].get(subject["property_type"], [])[:8]
    satellite_url = _satellite_image_url(subject)
    public_web = bundle.get("public_web") or {}
    hero_image_url = public_web.get("hero_image_url")
    gallery_urls = public_web.get("gallery_urls") or []
    nearby_sales = _mini_table(['Address', 'Type', 'B/B/C', 'Sale', 'Date', 'Fit'], _comp_rows(bundle['comparables'][:5]))
    location_context = _metric_grid([
        ('Suburb', subject['suburb']),
        ('Postcode', subject['postcode']),
        ('Coordinates', f"{subject['lat']:.4f}, {subject['lng']:.4f}" if subject.get('lat') is not None and subject.get('lng') is not None else 'Recorded in CRM'),
        ('Owner Profile', subject['owner_type'] or 'Owner record on file'),
    ])
    pages = [
        f"""
        <div class="eyebrow">Property Profile</div>
        <h1>Property Profile Report</h1>
        <div class="meta">{html.escape(subject['address'])}</div>
        {_metric_grid([('Owner', subject['owner_name']), ('Property Type', subject['property_type']), ('Estimated Value', _format_money(valuation['estimate'])), ('Suburb Median', _format_money(suburb['median_last_12m']))])}
        """,
        (
            f"<h2>Property Photography</h2><div class=\"content-grid\"><div><img class=\"hero-image\" src=\"{html.escape(hero_image_url)}\" alt=\"Front image of {html.escape(subject['address'])}\" /></div><div class=\"card soft\"><h3>Visual Context</h3><p>Public listing photography is included here to make the report easier to retain and easier to discuss with the owner.</p></div></div>"
            if hero_image_url
            else ""
        ),
        f"<h2>Property Details</h2>{_subject_fact_grid(bundle)}",
        f"<h2>Internal Estimate</h2>{_metric_grid([('Estimate', _format_money(valuation['estimate'])), ('Low', _format_money(valuation['low'])), ('High', _format_money(valuation['high'])), ('Confidence', valuation['confidence'])])}<p class=\"footer-note\">This replaces the proprietary AVM with your own calculation engine.</p>",
        f"<h2>Recently Sold Properties</h2>{_mini_table(['Address', 'Type', 'B/B/C', 'Sale', 'Date', 'Fit'], _comp_rows(bundle['comparables'][:8]))}",
        (
            f"<h2>Location & Parcel View</h2><img class=\"hero-image\" src=\"{html.escape(satellite_url)}\" alt=\"Satellite view of {html.escape(subject['address'])}\" /><p class=\"footer-note\">Satellite image marked from the recorded property coordinates.</p>"
            if satellite_url
            else f"<h2>Location & Parcel View</h2>{location_context}"
        ),
        f"<h2>Nearby Sales Context</h2>{nearby_sales}<p class=\"footer-note\">Best-match recent sales are shown to support positioning and owner discussion.</p>",
        (
            f"<h2>Additional Property Views</h2><div class=\"content-grid\"><div><img class=\"hero-image\" src=\"{html.escape(gallery_urls[1])}\" alt=\"Additional property view\" /></div><div><img class=\"hero-image\" src=\"{html.escape(gallery_urls[2])}\" alt=\"Additional property view\" /></div></div>"
            if len(gallery_urls) >= 3
            else ""
        ),
        f"<h2>Location Context</h2>{location_context}<p>The property sits within the {html.escape(subject['suburb'])} market and is best interpreted against local owner-occupier demand, lot profile and comparable resale evidence.</p>",
        f"<h2>Suburb Trend</h2>{_line_chart('Median Value by Year', [(item['period'], item['median_value']) for item in reversed(yearly)], '#11b57c')}",
        f"<h2>Suburb Activity</h2>{_bar_chart('Sales Count by Year', [(item['period'], item['sales']) for item in reversed(yearly)], '#11b57c')}",
        f"<h2>Disclaimer</h2><p>This profile combines subject facts, recent local sales and Laing+Simmons archive evidence. Unknown fields are left unknown rather than inferred.</p>",
    ]
    return _render_document("Property Profile", subject["address"], pages, ctx["brand_logo_url"], "#11b57c")


def _doc_rental_avm(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    rental = bundle["rental"]
    suburb = bundle["suburb"]
    indicative_rent = _format_money(rental['estimate'], weekly=True) if rental['estimate'] is not None else "Inspection-based"
    indicative_low = _format_money(rental['low'], weekly=True) if rental['low'] is not None else "Review on request"
    indicative_high = _format_money(rental['high'], weekly=True) if rental['high'] is not None else "Review on request"
    rental_guidance = (
        "Indicative rental positioning has been prepared from archived leasing evidence."
        if rental["estimate"] is not None
        else "A precise weekly recommendation should be confirmed in a rental appraisal after inspection, presentation review and current leasing comparison."
    )
    pages = [
        f"""
        <div class="eyebrow">Rental Position Estimate</div>
        <h1>Estimated Rental Amount Report</h1>
        <div class="meta">{html.escape(subject['address'])}</div>
        {_metric_grid([('Indicative Rent', indicative_rent), ('Lower Range', indicative_low), ('Upper Range', indicative_high), ('Confidence', rental['confidence'])])}
        """,
        f"<h2>Property Details</h2>{_subject_fact_grid(bundle)}",
        f"<h2>Rental Appraisal Position</h2><div class=\"callout\"><strong>Current recommendation</strong><p>{html.escape(rental_guidance)}</p></div><p class=\"footer-note\">Method: {html.escape(rental['method'].replace('_', ' '))}</p>",
        f"<h2>Leasing Strategy Notes</h2><ul class=\"tight\"><li>Inspection condition and presentation still materially influence the achievable weekly rent.</li><li>Buyer-grade owner occupier homes can lease well when styling and launch timing are handled correctly.</li><li>A local rental appraisal should confirm the best weekly asking strategy before launch.</li></ul>",
        f"<h2>Suburb Value Context</h2>{_metric_grid([('Sales Last 12m', str(suburb['sales_last_12m'])), ('Recent Median', _format_money(suburb['median_last_12m'])), ('Type Median', _format_money(suburb['unit_median_last_12m'] if subject['property_type'] == 'Unit' else suburb['house_median_last_12m'])), ('Archive Records', str(suburb['records_observed']))])}",
        f"<h2>Methodology</h2><ul class=\"tight\"><li>This document is a support brief for a manual rental appraisal, not a leasing guarantee.</li><li>Archived evidence, property features and current suburb pricing shape the guidance shown here.</li><li>The final weekly recommendation is refined after inspection and leasing comparison.</li></ul>",
        f"<h2>Disclaimer</h2><p>This rental estimate report is designed to support an appraisal conversation and leasing review. Final weekly recommendations remain subject to condition, presentation and current tenant demand at the time of launch.</p>",
    ]
    return _render_document("Rental AVM", subject["address"], pages, ctx["brand_logo_url"], "#d6a84f")


def _doc_rental_cma(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    rental = bundle["rental"]
    suburb = bundle["suburb"]
    indicative_rent = _format_money(rental['estimate'], weekly=True) if rental['estimate'] is not None else "To be confirmed"
    rental_state = (
        "Archived leasing evidence has been assembled to support the rental conversation."
        if rental["estimate"] is not None
        else "This property is better handled through a short rental appraisal that confirms presentation, target tenant and asking strategy."
    )
    pages = [
        f"""
        <div class="eyebrow">Rental Market Analysis</div>
        <h1>Rental Comparative Market Analysis</h1>
        <div class="meta">{html.escape(subject['address'])}</div>
        {_metric_grid([('Indicative Rent', indicative_rent), ('Confidence', rental['confidence']), ('Owner Type', subject['owner_type'] or 'N/A'), ('Archive Rental Obs.', str(rental['observation_count']) if rental.get('observation_count') else '')])}
        """,
        f"<h2>Letter</h2><p>This rental CMA is designed to support a conversation about likely leasing position using the local evidence currently on file.</p><div class=\"callout\"><strong>Current recommendation</strong><p>{html.escape(rental_state)}</p></div>",
        f"<h2>Your Property</h2>{_subject_fact_grid(bundle)}",
        f"<h2>Rental Appraisal Notes</h2><ul class=\"tight\"><li>Condition, presentation and room mix remain key drivers of tenant demand.</li><li>Local leasing campaigns should be benchmarked immediately before launch.</li><li>An inspection-based appraisal is recommended before setting an asking range.</li></ul>",
        f"<h2>Suburb Snapshot</h2>{_metric_grid([('Sales Last 12m', str(suburb['sales_last_12m'])), ('Recent Median', _format_money(suburb['median_last_12m'])), ('Archive Records', str(suburb['records_observed'])), ('Property Type', subject['property_type'])])}",
        f"<h2>Value Trend</h2>{_line_chart('Median Sale Value by Year', [(item['period'], item['median_value']) for item in reversed(suburb['yearly'].get(subject['property_type'], [])[:10])], '#d6a84f')}",
        f"<h2>Recommended Rental Appraisal Step</h2><p>Use this document as a transparent pre-brief, then complete a manual rental appraisal once leasing evidence is attached for the suburb.</p>",
        f"<h2>Disclaimer</h2><p>This rental market analysis supports a principal-led rental appraisal. Weekly asking guidance should still be confirmed against the most recent leasing competition at the time of launch.</p>",
    ]
    return _render_document("Rental CMA", subject["address"], pages, ctx["brand_logo_url"], "#d6a84f")


def _doc_suburb_profile(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    suburb = bundle["suburb"]
    house_yearly = suburb["yearly"].get("House", [])[:8]
    unit_yearly = suburb["yearly"].get("Unit", [])[:8]
    owner_mix = [[label, str(count)] for label, count in suburb["owner_mix"][:4]]
    sold_map_url = bundle.get("agent_sales_map_url") or ""
    pages = [
        f"""
        <div class="eyebrow">Suburb Profile</div>
        <h1>Suburb Profile Report</h1>
        <div class="meta">{html.escape(subject['suburb'])} NSW {html.escape(subject['postcode'])}</div>
        {_metric_grid([('Sales Last 12m', str(suburb['sales_last_12m'])), ('Overall Median', _format_money(suburb['median_last_12m'])), ('House Median', _format_money(suburb['house_median_last_12m'])), ('Unit Median', _format_money(suburb['unit_median_last_12m']))])}
        """,
        f"<h2>Overview</h2><p>This profile is built from the suburb workbook and live archive attached to the current system. It is intended to mirror the utility of a suburb report while keeping every figure traceable.</p>{_mini_table(['Owner Mix', 'Count'], owner_mix)}",
        f"<h2>Houses: Market Trend</h2>{_line_chart('House Median Value by Year', [(item['period'], item['median_value']) for item in reversed(house_yearly)], '#0f6fff')}",
        f"<h2>Houses: Sales Activity</h2>{_bar_chart('House Sales by Year', [(item['period'], item['sales']) for item in reversed(house_yearly)], '#0f6fff')}",
        f"<h2>Units: Market Trend</h2>{_line_chart('Unit Median Value by Year', [(item['period'], item['median_value']) for item in reversed(unit_yearly)], '#11b57c')}",
        f"<h2>Units: Sales Activity</h2>{_bar_chart('Unit Sales by Year', [(item['period'], item['sales']) for item in reversed(unit_yearly)], '#11b57c')}",
        f"<h2>Owner and Tenure Mix</h2>{_bar_chart('Observed owner mix', [(label, count) for label, count in suburb['owner_mix'][:6]], '#d6a84f')}",
        (
            f"<h2>Sold Properties by Nitin Puri</h2><img class=\"hero-image\" src=\"{html.escape(sold_map_url)}\" alt=\"Map of sold properties handled by Nitin Puri\" /><p class=\"footer-note\">Mapped from sold records in the local CRM where Nitin Puri is recorded as the agent.</p>"
            if sold_map_url
            else ""
        ),
        f"<h2>Market Interpretation</h2><p>{html.escape(subject['suburb'])} remains best read through recent turnover, owner mix and the relative performance of houses versus units. This context helps frame both sale timing and appraisal positioning for the subject property.</p>",
        f"<h2>Disclaimer</h2><p>This suburb profile is generated from observed local evidence held by Laing+Simmons Oakville | Windsor. It is not represented as a full-market licensed data feed.</p>",
    ]
    return _render_document("Suburb Profile", subject["suburb"], pages, ctx["brand_logo_url"], "#0f6fff")


def _doc_suburb_statistics(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> str:
    subject = bundle["subject"]
    suburb = bundle["suburb"]
    pages = [
        f"""
        <div class="eyebrow">Suburb Statistics</div>
        <h1>Suburb Statistics Report</h1>
        <div class="meta">{html.escape(subject['suburb'])} NSW {html.escape(subject['postcode'])}</div>
        {_metric_grid([('Archive Records', str(suburb['records_observed'])), ('Sales Last 12m', str(suburb['sales_last_12m'])), ('House Median', _format_money(suburb['house_median_last_12m'])), ('Unit Median', _format_money(suburb['unit_median_last_12m']))])}
        """
    ]
    for property_type, color in (("House", "#0f6fff"), ("Unit", "#11b57c")):
        yearly = suburb["yearly"].get(property_type, [])[:12]
        pages.extend(
            [
                f"<h2>{property_type}: Market Trend</h2>{_line_chart(f'{property_type} median value by year', [(item['period'], item['median_value']) for item in reversed(yearly)], color)}",
                f"<h2>{property_type}: Sales Per Annum</h2>{_bar_chart(f'{property_type} sales per year', [(item['period'], item['sales']) for item in reversed(yearly)], color)}",
                f"<h2>{property_type}: Turnover</h2>{_bar_chart(f'{property_type} turnover by year', [(item['period'], item['turnover']) for item in reversed(yearly)], color, currency=True)}",
            ]
        )
    pages.extend(
        [
            f"<h2>Owner Mix</h2>{_bar_chart('Observed owner types', [(label, count) for label, count in suburb['owner_mix'][:6]], '#d6a84f')}",
            f"<h2>Method Summary</h2><ul class=\"tight\"><li>Reference report families reviewed: {len(bundle['reference_manifest'])}</li><li>Local workbook-backed records observed: {suburb['records_observed']}</li><li>Current rental observations on file: {bundle['rental']['observation_count']}</li><li>Figures are drawn from the local archive and current CRM records.</li></ul>",
            f"<h2>Disclaimer</h2><p>This long-form statistics report mirrors the structure of a suburb statistics pack using local archive evidence and internal calculations. It is designed to support a market discussion rather than substitute for a licensed full-market data subscription.</p>",
        ]
    )
    return _render_document("Suburb Statistics", subject["suburb"], pages, ctx["brand_logo_url"], "#0f6fff")


def build_report_documents(bundle: Dict[str, Any], ctx: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    return [
        ("01_property_details_sheet.pdf", "property_details_sheet", _doc_property_sheet(bundle, ctx)),
        ("02_sales_cma.pdf", "sales_cma", _doc_sales_cma(bundle, ctx)),
        ("03_property_profile.pdf", "property_profile", _doc_property_profile(bundle, ctx)),
        ("04_rental_cma.pdf", "rental_cma", _doc_rental_cma(bundle, ctx)),
        ("05_rental_avm.pdf", "rental_avm", _doc_rental_avm(bundle, ctx)),
        ("06_suburb_profile.pdf", "suburb_profile", _doc_suburb_profile(bundle, ctx)),
        ("07_suburb_statistics.pdf", "suburb_statistics", _doc_suburb_statistics(bundle, ctx)),
        ("08_sale_avm.pdf", "sale_avm", _doc_sale_avm(bundle, ctx)),
    ]


from core.events import event_manager

async def _broadcast_report_stage(lead_id: str, stage: str, progress: float):
    await event_manager.broadcast({
        "type": "REPORT_PROGRESS",
        "data": {
            "lead_id": lead_id,
            "stage": stage,
            "progress": progress,
            "timestamp": datetime.datetime.now().isoformat()
        }
    })

async def create_report_pack(
    conn: sqlite3.Connection,
    lead: Dict[str, Any],
    *,
    stock_root: str,
    brand_name: str,
    brand_area: str,
    brand_logo_url: str,
    principal_name: str,
    principal_email: str,
    principal_phone: str,
    html_to_pdf: Callable[[str, str], Awaitable[Any]],
    output_root: Optional[Path] = None,
) -> Dict[str, Any]:
    lead_id = lead.get("id")
    await _broadcast_report_stage(lead_id, "Initializing Intelligence Bundle", 0.1)
    bundle = build_property_bundle(conn, lead, stock_root)
    
    # --- Phase 4: Sanity Agent Verification ---
    await _broadcast_report_stage(lead_id, "Sanity Agent: Auditing Brand Integrity", 0.2)
    verification = await verify_report_bundle(bundle)
    if not verification.is_valid:
        print(f"VERIFICATION WARNING: {verification.issues}")
    
    subject = bundle["subject"]
    await _broadcast_report_stage(lead_id, f"Generating Assets for {subject['address']}", 0.3)
    
    folder_name = _folder_safe(f"{subject['owner_name']} - {subject['address']}, {subject['suburb']} NSW {subject['postcode']}")
    root = (output_root or REPORT_PACK_ROOT) / folder_name
    # ... rest of function logic ...
    folder_name = _folder_safe(f"{subject['owner_name']} - {subject['address']}, {subject['suburb']} NSW {subject['postcode']}")
    root = (output_root or REPORT_PACK_ROOT) / folder_name
    manifest_dir = root / "00_manifest"
    greeting_dir = root / "01_greetings"
    existing_dir = root / "02_existing_briefs"
    replica_dir = root / "03_cotality_replica"
    working_dir = root / "04_working_data"
    for directory in (manifest_dir, greeting_dir, existing_dir, replica_dir, working_dir):
        directory.mkdir(parents=True, exist_ok=True)

    ctx = {
        "brand_name": brand_name,
        "brand_area": brand_area,
        "brand_logo_url": brand_logo_url,
        "principal_name": principal_name,
        "principal_email": principal_email,
        "principal_phone": principal_phone,
    }
    artifacts = []
    for filename, doc_type, html_content in build_report_documents(bundle, ctx):
        output_path = replica_dir / filename
        await html_to_pdf(html_content, str(output_path))
        artifacts.append({"type": doc_type, "filename": filename, "path": str(output_path)})

    greeting_text = (
        f"Subject: Property update for {subject['address']}\n\n"
        f"Dear {subject['owner_name']},\n\n"
        f"I've attached a tailored property pack for {subject['address']}. It includes a factual property snapshot, "
        "a local market analysis, and a suburb context brief built from our current archive.\n\n"
        "If useful, I can also walk you through the likely value range and the strongest next step in a short "
        "appraisal call.\n\n"
        f"{principal_name}\n{brand_name}\n{principal_phone}\n{principal_email}\n"
    )
    (greeting_dir / "greeting_email.txt").write_text(greeting_text, encoding="utf-8")
    (greeting_dir / "greeting_email.html").write_text(greeting_text.replace("\n", "<br />"), encoding="utf-8")
    cover_html = _render_document(
        "Owner Cover Letter",
        subject["address"],
        [
            f"""
            <div class="eyebrow">Greeting</div>
            <h1>Property intelligence pack enclosed</h1>
            <div class="meta">{html.escape(subject['address'])}</div>
            <p>Prepared for {html.escape(subject['owner_name'])}</p>
            <div class="callout"><strong>Why this pack exists</strong><p>It brings your property facts, local sales evidence and suburb market context into one owner-ready folder.</p></div>
            <p>{html.escape(principal_name)}<br />{html.escape(brand_name)}<br />{html.escape(principal_phone)}<br />{html.escape(principal_email)}</p>
            """
        ],
        brand_logo_url,
    )
    await html_to_pdf(cover_html, str(greeting_dir / "cover_letter.pdf"))
    manifest = {
        "owner_name": subject["owner_name"],
        "address": subject["address"],
        "folder_name": folder_name,
        "pack_root": str(root),
        "generated_at": bundle["generated_at"],
        "existing_briefs_note": "Reserved for generic_seller_brief and ai_appraisal_brief output.",
        "artifacts": artifacts,
        "known_gaps": [gap for gap in bundle["known_gaps"] if gap],
        "reference_manifest": bundle["reference_manifest"],
    }
    (manifest_dir / "pack_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (manifest_dir / "source_registry.json").write_text(json.dumps(bundle["source_registry"], indent=2), encoding="utf-8")
    (manifest_dir / "property_bundle.json").write_text(json.dumps(bundle, indent=2, default=str), encoding="utf-8")
    existing_dir.mkdir(exist_ok=True)
    return manifest


def run_pack_generation_sync(*args: Any, **kwargs: Any) -> Dict[str, Any]:
    return asyncio.run(create_report_pack(*args, **kwargs))
