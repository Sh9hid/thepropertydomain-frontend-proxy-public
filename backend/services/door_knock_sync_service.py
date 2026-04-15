from __future__ import annotations

import csv
import hashlib
import json
import logging
from pathlib import Path
from typing import Any

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.config import DOOR_KNOCK_SYNC_FILE, TEMP_DIR
from core.logic import get_deterministic_id
from core.utils import _dedupe_by_phone, _dedupe_text_list, now_iso

logger = logging.getLogger(__name__)

SYNC_HEADERS = [
    "lead_id",
    "address",
    "suburb",
    "postcode",
    "owner_name",
    "owner_type",
    "phone",
    "email",
    "source",
    "type",
    "queue",
    "tags",
    "notes",
    "updated_at",
]
LEGACY_HEADERS = [
    "Date",
    "Day",
    "Area / Street",
    "Address",
    "Lead Type",
    "Status",
    "Contact Name",
    "Phone",
    "Email",
    "Notes",
    "Action Required",
]

STATE_PATH = TEMP_DIR / "door_knock_sync_state.json"


def sync_enabled() -> bool:
    return bool((DOOR_KNOCK_SYNC_FILE or "").strip())


def _decode_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item or "").strip() for item in value if str(item or "").strip()]
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, list):
                return [str(item or "").strip() for item in parsed if str(item or "").strip()]
        except Exception:
            pass
        return [item.strip() for item in stripped.split(",") if item.strip()]
    return []


def _normalize_tag(token: str) -> str:
    raw = str(token or "").strip().lower().replace("-", "_").replace(" ", "_")
    if raw in {"doorknock", "door_knock"}:
        return "door_knock"
    if raw in {"rpdata", "rp_data", "rp_data_"}:
        return "rp_data"
    return raw.strip("_")


def _parse_house_token(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""
    lowered = text.lower()
    if lowered.startswith("house "):
        candidate = text[6:].strip()
        return candidate, f"House {candidate}".strip()
    return "", text


def _derive_address_from_legacy(row: dict[str, str]) -> str:
    street = str(row.get("Area / Street") or "").strip()
    house_num, fallback = _parse_house_token(row.get("Address") or "")
    if house_num and street:
        return f"{house_num} {street}".strip()
    if fallback and street:
        return f"{fallback} {street}".strip()
    return fallback or street


def _is_legacy_row(row: dict[str, str]) -> bool:
    lowered = {str(k or "").strip().lower() for k in row.keys()}
    return "area / street" in lowered and "lead type" in lowered


def _canonicalize_row(row: dict[str, str]) -> dict[str, str]:
    if _is_legacy_row(row):
        address = _derive_address_from_legacy(row)
        return {
            "lead_id": str(row.get("Lead ID") or row.get("lead_id") or "").strip(),
            "address": address,
            "suburb": str(row.get("Suburb") or row.get("suburb") or "").strip(),
            "postcode": str(row.get("Postcode") or row.get("postcode") or "").strip(),
            "owner_name": str(row.get("Contact Name") or "").strip(),
            "owner_type": str(row.get("Lead Type") or "").strip(),
            "phone": str(row.get("Phone") or "").strip(),
            "email": str(row.get("Email") or "").strip(),
            "source": "door_knock",
            "type": "Door Knock",
            "queue": "door_knock",
            "tags": str(row.get("Lead Type") or "").strip(),
            "notes": str(row.get("Notes") or "").strip(),
            "updated_at": str(row.get("updated_at") or "").strip(),
            "_format": "legacy",
        }
    next_row = {**row}
    next_row["_format"] = "sync"
    return next_row


def _parse_tags(value: str, owner_type: str, notes: str) -> list[str]:
    tags: list[str] = []
    for part in str(value or "").replace("/", ",").replace("|", ",").split(","):
        normalized = _normalize_tag(part)
        if normalized:
            tags.append(normalized)
    owner_token = _normalize_tag(owner_type)
    if owner_token:
        tags.append(owner_token)
    blob = f"{owner_type or ''} {notes or ''}".lower()
    if "builder" in blob:
        tags.append("builder")
    if "land" in blob:
        tags.append("land")
    tags.append("door_knock")
    return _dedupe_text_list(tags)


def _row_hash(row: dict[str, str]) -> str:
    payload = {
        "lead_id": row.get("lead_id", "").strip(),
        "address": row.get("address", "").strip().lower(),
        "suburb": row.get("suburb", "").strip().lower(),
        "postcode": row.get("postcode", "").strip(),
        "owner_name": row.get("owner_name", "").strip(),
        "owner_type": row.get("owner_type", "").strip().lower(),
        "phone": row.get("phone", "").strip(),
        "email": row.get("email", "").strip().lower(),
        "tags": row.get("tags", "").strip().lower(),
        "notes": row.get("notes", "").strip(),
    }
    return hashlib.md5(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def _row_key(row: dict[str, str]) -> str:
    lead_id = str(row.get("lead_id") or "").strip()
    if lead_id:
        return lead_id
    address = str(row.get("address") or "").strip().lower()
    suburb = str(row.get("suburb") or "").strip().lower()
    return hashlib.md5(f"{address}|{suburb}".encode("utf-8")).hexdigest()


def _sheet_path() -> Path:
    return Path(DOOR_KNOCK_SYNC_FILE).expanduser().resolve()


def _load_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            return [_canonicalize_row({k: str(v or "").strip() for k, v in row.items()}) for row in reader]
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows: list[dict[str, str]] = []
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(cell or "").strip() for cell in header_cells]
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[idx]: str(values[idx] or "").strip() for idx in range(min(len(headers), len(values))) if headers[idx]}
            if any(str(v).strip() for v in row.values()):
                rows.append(_canonicalize_row(row))
        wb.close()
        return rows
    raise ValueError(f"Unsupported sync file type: {path.suffix}")


def _write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    has_legacy = any(str(row.get("_format") or "") == "legacy" for row in rows)
    ordered_rows: list[dict[str, str]] = []
    headers = LEGACY_HEADERS if has_legacy else SYNC_HEADERS
    for row in rows:
        if has_legacy:
            street = ""
            full_address = str(row.get("address") or "").strip()
            house_num = ""
            parts = full_address.split(" ", 1)
            if len(parts) == 2 and any(ch.isdigit() for ch in parts[0]):
                house_num = parts[0]
                street = parts[1]
            lead_type = str(row.get("owner_type") or "").strip() or str(row.get("tags") or "").strip() or "Residential"
            status = str(row.get("status") or "").strip() or "Letter box drop off"
            ordered_rows.append(
                {
                    "Date": str(row.get("Date") or ""),
                    "Day": str(row.get("Day") or ""),
                    "Area / Street": str(row.get("Area / Street") or street),
                    "Address": str(row.get("Address") or (f"House {house_num}".strip() if house_num else full_address)),
                    "Lead Type": lead_type,
                    "Status": status,
                    "Contact Name": str(row.get("owner_name") or row.get("Contact Name") or ""),
                    "Phone": str(row.get("phone") or row.get("Phone") or ""),
                    "Email": str(row.get("email") or row.get("Email") or ""),
                    "Notes": str(row.get("notes") or row.get("Notes") or ""),
                    "Action Required": str(row.get("Action Required") or ""),
                }
            )
        else:
            ordered_rows.append({header: str(row.get(header, "") or "") for header in headers})
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers)
            writer.writeheader()
            writer.writerows(ordered_rows)
        return
    if suffix == ".xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in ordered_rows:
            ws.append([row.get(header, "") for header in headers])
        wb.save(path)
        wb.close()
        return
    raise ValueError(f"Unsupported sync file type: {path.suffix}")


def _load_state() -> dict[str, Any]:
    if not STATE_PATH.exists():
        return {"sheet_hashes": {}}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"sheet_hashes": {}}


def _save_state(state: dict[str, Any]) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=True, indent=2), encoding="utf-8")


def _lead_to_row(lead: dict[str, Any]) -> dict[str, str]:
    phones = _decode_list(lead.get("contact_phones"))
    emails = _decode_list(lead.get("contact_emails"))
    tags = _dedupe_text_list(_decode_list(lead.get("source_tags")))
    if "door_knock" not in tags:
        tags = ["door_knock", *tags]
    return {
        "lead_id": str(lead.get("id") or ""),
        "address": str(lead.get("address") or ""),
        "suburb": str(lead.get("suburb") or ""),
        "postcode": str(lead.get("postcode") or ""),
        "owner_name": str(lead.get("owner_name") or ""),
        "owner_type": str(lead.get("owner_type") or ""),
        "phone": phones[0] if phones else "",
        "email": emails[0] if emails else "",
        "source": "door_knock",
        "type": "Door Knock",
        "queue": "door_knock",
        "tags": ",".join(tags),
        "notes": str(lead.get("notes") or ""),
        "updated_at": str(lead.get("updated_at") or ""),
    }


async def _fetch_door_knock_leads(session: AsyncSession) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            text(
                """
                SELECT id, address, suburb, postcode, owner_name, owner_type, contact_phones, contact_emails,
                       source_tags, notes, updated_at, queue_bucket, trigger_type
                FROM leads
                WHERE COALESCE(queue_bucket, '') = 'door_knock'
                   OR LOWER(COALESCE(trigger_type, '')) LIKE '%door%'
                   OR LOWER(COALESCE(CAST(source_tags AS TEXT), '[]')) LIKE '%door_knock%'
                """
            )
        )
    ).mappings().all()
    return [dict(row) for row in rows]


async def _apply_sheet_row(session: AsyncSession, row: dict[str, str], now: str) -> bool:
    address = str(row.get("address") or "").strip()
    if not address:
        return False
    lead_id = str(row.get("lead_id") or "").strip() or get_deterministic_id(address)
    existing = (
        await session.execute(
            text(
                """
                SELECT id, contact_phones, contact_emails, source_tags, owner_name, owner_type, suburb, postcode, notes
                FROM leads
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": lead_id},
        )
    ).mappings().first()
    if existing is None:
        existing = (
            await session.execute(
                text(
                    """
                    SELECT id, contact_phones, contact_emails, source_tags, owner_name, owner_type, suburb, postcode, notes
                    FROM leads
                    WHERE LOWER(COALESCE(address, '')) = :address
                      AND (:suburb = '' OR LOWER(COALESCE(suburb, '')) = :suburb)
                    LIMIT 1
                    """
                ),
                {"address": address.lower(), "suburb": str(row.get("suburb") or "").strip().lower()},
            )
        ).mappings().first()
    tags = _parse_tags(row.get("tags", ""), row.get("owner_type", ""), row.get("notes", ""))
    owner_name = str(row.get("owner_name") or "").strip()
    owner_type = str(row.get("owner_type") or "").strip()
    phone = str(row.get("phone") or "").strip()
    email = str(row.get("email") or "").strip()
    notes = str(row.get("notes") or "").strip()
    suburb = str(row.get("suburb") or "").strip()
    postcode = str(row.get("postcode") or "").strip()

    if existing is None:
        await session.execute(
            text(
                """
                INSERT INTO leads (
                    id, address, suburb, postcode, owner_name, owner_type, contact_phones, contact_emails,
                    source_tags, trigger_type, queue_bucket, record_type, status, notes, stage_note, created_at, updated_at
                ) VALUES (
                    :id, :address, :suburb, :postcode, :owner_name, :owner_type, :contact_phones, :contact_emails,
                    :source_tags, :trigger_type, :queue_bucket, :record_type, :status, :notes, :stage_note, :created_at, :updated_at
                )
                """
            ),
            {
                "id": lead_id,
                "address": address,
                "suburb": suburb,
                "postcode": postcode,
                "owner_name": owner_name or "Owner record pending",
                "owner_type": owner_type or None,
                "contact_phones": json.dumps(_dedupe_by_phone([phone] if phone else [])),
                "contact_emails": json.dumps(_dedupe_text_list([email] if email else [])),
                "source_tags": json.dumps(tags),
                "trigger_type": "Door Knock",
                "queue_bucket": "door_knock",
                "record_type": "manual_entry",
                "status": "captured",
                "notes": notes,
                "stage_note": notes,
                "created_at": now,
                "updated_at": now,
            },
        )
        return True

    current = dict(existing)
    next_phones = _dedupe_by_phone([*_decode_list(current.get("contact_phones")), *([phone] if phone else [])])
    next_emails = _dedupe_text_list([*_decode_list(current.get("contact_emails")), *([email] if email else [])])
    next_tags = _dedupe_text_list([*_decode_list(current.get("source_tags")), *tags])
    await session.execute(
        text(
            """
            UPDATE leads
            SET owner_name = CASE WHEN :owner_name != '' THEN :owner_name ELSE owner_name END,
                owner_type = CASE WHEN :owner_type != '' THEN :owner_type ELSE owner_type END,
                suburb = CASE WHEN :suburb != '' THEN :suburb ELSE suburb END,
                postcode = CASE WHEN :postcode != '' THEN :postcode ELSE postcode END,
                contact_phones = :contact_phones,
                contact_emails = :contact_emails,
                source_tags = :source_tags,
                trigger_type = 'Door Knock',
                queue_bucket = 'door_knock',
                notes = CASE WHEN :notes != '' THEN :notes ELSE notes END,
                stage_note = CASE WHEN :notes != '' THEN :notes ELSE stage_note END,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "owner_name": owner_name,
            "owner_type": owner_type,
            "suburb": suburb,
            "postcode": postcode,
            "contact_phones": json.dumps(next_phones),
            "contact_emails": json.dumps(next_emails),
            "source_tags": json.dumps(next_tags),
            "notes": notes,
            "updated_at": now,
            "id": str(current.get("id")),
        },
    )
    return True


async def run_door_knock_sheet_sync_once(session: AsyncSession) -> dict[str, int]:
    if not sync_enabled():
        return {"imported": 0, "exported": 0, "total_sheet_rows": 0, "total_leads": 0}

    path = _sheet_path()
    now = now_iso()
    sheet_rows = _load_rows(path)
    state = _load_state()
    sheet_hashes: dict[str, str] = dict(state.get("sheet_hashes") or {})

    imported = 0
    for row in sheet_rows:
        key = _row_key(row)
        digest = _row_hash(row)
        if sheet_hashes.get(key) == digest:
            continue
        changed = await _apply_sheet_row(session, row, now)
        if changed:
            imported += 1
        sheet_hashes[key] = digest

    leads = await _fetch_door_knock_leads(session)
    rows_by_key = {_row_key(row): row for row in sheet_rows}
    exported = 0
    for lead in leads:
        row = _lead_to_row(lead)
        key = _row_key(row)
        existing = rows_by_key.get(key)
        if existing is None:
            rows_by_key[key] = row
            exported += 1
            continue
        merged = {**existing, **row}
        if _is_legacy_row(existing):
            merged["_format"] = "legacy"
            if not merged.get("Area / Street"):
                parts = str(row.get("address") or "").split(" ", 1)
                merged["Area / Street"] = parts[1] if len(parts) == 2 else str(row.get("address") or "")
            if not merged.get("Address"):
                parts = str(row.get("address") or "").split(" ", 1)
                house = parts[0] if parts and any(ch.isdigit() for ch in parts[0]) else ""
                merged["Address"] = f"House {house}".strip() if house else str(row.get("address") or "")
            merged["Contact Name"] = row.get("owner_name", "")
            merged["Phone"] = row.get("phone", "")
            merged["Email"] = row.get("email", "")
            merged["Notes"] = row.get("notes", "")
            if not merged.get("Lead Type"):
                merged["Lead Type"] = row.get("owner_type") or "Residential"
        if merged != existing:
            rows_by_key[key] = merged
            exported += 1

    merged_rows = list(rows_by_key.values())
    if exported > 0 or (not path.exists() and leads):
        _write_rows(path, merged_rows)

    final_hashes = {_row_key(row): _row_hash(row) for row in merged_rows}
    _save_state({"sheet_hashes": final_hashes, "synced_at": now})
    await session.commit()

    return {
        "imported": imported,
        "exported": exported,
        "total_sheet_rows": len(merged_rows),
        "total_leads": len(leads),
    }
