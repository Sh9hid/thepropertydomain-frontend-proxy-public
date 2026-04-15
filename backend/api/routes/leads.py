import csv
import datetime
import html
import asyncio
import hashlib
import hmac
import io
import json
import logging
import os
import re
import uuid
import smtplib

logger = logging.getLogger(__name__)
from base64 import b64encode
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request
from urllib.parse import urlencode

import httpx
from fastapi import APIRouter, Depends, HTTPException, Security, Request, BackgroundTasks, File, UploadFile, Form
from zoneinfo import ZoneInfo

from api.routes._deps import APIKeyDep, SessionDep
from api.routes.leads_support import (
    BulkLeadUpdate,
    DirectSendEmailRequest,
    DirectSendSMSRequest,
    LeadFollowupPayload,
    LeadPatchPayload,
    LogCallRequest,
    attach_deterministic_intelligence,
    delta_values,
    rank_leads_for_hit_list,
)
from core.config import (
    API_KEY, api_key_header, APP_TITLE, SYDNEY_TZ, STOCK_ROOT, 
    PROJECT_ROOT, PROJECT_LOG_PATH, BRAND_NAME, BRAND_AREA, BRAND_LOGO_URL, 
    PRINCIPAL_NAME, PRINCIPAL_EMAIL, PRINCIPAL_PHONE, PROJECT_MEMORY_RULE, 
    BACKGROUND_SEND_POLL_SECONDS, PRIMARY_STRIKE_SUBURB, SECONDARY_STRIKE_SUBURBS,
    SMS_BRIDGE_URL, USE_POSTGRES, GENERATED_REPORTS_ROOT, build_public_url
)
from core.utils import (
    now_sydney, now_iso, format_sydney, parse_client_datetime, 
    _first_non_empty, _safe_int, _format_moneyish, _parse_json_list, 
    _encode_value, _decode_row, _dedupe_text_list, _normalize_phone, 
    _dedupe_by_phone, _parse_iso_datetime, _parse_calendar_date, 
    _month_range_from_date, _bool_db
)
from services.scoring import _trigger_bonus, _status_penalty, _score_lead
from models.schemas import *
from core.logic import *

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select, col, or_, and_
from core.database import (
    _fetch_joined_task, _get_lead_or_404, get_session,
    _task_to_dict, _appointment_to_dict, _sold_event_to_dict
)
from models.sql_models import Lead as SQLLead
from services.automations import (
    _schedule_task, _refresh_lead_next_action, _supersede_auto_tasks,
    _schedule_callback_cadence, _schedule_enrichment_task, 
    _schedule_nurture_cadence, _schedule_booked_followthrough
)
from services.integrations import send_email_service
from models.schemas import SendEmailRequest

from core.security import get_api_key
from services.funnel_service import assert_outreach_allowed, resolve_outreach_purpose
from services.lead_summary_service import (
    get_lead_detail_payload,
    list_ranked_leads,
    list_lead_summaries,
)
from services.lead_search_service import search_leads_hybrid, refresh_lead_search_index
try:
    from services.followup_service import get_followup_state, update_followup_preferences
except Exception:
    async def get_followup_state(*a, **kw): return {}
    async def update_followup_preferences(*a, **kw): return {}
try:
    from services.door_knock_sync_service import run_door_knock_sheet_sync_once, sync_enabled as door_knock_sync_enabled
except Exception:
    async def run_door_knock_sheet_sync_once(*a, **kw): return {"imported": 0, "exported": 0, "total_sheet_rows": 0, "total_leads": 0}
    def door_knock_sync_enabled() -> bool: return False
try:
    from services.lead_read_cache import invalidate_lead_read_models
except Exception:
    async def invalidate_lead_read_models(*a, **kw): pass
try:
    from services.hermes_lead_ops_service import refresh_hermes_for_lead, refresh_hermes_for_leads
except Exception:
    async def refresh_hermes_for_lead(*a, **kw): pass
    async def refresh_hermes_for_leads(*a, **kw): pass
try:
    from services.address_identity import build_address_identity, classify_match
except Exception:
    def build_address_identity(*a, **kw): return {}
    def classify_match(*a, **kw): return "unknown"
from services.pipeline_guard import assert_status_transition_allowed
try:
    from services.lead_hygiene import apply_precall_hygiene
except Exception:
    async def apply_precall_hygiene(*args, **kwargs):
        return {"processed": 0, "updated": 0, "errors": 0, "status": "fallback"}

router = APIRouter()

# Recordings endpoint moved to api/routes/recordings.py


@router.get("/api/leads")
async def get_leads(
    limit: int = 100, 
    offset: int = 0, 
    search: str = None, 
    is_fresh: bool = False,
    signal_status: Optional[str] = None,
    min_dom: Optional[int] = None,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    try:
        return await list_ranked_leads(
            session,
            limit=limit,
            offset=offset,
            search=search,
            is_fresh=is_fresh,
            signal_status=signal_status,
            min_dom=min_dom,
        )
    except Exception:
        logger.exception("get_leads failed; using emergency fallback query")
        safe_limit = max(1, min(int(limit or 100), 500))
        safe_offset = max(0, int(offset or 0))
        params: dict[str, Any] = {"limit": safe_limit, "offset": safe_offset}
        try:
            is_sqlite = bool(session.bind and str(session.bind.url).startswith("sqlite"))
            if is_sqlite:
                col_rows = await session.execute(text("PRAGMA table_info(leads)"))
                columns = {str(row[1]).lower() for row in col_rows.fetchall()}
            else:
                col_rows = await session.execute(
                    text(
                        """
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_name = 'leads'
                        """
                    )
                )
                columns = {str(row[0]).lower() for row in col_rows.fetchall()}

            where_parts: list[str] = []
            if str(search or "").strip():
                token = f"%{str(search).strip().lower()}%"
                search_cols = [
                    col_name
                    for col_name in (
                        "address",
                        "owner_name",
                        "suburb",
                        "canonical_address",
                        "postcode",
                        "contact_phones",
                        "contact_emails",
                        "notes",
                        "source",
                        "type",
                    )
                    if col_name in columns
                ]
                if search_cols:
                    where_parts.append(
                        "("
                        + " OR ".join(
                            [f"LOWER(COALESCE(CAST({col_name} AS TEXT), '')) LIKE :token" for col_name in search_cols]
                        )
                        + ")"
                    )
                    params["token"] = token

            where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""

            order_parts: list[str] = []
            if "updated_at" in columns:
                order_parts.append("updated_at DESC")
            if "created_at" in columns:
                order_parts.append("created_at DESC")
            if "id" in columns:
                order_parts.append("id ASC")
            order_sql = f"ORDER BY {', '.join(order_parts)}" if order_parts else ""

            rows = (
                await session.execute(
                    text(
                        f"""
                        SELECT *
                        FROM leads
                        {where_sql}
                        {order_sql}
                        LIMIT :limit OFFSET :offset
                        """
                    ),
                    params,
                )
            ).mappings().all()

            total = int(
                (
                    await session.execute(
                        text(f"SELECT COUNT(*) AS cnt FROM leads {where_sql}"),
                        {k: v for k, v in params.items() if k == "token"},
                    )
                ).scalar_one()
            )

            leads = [_hydrate_lead(dict(row)) for row in rows]
            return {"leads": leads, "total": total, "fallback": True}
        except Exception:
            logger.exception("get_leads emergency fallback failed; returning empty payload")
            return {"leads": [], "total": 0, "fallback": True}


@router.get("/api/leads/search")
async def search_leads(
    q: str,
    limit: int = 50,
    signal_status: Optional[str] = None,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    result = await search_leads_hybrid(
        session,
        q=q,
        limit=limit,
        signal_status=signal_status,
    )
    return result


@router.post("/api/leads/search/rebuild-index")
async def rebuild_lead_search_index(
    batch_size: int = 3000,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    result = await refresh_lead_search_index(session, batch_size=batch_size)
    return result


@router.get("/api/leads/get_next_leads_to_call")
async def get_next_leads_to_call(
    limit: int = 25,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    safe_limit = max(1, min(int(limit or 25), 100))
    await apply_precall_hygiene(session, limit=1200)
    rows = (
        await session.execute(
            text(
                """
                SELECT *
                FROM leads
                WHERE COALESCE(status, 'captured') NOT IN ('converted', 'dropped', 'appt_booked', 'mortgage_appt_booked')
                ORDER BY COALESCE(call_today_score, 0) DESC, COALESCE(heat_score, 0) DESC, COALESCE(updated_at, created_at, '') DESC, id ASC
                LIMIT :limit
                """
            ),
            {"limit": safe_limit},
        )
    ).mappings().all()
    leads = await attach_deterministic_intelligence(session, [_hydrate_lead(dict(row)) for row in rows])
    call_ready = [lead for lead in leads if lead.get("lead_state") == "ready_to_call"]
    call_ready.sort(
        key=lambda lead: (
            -(float(lead.get("priority_rank") or 0)),
            -int(lead.get("call_today_score") or 0),
            -int(lead.get("heat_score") or 0),
            str(lead.get("id") or ""),
        )
    )
    return {"total": len(call_ready), "leads": call_ready}


@router.get("/api/leads/summary")
async def get_lead_summary(
    limit: int = 100,
    offset: int = 0,
    search: str = None,
    is_fresh: bool = False,
    signal_status: Optional[str] = None,
    min_dom: Optional[int] = None,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    return await list_lead_summaries(
        session,
        limit=limit,
        offset=offset,
        search=search,
        is_fresh=is_fresh,
        signal_status=signal_status,
        min_dom=min_dom,
    )


@router.get("/api/leads/{lead_id}")
async def get_lead_detail(lead_id: str, session: SessionDep = None, api_key: APIKeyDep = ""):
    from services.hermes_lead_ops_service import get_hermes_lead

    lead = await session.get(SQLLead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    hydrated = _hydrate_lead(lead.model_dump())
    enriched = (await attach_deterministic_intelligence(session, [hydrated]))[0]
    try:
        enriched["hermes"] = await get_hermes_lead(session, lead_id)
    except Exception:
        enriched["hermes"] = None
    return enriched


@router.get("/api/leads/{lead_id}/detail")
async def get_lead_detail_expanded(
    lead_id: str,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    from services.hermes_lead_ops_service import get_hermes_lead

    payload = await get_lead_detail_payload(session, lead_id, include_timeline=True)
    if isinstance(payload, dict):
        try:
            payload["hermes"] = await get_hermes_lead(session, lead_id)
        except Exception:
            payload["hermes"] = None
    return payload


@router.get("/api/leads/{lead_id}/follow-up")
async def get_lead_followup(
    lead_id: str,
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    try:
        return await get_followup_state(session, lead_id)
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.patch("/api/leads/{lead_id}/follow-up")
async def patch_lead_followup(
    lead_id: str,
    payload: "LeadFollowupPayload",
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    try:
        result = await update_followup_preferences(
            session,
            lead_id,
            payload.model_dump(exclude_unset=True),
            actor="operator",
            source="api",
        )
        await refresh_hermes_for_lead(session, lead_id, actor="lead_followup_patch")
        return result
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

@router.patch("/api/leads/{lead_id}")
async def patch_lead(lead_id: str, payload: LeadPatchPayload, session: AsyncSession = Depends(get_session), api_key: str = Depends(get_api_key)):
    """Update mutable fields on a lead (inline edit from EntityOS)."""
    now = now_iso()
    body = payload.model_dump()
    manual_note = (body.pop("notes", "") or "").strip()
    updates = {k: v for k, v in body.items() if v is not None}
    if not updates and not manual_note:
        raise HTTPException(status_code=400, detail="No fields to update")

    if "contact_phones" in updates:
        updates["contact_phones"] = _dedupe_by_phone(updates.get("contact_phones"))
    if "contact_emails" in updates:
        updates["contact_emails"] = _dedupe_text_list(updates.get("contact_emails"))

    if USE_POSTGRES:
        lead = await session.get(SQLLead, lead_id)
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        current_status = lead.status or "captured"
        target_status = updates.get("status") or current_status
        if target_status != current_status:
            transition_view = {**lead.model_dump(), **updates}
            assert_status_transition_allowed(
                transition_view,
                target_status,
                source="lead_patch",
                appointment_at=(
                    str(updates.get("next_action_at") or "")
                    or str(updates.get("follow_up_due_at") or "")
                    or str(lead.next_action_at or "")
                    or str(lead.follow_up_due_at or "")
                ),
            )
        note_history = list(lead.stage_note_history or [])
        activity_log = list(lead.activity_log or [])
        generated_notes: List[str] = []

        if "owner_name" in updates and (lead.owner_name or "").strip() != (updates["owner_name"] or "").strip():
            old_owner = (lead.owner_name or "blank").strip()
            new_owner = (updates["owner_name"] or "blank").strip()
            note = f"Owner updated - {old_owner} -> {new_owner}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("owner_updated", note, target_status, "profile"))

        if "contact_phones" in updates:
            previous_phones = _dedupe_by_phone(lead.contact_phones or [])
            next_phones = updates["contact_phones"] or []
            added_phones, removed_phones = delta_values(previous_phones, next_phones, _normalize_phone)
            for phone in added_phones:
                note = f"Phone added - {phone}"
                generated_notes.append(note)
                note_history = _append_stage_note(note_history, note, target_status)
                activity_log = _append_activity(activity_log, _build_activity_entry("phone_added", note, target_status, "profile", recipient=phone))
            for phone in removed_phones:
                note = f"Phone removed - {phone}"
                generated_notes.append(note)
                note_history = _append_stage_note(note_history, note, target_status)
                activity_log = _append_activity(activity_log, _build_activity_entry("phone_removed", note, target_status, "profile", recipient=phone))

        if "contact_emails" in updates:
            previous_emails = _dedupe_text_list(lead.contact_emails or [])
            next_emails = updates["contact_emails"] or []
            added_emails, removed_emails = delta_values(previous_emails, next_emails, lambda value: str(value or "").strip().lower())
            for email in added_emails:
                note = f"Email added - {email}"
                generated_notes.append(note)
                note_history = _append_stage_note(note_history, note, target_status)
                activity_log = _append_activity(activity_log, _build_activity_entry("email_added", note, target_status, "profile", recipient=email))
            for email in removed_emails:
                note = f"Email removed - {email}"
                generated_notes.append(note)
                note_history = _append_stage_note(note_history, note, target_status)
                activity_log = _append_activity(activity_log, _build_activity_entry("email_removed", note, target_status, "profile", recipient=email))

        if "status" in updates and target_status != current_status:
            note = f"Status changed - {current_status} -> {target_status}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("status_change", note, target_status, "status"))

        if manual_note:
            generated_notes.append(manual_note)
            note_history = _append_stage_note(note_history, manual_note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("note", manual_note, target_status, "crm_note"))

        for k, v in updates.items():
            setattr(lead, k, v)
        if generated_notes:
            lead.stage_note = generated_notes[-1]
            lead.stage_note_history = note_history
            lead.activity_log = activity_log
        if target_status == "contacted" and current_status != "contacted":
            lead.last_contacted_at = now
        lead.updated_at = now
        await session.commit()
        await refresh_hermes_for_lead(session, lead_id, actor="lead_patch")
        await session.refresh(lead)
        invalidate_lead_read_models([lead_id])
        return _hydrate_lead(lead.model_dump())

    # SQLite path
    lead_row = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()
    if not lead_row:
        raise HTTPException(status_code=404, detail="Lead not found")
    lead = _decode_row(dict(lead_row))
    current_status = lead.get("status") or "captured"
    target_status = updates.get("status") or current_status
    if target_status != current_status:
        transition_view = {**lead, **updates}
        assert_status_transition_allowed(
            transition_view,
            target_status,
            source="lead_patch",
            appointment_at=(
                str(updates.get("next_action_at") or "")
                or str(updates.get("follow_up_due_at") or "")
                or str(lead.get("next_action_at") or "")
                or str(lead.get("follow_up_due_at") or "")
            ),
        )
    note_history = lead.get("stage_note_history")
    activity_log = lead.get("activity_log")
    generated_notes: List[str] = []

    if "owner_name" in updates and (lead.get("owner_name") or "").strip() != (updates["owner_name"] or "").strip():
        old_owner = (lead.get("owner_name") or "blank").strip()
        new_owner = (updates["owner_name"] or "blank").strip()
        note = f"Owner updated - {old_owner} -> {new_owner}"
        generated_notes.append(note)
        note_history = _append_stage_note(note_history, note, target_status)
        activity_log = _append_activity(activity_log, _build_activity_entry("owner_updated", note, target_status, "profile"))

    if "contact_phones" in updates:
        previous_phones = _dedupe_by_phone(lead.get("contact_phones"))
        next_phones = updates["contact_phones"] or []
        added_phones, removed_phones = delta_values(previous_phones, next_phones, _normalize_phone)
        for phone in added_phones:
            note = f"Phone added - {phone}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("phone_added", note, target_status, "profile", recipient=phone))
        for phone in removed_phones:
            note = f"Phone removed - {phone}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("phone_removed", note, target_status, "profile", recipient=phone))

    if "contact_emails" in updates:
        previous_emails = _dedupe_text_list(lead.get("contact_emails"))
        next_emails = updates["contact_emails"] or []
        added_emails, removed_emails = delta_values(previous_emails, next_emails, lambda value: str(value or "").strip().lower())
        for email in added_emails:
            note = f"Email added - {email}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("email_added", note, target_status, "profile", recipient=email))
        for email in removed_emails:
            note = f"Email removed - {email}"
            generated_notes.append(note)
            note_history = _append_stage_note(note_history, note, target_status)
            activity_log = _append_activity(activity_log, _build_activity_entry("email_removed", note, target_status, "profile", recipient=email))

    if "status" in updates and target_status != current_status:
        note = f"Status changed - {current_status} -> {target_status}"
        generated_notes.append(note)
        note_history = _append_stage_note(note_history, note, target_status)
        activity_log = _append_activity(activity_log, _build_activity_entry("status_change", note, target_status, "status"))

    if manual_note:
        generated_notes.append(manual_note)
        note_history = _append_stage_note(note_history, manual_note, target_status)
        activity_log = _append_activity(activity_log, _build_activity_entry("note", manual_note, target_status, "crm_note"))

    if generated_notes:
        updates["stage_note"] = generated_notes[-1]
        updates["stage_note_history"] = json.dumps(note_history or [])
        updates["activity_log"] = json.dumps(activity_log or [])
    if target_status == "contacted" and current_status != "contacted":
        updates["last_contacted_at"] = now

    set_parts = ", ".join(f"{k} = :{k}" for k in updates)
    params = {**updates, "lead_id": lead_id, "now": now}
    await session.execute(
        text(f"UPDATE leads SET {set_parts}, updated_at = :now WHERE id = :lead_id"),
        params,
    )
    await session.commit()
    await refresh_hermes_for_lead(session, lead_id, actor="lead_patch")
    row = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Lead not found")
    invalidate_lead_read_models([lead_id])
    return _hydrate_lead(dict(row))

@router.post("/api/leads/bulk-update")
async def bulk_update_leads(
    payload: BulkLeadUpdate,
    session: AsyncSession = Depends(get_session),
    api_key: str = Depends(get_api_key),
):
    """Bulk update status, route_queue, or follow_up_due_at on multiple leads."""
    allowed_fields = {"status", "route_queue", "follow_up_due_at"}
    update = {k: v for k, v in payload.update.items() if k in allowed_fields}
    if not update:
        raise HTTPException(status_code=400, detail="No valid update fields provided. Allowed: status, route_queue, follow_up_due_at")
    if not payload.lead_ids:
        raise HTTPException(status_code=400, detail="lead_ids must not be empty")

    now = now_iso()
    updated_ids: List[str] = []

    failed: List[Dict[str, Any]] = []
    status_target = str(update.get("status") or "").strip()

    from sqlalchemy import select as sa_select
    rows = (
        await session.execute(
            sa_select(SQLLead).where(SQLLead.id.in_(list(payload.lead_ids)))
        )
    ).scalars().all()
    row_by_id = {str(row.id or ""): row.model_dump() for row in rows}

    for lead_id in payload.lead_ids:
        lead_row = row_by_id.get(str(lead_id))
        if not lead_row:
            failed.append(
                {
                    "lead_id": str(lead_id),
                    "code": "lead_not_found",
                    "message": "Lead not found",
                }
            )
            continue

        if status_target:
            current_status = str(lead_row.get("status") or "captured")
            if status_target != current_status:
                transition_view = {**lead_row, **update}
                try:
                    assert_status_transition_allowed(
                        transition_view,
                        status_target,
                        source="lead_bulk_patch",
                        appointment_at=(
                            str(update.get("next_action_at") or "")
                            or str(update.get("follow_up_due_at") or "")
                            or str(lead_row.get("next_action_at") or "")
                            or str(lead_row.get("follow_up_due_at") or "")
                        ),
                    )
                except HTTPException as exc:
                    failed.append(
                        {
                            "lead_id": str(lead_id),
                            "code": "pipeline_stage_requirements_missing",
                            "detail": exc.detail,
                        }
                    )
                    continue

        set_fields = {**update, "updated_at": now}
        set_clause = ", ".join(f"{key} = :{key}" for key in set_fields.keys())
        params = {**set_fields, "lead_id": str(lead_id)}
        await session.execute(
            text(f"UPDATE leads SET {set_clause} WHERE id = :lead_id"),
            params,
        )
        updated_ids.append(str(lead_id))

    await session.commit()
    if updated_ids:
        await refresh_hermes_for_leads(session, updated_ids, actor="lead_bulk_patch")

    invalidate_lead_read_models(updated_ids)
    return {
        "updated": len(updated_ids),
        "lead_ids": updated_ids,
        "failed": failed,
    }


@router.post("/api/leads/admin/normalize-migrated-sources")
async def normalize_migrated_sources(
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    """
    Production data cleanup for migrated door-knock rows:
    1) trigger_type Manual/manual_entry -> RP Data
    2) source_tags reordered to start with rp_data, door_knock and remove manual/legacy tags
    3) scrub 'mansi' (case-insensitive) from notes/stage_note/history/activity text
    """
    try:
        pre = (
            await session.execute(
                text(
                    """
                    SELECT
                        COUNT(*) AS total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(trigger_type, '')) IN ('manual', 'manual_entry')
                        ) AS manual_total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(CAST(source_tags AS text), '[]')) LIKE '%manual%'
                        ) AS manual_tag_total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(notes, '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(stage_note, '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(CAST(stage_note_history AS text), '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(CAST(activity_log AS text), '')) LIKE '%mansi%'
                        ) AS mansi_rows
                    FROM leads
                    """
                )
            )
        ).mappings().one()

        manual_to_rp = (
            await session.execute(
                text(
                    """
                    UPDATE leads
                    SET trigger_type = 'RP Data',
                        updated_at = :now
                    WHERE LOWER(COALESCE(trigger_type, '')) IN ('manual', 'manual_entry')
                    """
                ),
                {"now": now_iso()},
            )
        ).rowcount or 0

        source_reordered = (
            await session.execute(
                text(
                    """
                    UPDATE leads l
                    SET source_tags = (
                        SELECT COALESCE(
                            jsonb_agg(v ORDER BY min_ord),
                            '[]'::jsonb
                        )
                        FROM (
                            SELECT v, MIN(ord) AS min_ord
                            FROM unnest(
                                ARRAY['rp_data', 'door_knock'] ||
                                COALESCE(
                                    ARRAY(
                                        SELECT jsonb_array_elements_text(
                                            CASE
                                                WHEN l.source_tags IS NULL THEN '[]'::jsonb
                                                ELSE l.source_tags::jsonb
                                            END
                                        )
                                    ),
                                    ARRAY[]::text[]
                                )
                            ) WITH ORDINALITY AS t(v, ord)
                            WHERE LOWER(v) NOT IN ('manual', 'manual_entry', 'legacy_migration')
                            GROUP BY v
                        ) ranked
                    ),
                    updated_at = :now
                    WHERE LOWER(COALESCE(CAST(source_tags AS text), '[]')) LIKE '%door_knock%'
                       OR LOWER(COALESCE(CAST(source_tags AS text), '[]')) LIKE '%rp_data%'
                       OR LOWER(COALESCE(trigger_type, '')) = 'rp data'
                    """
                ),
                {"now": now_iso()},
            )
        ).rowcount or 0

        mansi_scrubbed = (
            await session.execute(
                text(
                    """
                    UPDATE leads
                    SET notes = REGEXP_REPLACE(COALESCE(notes, ''), '(?i)mansi', '', 'g'),
                        stage_note = REGEXP_REPLACE(COALESCE(stage_note, ''), '(?i)mansi', '', 'g'),
                        stage_note_history = CASE
                            WHEN stage_note_history IS NULL THEN stage_note_history
                            ELSE REGEXP_REPLACE(CAST(stage_note_history AS text), '(?i)mansi', '', 'g')::jsonb
                        END,
                        activity_log = CASE
                            WHEN activity_log IS NULL THEN activity_log
                            ELSE REGEXP_REPLACE(CAST(activity_log AS text), '(?i)mansi', '', 'g')::jsonb
                        END,
                        updated_at = :now
                    WHERE LOWER(COALESCE(notes, '')) LIKE '%mansi%'
                       OR LOWER(COALESCE(stage_note, '')) LIKE '%mansi%'
                       OR LOWER(COALESCE(CAST(stage_note_history AS text), '')) LIKE '%mansi%'
                       OR LOWER(COALESCE(CAST(activity_log AS text), '')) LIKE '%mansi%'
                    """
                ),
                {"now": now_iso()},
            )
        ).rowcount or 0

        await session.commit()
        _invalidate_analytics_cache()

        post = (
            await session.execute(
                text(
                    """
                    SELECT
                        COUNT(*) AS total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(trigger_type, '')) IN ('manual', 'manual_entry')
                        ) AS manual_total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(CAST(source_tags AS text), '[]')) LIKE '%manual%'
                        ) AS manual_tag_total,
                        COUNT(*) FILTER (
                            WHERE LOWER(COALESCE(notes, '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(stage_note, '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(CAST(stage_note_history AS text), '')) LIKE '%mansi%'
                               OR LOWER(COALESCE(CAST(activity_log AS text), '')) LIKE '%mansi%'
                        ) AS mansi_rows
                    FROM leads
                    """
                )
            )
        ).mappings().one()

        sample = (
            await session.execute(
                text(
                    """
                    SELECT id, trigger_type, source_tags, notes, stage_note
                    FROM leads
                    WHERE LOWER(COALESCE(CAST(source_tags AS text), '[]')) LIKE '%door_knock%'
                    ORDER BY updated_at DESC NULLS LAST
                    LIMIT 10
                    """
                )
            )
        ).mappings().all()

        return {
            "ok": True,
            "updated": {
                "manual_to_rp": int(manual_to_rp),
                "source_reordered": int(source_reordered),
                "mansi_scrubbed": int(mansi_scrubbed),
            },
            "before": dict(pre),
            "after": dict(post),
            "sample": [dict(row) for row in sample],
        }
    except Exception as exc:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"cleanup_failed: {exc}") from exc


@router.get("/api/leads/{lead_id}/terminal")
async def get_lead_terminal(lead_id: str, api_key: APIKeyDep = "", session: SessionDep = None, background_tasks: BackgroundTasks = None):
    from services.property_terminal_service import get_property_terminal
    from services.underwriter_service import get_or_generate_brief

    result = await get_property_terminal(session, lead_id)

    # Attach brief — fetch from cache or generate async
    try:
        lead_row = (await session.execute(
            text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id}
        )).mappings().first()
        if lead_row:
            brief = await get_or_generate_brief(lead_id, dict(lead_row), session)
            result["brief"] = brief
    except Exception as _exc:
        result["brief"] = None

    return result


@router.get("/api/leads/{lead_id}/tools")
async def get_lead_tools(lead_id: str, api_key: APIKeyDep = "", session: SessionDep = None):
    from services.lead_tools_service import build_lead_tools_payload_for_id

    payload = await build_lead_tools_payload_for_id(session, lead_id)
    if not payload:
        raise HTTPException(status_code=404, detail="Lead not found")
    return payload


@router.get("/api/tools/suburb-opportunity")
async def get_suburb_opportunity(suburb: str, api_key: APIKeyDep = "", session: SessionDep = None):
    from services.lead_tools_service import build_suburb_opportunity_payload

    suburb_name = str(suburb or "").strip()
    if not suburb_name:
        raise HTTPException(status_code=400, detail="suburb is required")
    return await build_suburb_opportunity_payload(session, suburb_name)


@router.get("/api/leads/{lead_id}/timeline")
async def get_lead_timeline_endpoint(lead_id: str, limit: int = 50, api_key: APIKeyDep = "", session: SessionDep = None):
    from services.timeline_service import get_lead_timeline
    events = await get_lead_timeline(lead_id, session, limit=limit)
    return {"lead_id": lead_id, "events": events, "count": len(events)}


@router.post("/api/leads/invalidate-briefs")
async def invalidate_lead_briefs(
    body: dict,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Force brief cache eviction for a list of lead IDs or all leads in a suburb."""
    from services.underwriter_service import invalidate_brief
    lead_ids = body.get("lead_ids", [])
    if not lead_ids and body.get("suburb"):
        rows = (await session.execute(
            text("SELECT id FROM leads WHERE LOWER(suburb) = LOWER(:suburb)"),
            {"suburb": body["suburb"]},
        )).mappings().all()
        lead_ids = [r["id"] for r in rows]
    invalidate_lead_read_models(lead_ids)
    return {"invalidated": len(lead_ids), "lead_ids": lead_ids}


@router.get("/api/leads/{lead_id}/tasks")
async def get_lead_tasks(lead_id: str, api_key: APIKeyDep = "", session: SessionDep = None):
    
    task_result = await session.execute(
        text("SELECT * FROM tasks WHERE lead_id = :lead_id ORDER BY CASE WHEN status = 'pending' THEN 0 ELSE 1 END, due_at ASC, created_at DESC"),
        {"lead_id": lead_id},
    )
    rows = task_result.mappings().all()
    return [_task_to_dict(dict(row)) for row in rows]


@router.post("/api/leads/{lead_id}/tasks")
async def create_lead_task(lead_id: str, body: TaskRequest, api_key: APIKeyDep = "", session: SessionDep = None):
    
    (await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first()
    task_id = body.id or hashlib.md5(f"{lead_id}:{body.title}:{body.due_at}:{now_iso()}".encode()).hexdigest()
    due_at = parse_client_datetime(body.due_at)
    now = now_iso()
    await session.execute(
        text("""
        INSERT INTO tasks (
            id, lead_id, title, task_type, action_type, channel, due_at, status, notes, related_report_id,
            approval_status, message_subject, message_preview, rewrite_reason, superseded_by, cadence_name,
            cadence_step, auto_generated, priority_bucket, completed_at, created_at, updated_at
        )
        VALUES (:id, :lead_id, :title, :task_type, :action_type, :channel, :due_at, :status, :notes,
                :related_report_id, :approval_status, :message_subject, :message_preview, :rewrite_reason,
                :superseded_by, :cadence_name, :cadence_step, :auto_generated, :priority_bucket,
                :completed_at, :created_at, :updated_at)
        """),
        {
            "id": task_id,
            "lead_id": lead_id,
            "title": body.title,
            "task_type": body.task_type,
            "action_type": body.task_type,
            "channel": body.channel or "",
            "due_at": due_at,
            "status": body.status,
            "notes": body.notes or "",
            "related_report_id": body.related_report_id or "",
            "approval_status": "pending" if (body.channel or "") in {"sms", "email"} else "not_required",
            "message_subject": "",
            "message_preview": "",
            "rewrite_reason": "Manual operator task",
            "superseded_by": "",
            "cadence_name": "",
            "cadence_step": 0,
            "auto_generated": 0,
            "priority_bucket": "send_now" if (body.channel or "") in {"sms", "email"} else "follow_up",
            "completed_at": None,
            "created_at": now,
            "updated_at": now,
        },
    )
    updated_lead = await _append_activity_and_commit(
        session,
        lead_id,
        "task_scheduled",
        f"{body.title} scheduled for {due_at}",
        body.channel,
        body.task_type.replace("_", " ").title(),
    )
    task_row = (await session.execute(text("SELECT * FROM tasks WHERE id = :id"), {"id": task_id})).mappings().first()

    return {"status": "ok", "task": _task_to_dict(dict(task_row) if task_row else {}), "lead": updated_lead}


@router.get("/api/leads/{lead_id}/appointments")
async def get_lead_appointments(lead_id: str, api_key: APIKeyDep = "", session: SessionDep = None):
    
    appt_result = await session.execute(
        text("SELECT * FROM appointments WHERE lead_id = :lead_id ORDER BY starts_at ASC, created_at DESC"),
        {"lead_id": lead_id},
    )
    rows = appt_result.mappings().all()
    return [_appointment_to_dict(dict(row)) for row in rows]


@router.post("/api/leads/{lead_id}/appointments")
async def create_lead_appointment(lead_id: str, body: AppointmentRequest, api_key: APIKeyDep = "", session: SessionDep = None):
    
    (await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first()
    appointment_id = body.id or hashlib.md5(f"{lead_id}:{body.title}:{body.starts_at}:{now_iso()}".encode()).hexdigest()
    starts_at = parse_client_datetime(body.starts_at)
    now = now_iso()
    await session.execute(
        text("""
        INSERT INTO appointments (id, lead_id, title, starts_at, status, location, notes, cadence_name, auto_generated, created_at, updated_at)
        VALUES (:id, :lead_id, :title, :starts_at, :status, :location, :notes, :cadence_name, :auto_generated, :created_at, :updated_at)
        """),
        {
            "id": appointment_id,
            "lead_id": lead_id,
            "title": body.title,
            "starts_at": starts_at,
            "status": body.status,
            "location": body.location or "",
            "notes": body.notes or "",
            "cadence_name": "",
            "auto_generated": 0,
            "created_at": now,
            "updated_at": now,
        },
    )
    updated_lead = await _append_activity_and_commit(
        session,
        lead_id,
        "appointment_booked",
        f"{body.title} booked for {starts_at}",
        "appointment",
        body.title,
    )
    appointment = (await session.execute(text("SELECT * FROM appointments WHERE id = :id"), {"id": appointment_id})).mappings().first()
    await refresh_hermes_for_lead(session, lead_id, actor="appointment_create")
    
    return {"status": "ok", "appointment": _appointment_to_dict(appointment), "lead": updated_lead}


@router.post("/api/leads/{lead_id}/outcome")
async def apply_lead_outcome(lead_id: str, body: LeadOutcomeRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    from services.metrics_service import build_call_log_row, insert_call_log_row

    lead_row = (await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first()
    lead = _hydrate_lead(lead_row)
    outcome = _normalize_token(body.outcome).replace(" ", "_")
    if not outcome:
        
        raise HTTPException(status_code=400, detail="Outcome is required")

    task = None
    if body.task_id:
        task = (await session.execute(text("SELECT * FROM tasks WHERE id = :task_id AND lead_id = :lead_id"), {"task_id": body.task_id, "lead_id": lead_id})).mappings().first()
        if task and task["status"] == "pending":
            await session.execute(
                text("UPDATE tasks SET status = 'completed', completed_at = :completed_at, updated_at = :updated_at WHERE id = :id"),
                {"completed_at": now_iso(), "updated_at": now_iso(), "id": body.task_id},
            )

    cancelled = 0
    created = 0
    now = now_iso()
    next_action_due = None
    appointment_at_for_validation = None
    note = (body.note or "").strip() or outcome.replace("_", " ").title()
    lead["last_outcome"] = outcome
    lead["last_outcome_at"] = now
    if body.owner_verified is not None:
        lead["owner_verified"] = body.owner_verified
    if body.preferred_channel:
        lead["preferred_channel"] = body.preferred_channel
    if body.objection_reason:
        lead["objection_reason"] = body.objection_reason

    status_target = lead.get("status") or "captured"
    if outcome == "no_answer":
        status_target = "outreach_ready" if status_target in {"captured", "qualified"} else status_target
        lead["queue_bucket"] = "active"
        lead["cadence_name"] = lead.get("cadence_name") or ("hot_seller_10_day" if _infer_strike_zone(lead) == "primary" else "warm_seller_30_day")
        task_step = int(task["cadence_step"]) if task else 0
        if _lead_has_phone(lead) and task_step in {1, 4}:
            bundle = _message_bundle(lead, "missed_call_sms")
            await _schedule_task(
                session,
                lead,
                title="Missed-call SMS",
                due_at=now_sydney() + datetime.timedelta(minutes=5),
                task_type="sms",
                channel="sms",
                cadence_name=task["cadence_name"] if task else lead["cadence_name"],
                cadence_step=task_step,
                priority_bucket="send_now",
                message_preview=bundle["body"],
                rewrite_reason="No answer outcome triggered a short SMS follow-up",
            )
            created += 1
        if _lead_has_email(lead) and task_step in {3, 6}:
            email_key = "step_back_email" if task_step == 6 else "market_email"
            bundle = _message_bundle(lead, email_key)
            await _schedule_task(
                session,
                lead,
                title="Market update email" if task_step != 6 else "Step-back email",
                due_at=now_sydney() + datetime.timedelta(minutes=20),
                task_type="email",
                channel="email",
                cadence_name=task["cadence_name"] if task else lead["cadence_name"],
                cadence_step=task_step,
                priority_bucket="send_now",
                message_subject=bundle["subject"],
                message_preview=bundle["body"],
                rewrite_reason="No answer outcome triggered a value-led email",
            )
            created += 1
    elif outcome in {"send_info", "question"}:
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "contacted"
        lead["queue_bucket"] = "callback_due"
        lead["cadence_name"] = "callback_plan"
        callback_at = _parse_iso_datetime(parse_client_datetime(body.callback_at)) if body.callback_at else _next_business_slot(2, 11, 0)
        next_action_due = callback_at.isoformat()
        channel = lead.get("preferred_channel") or _default_preferred_channel(lead)
        bundle = _message_bundle(lead, "question_reply" if outcome == "question" else "market_email")
        if channel in {"call", "enrichment"}:
            channel = "email" if _lead_has_email(lead) else "sms" if _lead_has_phone(lead) else "enrichment"
        if channel == "email" and _lead_has_email(lead):
            await _schedule_task(
                session,
                lead,
                title="Answer / send requested info",
                due_at=now_sydney() + datetime.timedelta(minutes=5),
                task_type="email",
                channel="email",
                cadence_name="callback_plan",
                cadence_step=1,
                priority_bucket="send_now",
                message_subject=bundle["subject"],
                message_preview=bundle["body"],
                rewrite_reason="Owner requested information first",
            )
            created += 1
        elif channel == "sms" and _lead_has_phone(lead):
            await _schedule_task(
                session,
                lead,
                title="Answer / send requested info",
                due_at=now_sydney() + datetime.timedelta(minutes=5),
                task_type="sms",
                channel="sms",
                cadence_name="callback_plan",
                cadence_step=1,
                priority_bucket="send_now",
                message_preview=bundle["body"],
                rewrite_reason="Owner requested information first",
            )
            created += 1
        created += await _schedule_callback_cadence(session, lead, callback_at, now_sydney() + datetime.timedelta(days=60), "Callback after info send")
        lead["do_not_contact_until"] = callback_at.isoformat()
    elif outcome in {"not_now", "call_back"}:
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "contacted"
        lead["queue_bucket"] = "callback_due"
        lead["cadence_name"] = "callback_plan"
        callback_at = _parse_iso_datetime(parse_client_datetime(body.callback_at)) if body.callback_at else _next_business_slot(21, 10, 0)
        lead["do_not_contact_until"] = callback_at.isoformat()
        next_action_due = callback_at.isoformat()
        created += await _schedule_callback_cadence(session, lead, callback_at, now_sydney() + datetime.timedelta(days=60), "Owner asked for later follow-up")
    elif outcome in {"wrong_person", "not_me", "wrong_number"}:
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "qualified"
        lead["queue_bucket"] = "enrichment"
        lead["cadence_name"] = "enrichment_queue"
        lead["owner_verified"] = False
        lead["contact_role"] = "non_owner" if outcome != "wrong_number" else "bad_number"
        created += await _schedule_enrichment_task(session, lead, _next_business_slot(1, 10, 15), "Outcome requires contact enrichment")
    elif outcome == "soft_no":
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "contacted"
        lead["queue_bucket"] = "nurture"
        lead["cadence_name"] = "monthly_nurture"
        lead["do_not_contact_until"] = _next_business_slot(30, 11, 0).isoformat()
        created += await _schedule_nurture_cadence(session, lead, now_sydney() + datetime.timedelta(days=60))
    elif outcome == "hard_no":
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "dropped"
        lead["queue_bucket"] = "suppressed"
        lead["cadence_name"] = ""
        lead["do_not_contact_until"] = (now_sydney() + datetime.timedelta(days=3650)).isoformat()
    elif outcome == "booked_appraisal":
        cancelled = await _supersede_auto_tasks(session, lead_id, outcome, preserve_task_id=body.task_id)
        status_target = "appt_booked"
        lead["queue_bucket"] = "booked"
        lead["cadence_name"] = "booked_appraisal"
        appointment_at = _parse_iso_datetime(parse_client_datetime(body.appointment_at)) if body.appointment_at else _next_business_slot(1, 16, 0)
        appointment_at_for_validation = appointment_at.isoformat()
        next_action_due = appointment_at.isoformat()
        created += await _schedule_booked_followthrough(session, lead, appointment_at, body.appointment_location or "Phone / on-site")
    else:
        status_target = "contacted"

    assert_status_transition_allowed(
        lead,
        status_target,
        source="lead_outcome",
        appointment_at=appointment_at_for_validation,
    )

    note_history = _append_stage_note(lead.get("stage_note_history"), note, status_target, "workflow", outcome.replace("_", " ").title())
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry("workflow_outcome", note, status_target, "workflow", outcome.replace("_", " ").title()),
    )
    lead["touches_14d"] = _recent_touch_count(activity_log, 14)
    lead["touches_30d"] = _recent_touch_count(activity_log, 30)
    await session.execute(
        text("""
        UPDATE leads
        SET status = :status, stage_note = :stage_note, stage_note_history = :stage_note_history,
            activity_log = :activity_log, last_contacted_at = :last_contacted_at,
            last_outbound_at = CASE WHEN :outcome_check IN ('no_answer', 'send_info', 'question') THEN :outbound_now ELSE last_outbound_at END,
            queue_bucket = :queue_bucket, lead_archetype = :lead_archetype, contactability_status = :contactability_status,
            owner_verified = :owner_verified, contact_role = :contact_role,
            cadence_name = :cadence_name, cadence_step = :cadence_step, last_outcome = :last_outcome,
            last_outcome_at = :last_outcome_at, objection_reason = :objection_reason,
            preferred_channel = :preferred_channel, strike_zone = :strike_zone,
            touches_14d = :touches_14d, touches_30d = :touches_30d,
            do_not_contact_until = :do_not_contact_until, updated_at = :updated_at
        WHERE id = :id
        """),
        {
            "status": status_target,
            "stage_note": note,
            "stage_note_history": json.dumps(note_history),
            "activity_log": json.dumps(activity_log),
            "last_contacted_at": now,
            "outcome_check": outcome,
            "outbound_now": now,
            "queue_bucket": lead.get("queue_bucket") or "",
            "lead_archetype": lead.get("lead_archetype") or _infer_lead_archetype(lead),
            "contactability_status": lead.get("contactability_status") or _infer_contactability_status(lead),
            "owner_verified": 1 if lead.get("owner_verified") else 0,
            "contact_role": lead.get("contact_role") or "",
            "cadence_name": lead.get("cadence_name") or "",
            "cadence_step": int(lead.get("cadence_step") or 0),
            "last_outcome": outcome,
            "last_outcome_at": now,
            "objection_reason": lead.get("objection_reason") or "",
            "preferred_channel": lead.get("preferred_channel") or _default_preferred_channel(lead),
            "strike_zone": lead.get("strike_zone") or _infer_strike_zone(lead),
            "touches_14d": int(lead.get("touches_14d") or 0),
            "touches_30d": int(lead.get("touches_30d") or 0),
            "do_not_contact_until": lead.get("do_not_contact_until"),
            "updated_at": now,
            "id": lead_id,
        },
    )
    await insert_call_log_row(
        session,
        build_call_log_row(
            lead_id=lead_id,
            lead_address=lead.get("address", ""),
            outcome=outcome,
            note=note,
            user_id=body.user_id,
            timestamp=now,
            next_action_due=next_action_due,
            provider="workflow",
            direction="outbound",
            raw_payload=json.dumps(
                {
                    "task_id": body.task_id,
                    "callback_at": body.callback_at,
                    "appointment_at": body.appointment_at,
                    "appointment_location": body.appointment_location,
                    "preferred_channel": body.preferred_channel,
                }
            ),
        ),
    )
    await _refresh_lead_next_action(session, lead_id)
    await session.commit()
    await refresh_hermes_for_lead(session, lead_id, actor="lead_outcome")
    updated_lead = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    rewrite = {
        "reason": f"Calendar updated because outcome changed to {outcome.replace('_', ' ')}",
        "cancelled_tasks": cancelled,
        "created_tasks": created,
        "next_action_at": updated_lead.get("next_action_at"),
        "next_action_title": updated_lead.get("next_action_title"),
    }
    
    return {"status": "ok", "lead": updated_lead, "rewrite": rewrite}


@router.post("/api/leads/{lead_id}/advance")
async def advance_lead(lead_id: str, body: LeadAdvanceRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    current_status = lead.get("status") or "captured"
    target_status = body.status.strip()
    _validate_next_status(current_status, target_status)
    assert_status_transition_allowed(
        lead,
        target_status,
        source="lead_advance",
    )
    note_value = (body.note or "").strip()
    note_history = lead.get("stage_note_history")
    activity_log = lead.get("activity_log")
    if note_value:
        note_history = _append_stage_note(note_history, note_value, target_status)
        activity_log = _append_activity(activity_log, _build_activity_entry("note", note_value, target_status, "crm_note"))
    elif target_status != current_status:
        activity_log = _append_activity(
            activity_log,
            _build_activity_entry("status_change", f"Status changed from {current_status} to {target_status}", target_status, "status"),
        )
    now = now_iso()
    await session.execute(
        text("""
        UPDATE leads
        SET status = :status, stage_note = CASE WHEN :note_check != '' THEN :stage_note ELSE stage_note END,
            stage_note_history = :stage_note_history, activity_log = :activity_log,
            last_contacted_at = COALESCE(:last_contacted_at, last_contacted_at),
            updated_at = :updated_at
        WHERE id = :id
        """),
        {
            "status": target_status,
            "note_check": note_value,
            "stage_note": note_value,
            "stage_note_history": json.dumps(note_history or []),
            "activity_log": json.dumps(activity_log or []),
            "last_contacted_at": body.last_contacted_at or (now if target_status == "contacted" else None),
            "updated_at": now,
            "id": lead_id,
        },
    )
    await session.commit()
    await refresh_hermes_for_lead(session, lead_id, actor="lead_advance")
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    
    return {"status": "ok", "from": current_status, "to": target_status, "lead": updated}


@router.post("/api/leads/{lead_id}/generate_outreach")
async def generate_outreach(lead_id: str, body: OutreachPackRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    
    return {"lead_id": lead_id, "status": lead.get("status", "captured"), "pack": _build_outreach_pack(lead, body.tone)}


def _normalize_source_tag(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    if normalized in {"doorknock", "door_knock", "door-knock", "door_knocking"}:
        return "door_knock"
    if not normalized:
        return "rp_data"
    return normalized


def _source_trigger_type(source_tag: str) -> str:
    tag = _normalize_source_tag(source_tag)
    if tag == "rp_data":
        return "RP Data"
    if tag == "door_knock":
        return "Door Knock"
    if tag == "builder":
        return "Builder"
    return tag.replace("_", " ").title() or "RP Data"


def _source_queue_bucket(source_tag: str) -> str:
    return "door_knock" if _normalize_source_tag(source_tag) == "door_knock" else "active"


def _normalize_address(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]+", " ", str(value or "").strip().lower())).strip()


def _decode_json_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text_value = value.strip()
        if not text_value:
            return []
        try:
            parsed = json.loads(text_value)
            return parsed if isinstance(parsed, list) else [parsed]
        except Exception:
            return [text_value]
    return [value]


def _merge_contacts(existing_contacts: list[dict[str, Any]], contact_name: str, phone: str, email: str) -> list[dict[str, Any]]:
    merged = [dict(item) for item in existing_contacts if isinstance(item, dict)]
    clean_name = str(contact_name or "").strip()
    clean_phone = _normalize_phone(phone or "")
    clean_email = str(email or "").strip().lower()
    if not clean_name and not clean_phone and not clean_email:
        return merged

    for item in merged:
        name_value = str(item.get("name") or "").strip().lower()
        phone_value = _normalize_phone(item.get("phone") or "")
        email_value = str(item.get("email") or "").strip().lower()
        if (clean_phone and phone_value and clean_phone == phone_value) or (clean_email and email_value and clean_email == email_value):
            if clean_name and not item.get("name"):
                item["name"] = clean_name
            if clean_phone and not item.get("phone"):
                item["phone"] = clean_phone
            if clean_email and not item.get("email"):
                item["email"] = clean_email
            return merged
        if clean_name and name_value and clean_name.lower() == name_value and (clean_phone == phone_value or clean_email == email_value):
            return merged

    merged.append({
        "name": clean_name or "Contact",
        "phone": clean_phone,
        "email": clean_email,
        "gender": "unknown",
    })
    return merged


def _prioritize_contact(contacts: list[dict[str, Any]], contact_name: str, phone: str, email: str) -> list[dict[str, Any]]:
    clean_name = str(contact_name or "").strip().lower()
    clean_phone = _normalize_phone(phone or "")
    clean_email = str(email or "").strip().lower()
    if not contacts:
        return contacts
    for idx, item in enumerate(contacts):
        if not isinstance(item, dict):
            continue
        name_value = str(item.get("name") or "").strip().lower()
        phone_value = _normalize_phone(item.get("phone") or "")
        email_value = str(item.get("email") or "").strip().lower()
        matches = (
            (clean_phone and phone_value and clean_phone == phone_value)
            or (clean_email and email_value and clean_email == email_value)
            or (clean_name and name_value and clean_name == name_value)
        )
        if matches:
            if idx == 0:
                return contacts
            reordered = list(contacts)
            reordered.insert(0, reordered.pop(idx))
            return reordered
    return contacts


def _append_notes_text(existing_notes: Any, new_note: str) -> str:
    note_text = str(new_note or "").strip()
    if not note_text:
        return str(existing_notes or "").strip()
    current = str(existing_notes or "").strip()
    if not current:
        return note_text
    if note_text.lower() in current.lower():
        return current
    return f"{current}\n{note_text}"


def _invalidate_analytics_cache() -> None:
    try:
        from api.routes import analytics as analytics_routes

        cache_obj = getattr(analytics_routes, "_cache", None)
        if isinstance(cache_obj, dict):
            cache_obj.pop("analytics", None)
    except Exception:
        pass


@router.post("/api/leads/manual")
async def create_manual_lead(body: ManualLeadRequest, db: AsyncSession = Depends(get_session), api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    scenario = body.scenario or body.notes or f"Manual lead added for {body.address}."
    images = _dedupe_text_list([body.main_image, *body.property_images])
    lead_id = get_deterministic_id(body.address)
    now = now_iso()
    default_source_tag = _normalize_source_tag(body.source or "rp_data")
    manual_source_tags = _dedupe_text_list([_normalize_source_tag(tag) for tag in (body.source_tags or [default_source_tag]) if str(tag or "").strip()])
    if not manual_source_tags:
        manual_source_tags = [default_source_tag]
    inferred_manual_tags = _extract_import_tags("", body.owner_type, body.notes or "")
    manual_source_tags = _dedupe_text_list([*manual_source_tags, *inferred_manual_tags])
    primary_source_tag = manual_source_tags[0]
    queue_bucket_default = _source_queue_bucket(primary_source_tag)

    existing = (await db.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()
    owner_name = (body.owner_name or "").strip()
    note_text = (body.notes or scenario).strip()
    event_type = f"{primary_source_tag}_manual"
    activity_entry = _build_activity_entry(event_type, note_text or scenario, "captured", primary_source_tag)

    if existing:
        current = dict(existing)
        emails = _dedupe_text_list([*(_decode_json_list(current.get("contact_emails"))), *(body.contact_emails or [])])
        phones = _dedupe_by_phone([*(_decode_json_list(current.get("contact_phones"))), *(body.contact_phones or [])])
        contacts = _decode_json_list(current.get("contacts"))
        for candidate in body.contacts or []:
            if not isinstance(candidate, dict):
                continue
            contacts = _merge_contacts(
                contacts,
                str(candidate.get("name") or ""),
                str(candidate.get("phone") or ""),
                str(candidate.get("email") or ""),
            )
        contacts = _merge_contacts(
            contacts,
            owner_name,
            phones[0] if phones else "",
            emails[0] if emails else "",
        )
        contacts = _prioritize_contact(contacts, owner_name, phones[0] if phones else "", emails[0] if emails else "")
        tags = _dedupe_text_list([*(_decode_json_list(current.get("source_tags"))), *manual_source_tags])
        key_details = _dedupe_text_list([*(_decode_json_list(current.get("key_details"))), note_text])
        activity_log = _append_activity(_decode_json_list(current.get("activity_log")), activity_entry)
        stage_history = _append_stage_note(
            _decode_json_list(current.get("stage_note_history")),
            note_text or scenario,
            str(current.get("status") or "captured"),
            primary_source_tag,
        )

        await db.execute(
            text(
                """
                UPDATE leads
                SET owner_name = :owner_name,
                    owner_type = CASE WHEN :owner_type != '' THEN :owner_type ELSE owner_type END,
                    contact_emails = :emails,
                    contact_phones = :phones,
                    contacts = :contacts,
                    source_tags = :source_tags,
                    key_details = :key_details,
                    stage_note = :stage_note,
                    notes = :notes,
                    stage_note_history = :stage_note_history,
                    activity_log = :activity_log,
                    follow_up_due_at = COALESCE(:follow_up_due_at, follow_up_due_at),
                    queue_bucket = CASE
                        WHEN COALESCE(queue_bucket, '') IN ('', 'captured', 'new', 'door_knock') THEN :queue_bucket
                        ELSE queue_bucket
                    END,
                    scenario = CASE WHEN :scenario != '' THEN :scenario ELSE scenario END,
                    last_activity_type = :last_activity_type,
                    updated_at = :updated_at
                WHERE id = :id
                """
            ),
            {
                "owner_name": owner_name or current.get("owner_name") or "Owner record pending",
                "owner_type": (body.owner_type or "").strip(),
                "emails": json.dumps(emails),
                "phones": json.dumps(phones),
                "contacts": json.dumps(contacts),
                "source_tags": json.dumps(tags),
                "key_details": json.dumps(key_details),
                "stage_note": note_text or scenario,
                "notes": _append_notes_text(current.get("notes"), note_text or scenario),
                "stage_note_history": json.dumps(stage_history),
                "activity_log": json.dumps(activity_log),
                "follow_up_due_at": body.follow_up_due_at,
                "queue_bucket": queue_bucket_default,
                "scenario": scenario,
                "last_activity_type": body.last_activity_type or event_type,
                "updated_at": now,
                "id": lead_id,
            },
        )
        status = "merged"
    else:
        activity_log = [activity_entry]
        stage_history = _append_stage_note([], note_text or scenario, "captured", primary_source_tag)
        contacts = list(body.contacts or [])
        contacts = _merge_contacts(
            contacts,
            owner_name,
            (body.contact_phones or [""])[0] if body.contact_phones else "",
            (body.contact_emails or [""])[0] if body.contact_emails else "",
        )
        contacts = _prioritize_contact(
            contacts,
            owner_name,
            (body.contact_phones or [""])[0] if body.contact_phones else "",
            (body.contact_emails or [""])[0] if body.contact_emails else "",
        )
        trigger_type_value = (body.trigger_type or "").strip()
        if trigger_type_value.lower() in {"", "manual", "manual_entry"}:
            trigger_type_value = _source_trigger_type(primary_source_tag)
        new_lead = SQLLead(
            id=lead_id,
            address=body.address,
            suburb=body.suburb,
            postcode=body.postcode,
            owner_name=owner_name or "Owner record pending",
            owner_type=(body.owner_type or "").strip() or None,
            trigger_type=trigger_type_value,
            record_type="manual_entry",
            heat_score=55,
            scenario=scenario,
            strategic_value="Operator-created opportunity",
            contact_status="unreviewed",
            confidence_score=60,
            contacts=contacts,
            contact_emails=_dedupe_text_list(body.contact_emails),
            contact_phones=_dedupe_by_phone(body.contact_phones),
            lat=body.lat,
            lng=body.lng,
            date_found=now,
            key_details=[note_text or scenario],
            main_image=images[0] if images else body.main_image,
            property_images=images,
            description_deep=body.description_deep or scenario,
            conversion_strategy="Confirm details and establish next action.",
            summary_points=[scenario],
            horizon="review_now",
            last_checked=format_sydney(),
            exhaustive_summary=body.description_deep or scenario,
            likely_scenario=body.lifecycle_stage or "manual_entry",
            strategic_why="Manual operator entry",
            status=body.status or "captured",
            source_tags=manual_source_tags,
            source_evidence=body.source_evidence,
            linked_files=body.linked_files,
            lifecycle_stage=body.lifecycle_stage,
            queue_bucket=queue_bucket_default,
            stage_note=note_text or scenario,
            notes=note_text or scenario,
            stage_note_history=stage_history,
            activity_log=activity_log,
            follow_up_due_at=body.follow_up_due_at,
            last_activity_type=body.last_activity_type or event_type,
            created_at=now,
            updated_at=now,
        )
        db.add(new_lead)
        status = "created"

    await db.commit()
    _invalidate_analytics_cache()
    await refresh_hermes_for_lead(session, lead_id, actor="manual_lead_create")
    return {"status": status, "lead_id": lead_id}


@router.post("/api/leads/bulk-create")
async def bulk_create_leads(body: BulkCreateRequest, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_session), api_key: str = Depends(get_api_key)):
    FU_DAYS = {"ASAP": 0, "short_term": 14, "follow_up": 7, "mid_term": 30, "long_term": 90, "6_months": 180}
    now = now_iso()
    # If operator provides a date_added (e.g. door-knock collection date), use it as created_at
    effective_created_at = body.date_added.strip() if body.date_added and body.date_added.strip() else now
    # Validate the date format — fall back to now if invalid
    try:
        import datetime as _dt_check
        _dt_check.datetime.fromisoformat(effective_created_at.replace("Z", "+00:00")[:19])
    except (ValueError, TypeError):
        effective_created_at = now
    reader = csv.DictReader(io.StringIO(body.csv_data))
    created = 0
    merged = 0
    skipped = 0
    enrich_lead_ids: list[str] = []
    for row in reader:
        address = f"{(row.get('house_no') or '').strip()} {(row.get('street') or '').strip()}, {(row.get('suburb') or '').strip()}".strip(", ").strip()
        if not address or address == ",":
            skipped += 1
            continue

        lead_id = get_deterministic_id(address)
        contact_name = (row.get("contact_name") or "").strip()
        phone = (row.get("phone") or "").strip()
        email = (row.get("email") or "").strip()
        notes = (row.get("notes") or "").strip()
        suburb = (row.get("suburb") or "").strip()
        source_raw = (row.get("source") or body.source or "rp_data").strip()
        source_tag = _normalize_source_tag(source_raw)
        source_queue_bucket = _source_queue_bucket(source_tag)
        import_tags = _extract_import_tags(row.get("tags", ""), row.get("owner_type", ""), notes)
        merged_source_tags = _dedupe_text_list([source_tag, *import_tags])
        owner_type_value = (row.get("owner_type") or "").strip()
        if not owner_type_value and "builder" in merged_source_tags:
            owner_type_value = "builder"

        follow_up_due_at = None
        follow_up_when = (row.get("follow_up_when") or "").strip()
        if (row.get("follow_up_needed") or "").strip().lower() in ("true", "1", "yes") and follow_up_when:
            days = FU_DAYS.get(follow_up_when, 7)
            import datetime as _dt
            follow_up_due_at = (_dt.datetime.utcnow() + _dt.timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%S")

        existing = (await db.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()

        if existing:
            current = dict(existing)
            emails = _dedupe_text_list([*(_decode_json_list(current.get("contact_emails"))), *([email] if email else [])])
            phones = _dedupe_by_phone([*(_decode_json_list(current.get("contact_phones"))), *([phone] if phone else [])])
            contacts = _merge_contacts(
                _decode_json_list(current.get("contacts")),
                contact_name, phone, email,
            )
            tags = _dedupe_text_list([*(_decode_json_list(current.get("source_tags"))), *merged_source_tags])
            updated_notes = _append_notes_text(current.get("notes"), notes)
            collected_label = f" (collected {effective_created_at[:10]})" if effective_created_at != now else ""
            activity_entry = _build_activity_entry(f"{source_tag}_csv_import", (notes or f"CSV import update for {address}") + collected_label, current.get("status") or "captured", source_tag)
            activity_log = _append_activity(_decode_json_list(current.get("activity_log")), activity_entry)

            await db.execute(
                text("""
                    UPDATE leads
                    SET owner_name = CASE WHEN :owner_name != '' THEN :owner_name ELSE owner_name END,
                        owner_type = CASE WHEN :owner_type != '' THEN :owner_type ELSE owner_type END,
                        contact_emails = :emails,
                        contact_phones = :phones,
                        contacts = :contacts,
                        source_tags = :source_tags,
                        notes = :notes,
                        activity_log = :activity_log,
                        queue_bucket = CASE
                            WHEN COALESCE(queue_bucket, '') IN ('', 'captured', 'new', 'door_knock') THEN :queue_bucket
                            ELSE queue_bucket
                        END,
                        follow_up_due_at = COALESCE(:follow_up_due_at, follow_up_due_at),
                        updated_at = :updated_at
                    WHERE id = :id
                """),
                {
                    "owner_name": contact_name,
                    "owner_type": owner_type_value,
                    "emails": json.dumps(emails),
                    "phones": json.dumps(phones),
                    "contacts": json.dumps(contacts),
                    "source_tags": json.dumps(tags),
                    "notes": updated_notes,
                    "activity_log": json.dumps(activity_log),
                    "queue_bucket": source_queue_bucket,
                    "follow_up_due_at": follow_up_due_at,
                    "updated_at": now,
                    "id": lead_id,
                },
            )
            merged += 1
            enrich_lead_ids.append(lead_id)
        else:
            contacts = _merge_contacts([], contact_name, phone, email)
            lead = SQLLead(
                id=lead_id,
                address=address,
                suburb=suburb,
                owner_name=contact_name or "Owner record pending",
                owner_type=owner_type_value or None,
                status=(row.get("status") or "captured").strip() or "captured",
                contact_phones=[phone] if phone else [],
                contact_emails=[email] if email else [],
                contacts=contacts,
                notes=notes,
                source_tags=merged_source_tags,
                follow_up_due_at=follow_up_due_at,
                queue_bucket=source_queue_bucket,
                record_type="manual_entry",
                heat_score=55,
                confidence_score=60,
                trigger_type=_source_trigger_type(source_tag),
                lifecycle_stage="manual_entry",
                created_at=effective_created_at,
                updated_at=now,
            )
            db.add(lead)
            created += 1
            enrich_lead_ids.append(lead_id)

    await db.commit()
    _invalidate_analytics_cache()

    # Background-enrich imported leads with Cotality/Domain data
    if enrich_lead_ids:
        async def _bg_enrich_leads(lead_ids: list[str]):
            try:
                from services.enrichment_service import enrichment_service
                for lid in lead_ids:
                    try:
                        await enrichment_service.enrich_lead(lid)
                    except Exception as exc:
                        logger.warning("Background enrichment failed for %s: %s", lid, exc)
            except Exception as exc:
                logger.warning("Background enrichment batch error: %s", exc)
        background_tasks.add_task(_bg_enrich_leads, enrich_lead_ids)

    return {"created": created, "merged": merged, "skipped": skipped}


@router.post("/api/sold-events")
async def create_sold_event(body: SoldEventRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    try:
        event_id = hashlib.md5(
            f"{body.address}:{body.sale_date or ''}:{body.sale_price or ''}:{body.source_url or body.source_name}".encode()
        ).hexdigest()
        matched_lead_ids = await _find_matching_leads_for_sold_event(session, body.address, body.suburb)
        match_reason = "exact_address_match" if matched_lead_ids else ""
        now = now_iso()
        await session.execute(
            text("""
            INSERT OR REPLACE INTO sold_events (
                id, address, suburb, postcode, sale_date, sale_price, lat, lng, source_name, source_url,
                match_reason, matched_lead_ids, created_at, updated_at
            ) VALUES (:id, :address, :suburb, :postcode, :sale_date, :sale_price, :lat, :lng, :source_name, :source_url,
                :match_reason, :matched_lead_ids, COALESCE((SELECT created_at FROM sold_events WHERE id = :id2), :created_at), :updated_at)
            """),
            {
                "id": event_id,
                "address": body.address,
                "suburb": body.suburb,
                "postcode": body.postcode,
                "sale_date": body.sale_date or "",
                "sale_price": body.sale_price or "",
                "lat": body.lat,
                "lng": body.lng,
                "source_name": body.source_name,
                "source_url": body.source_url,
                "match_reason": match_reason,
                "matched_lead_ids": json.dumps(matched_lead_ids),
                "id2": event_id,
                "created_at": now,
                "updated_at": now,
            },
        )
        for lead_id in matched_lead_ids:
            content = f"Recent sold event detected for {body.address}. Source: {body.source_url or body.source_name or 'manual'}"
            await session.execute(
                text("INSERT INTO notes (lead_id, note_type, content, created_at) VALUES (:lead_id, :note_type, :content, :created_at)"),
                {"lead_id": lead_id, "note_type": "sold_event", "content": content, "created_at": now},
            )
            lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
            activity_log = _append_activity(
                lead.get("activity_log"),
                _build_activity_entry("sold_event", content, lead.get("status"), "sold_event", body.source_name, body.source_url),
            )
            await session.execute(
                text("UPDATE leads SET activity_log = :activity_log, updated_at = :updated_at WHERE id = :id"),
                {"activity_log": json.dumps(activity_log), "updated_at": now, "id": lead_id},
            )
        await session.commit()
        event = (await session.execute(text("SELECT * FROM sold_events WHERE id = :id"), {"id": event_id})).mappings().first()
        return {"status": "created", "event": _sold_event_to_dict(event)}
    finally:
        pass

@router.get("/api/sold-events")
async def get_sold_events(limit: int = 200, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    rows = (await session.execute(text("SELECT * FROM sold_events ORDER BY created_at DESC LIMIT :limit"), {"limit": limit})).mappings().all()

    return [_sold_event_to_dict(row) for row in rows]


@router.get("/api/sold-events/recent")
async def get_recent_sold_events(limit: int = 50, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):

    rows = (await session.execute(text("SELECT * FROM sold_events ORDER BY created_at DESC LIMIT :limit"), {"limit": limit})).mappings().all()
    
    return [_sold_event_to_dict(row) for row in rows]


@router.post("/api/leads/{lead_id}/activity")
async def log_lead_activity(lead_id: str, body: ActivityLogRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry(body.activity_type, body.note, body.status or lead.get("status"), body.channel, body.subject, body.recipient),
    )
    last_inbound_at = lead.get("last_inbound_at")
    last_outbound_at = lead.get("last_outbound_at")
    if "received" in body.activity_type or "inbound" in body.activity_type:
        last_inbound_at = now_iso()
    if "sent" in body.activity_type or "opened" in body.activity_type or body.channel in {"email", "sms", "zoom"}:
        last_outbound_at = now_iso()
    await session.execute(
        text("UPDATE leads SET activity_log = :activity_log, last_inbound_at = :last_inbound_at, last_outbound_at = :last_outbound_at, updated_at = :updated_at WHERE id = :id"),
        {"activity_log": json.dumps(activity_log), "last_inbound_at": last_inbound_at, "last_outbound_at": last_outbound_at, "updated_at": now_iso(), "id": lead_id},
    )
    await session.commit()
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    
    return {"status": "ok", "lead": updated}


@router.post("/api/leads/{lead_id}/send-email-account")
async def send_lead_email(lead_id: str, body: SendEmailRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    try:
        await assert_outreach_allowed(session, lead_id, "email", purpose=resolve_outreach_purpose(lead))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    account = (await session.execute(text("SELECT * FROM email_accounts WHERE id = :id"), {"id": body.account_id})).mappings().first()
    if not account:

        raise HTTPException(status_code=404, detail="Email account not found")
    account_data = dict(account)
    try:
        # Use centralized send_email_service (Graph priority)
        await asyncio.to_thread(send_email_service, account_data, body)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Email send failed: {exc}") from exc

    note_history = _append_stage_note(
        lead.get("stage_note_history"),
        body.body,
        lead.get("status") or "captured",
        "email",
        body.subject,
        body.recipient,
    )
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry("email_sent", body.body, lead.get("status"), "email", body.subject, body.recipient),
    )
    await session.execute(
        text("UPDATE leads SET stage_note_history = :stage_note_history, activity_log = :activity_log, last_outbound_at = :last_outbound_at, updated_at = :updated_at WHERE id = :id"),
        {"stage_note_history": json.dumps(note_history), "activity_log": json.dumps(activity_log), "last_outbound_at": now_iso(), "updated_at": now_iso(), "id": lead_id},
    )
    await session.commit()
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    
    return {"status": "sent", "lead": updated}

@router.post("/api/leads/{lead_id}/send-email")
async def send_lead_email_direct(lead_id: str, body: DirectSendEmailRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    """
    Send an email directly using Microsoft Graph (priority) or SMTP_HOST env vars.
    Does not require a saved email account record.
    """
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    try:
        await assert_outreach_allowed(session, lead_id, "email", purpose=resolve_outreach_purpose(lead))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    subject = body.subject or f"Property update — {lead.get('address', '')}"
    
    try:
        # Add tracking pixel + click wrapping
        tracked_body = body.body
        tracking_id = None
        try:
            from services.email_tracking import generate_tracking_id, wrap_email_with_tracking
            tracking_id = generate_tracking_id()
            backend_url = os.getenv("BASE_URL") or os.getenv("RENDER_EXTERNAL_URL") or "http://localhost:8001"
            tracked_body = wrap_email_with_tracking(body.body, tracking_id, backend_url)
            # Record send event
            from models.sql_models import EmailEvent
            send_event = EmailEvent(
                id=str(uuid.uuid4()),
                lead_id=lead_id,
                tracking_id=tracking_id,
                event_type="send",
                created_at=now_iso(),
            )
            session.add(send_event)
        except Exception:
            tracked_body = body.body

        email_req = SendEmailRequest(
            account_id="",
            recipient=body.recipient,
            subject=subject,
            body=tracked_body
        )
        # Use centralized send_email_service (Graph priority)
        await asyncio.to_thread(send_email_service, None, email_req)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Email send failed: {exc}") from exc

    note_history = _append_stage_note(
        lead.get("stage_note_history"), body.body,
        lead.get("status") or "captured", "email", subject, body.recipient,
    )
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry("email_sent", body.body, lead.get("status"), "email", subject, body.recipient),
    )
    await session.execute(
        text("UPDATE leads SET stage_note_history = :stage_note_history, activity_log = :activity_log, last_outbound_at = :last_outbound_at, updated_at = :updated_at WHERE id = :id"),
        {"stage_note_history": json.dumps(note_history), "activity_log": json.dumps(activity_log), "last_outbound_at": now_iso(), "updated_at": now_iso(), "id": lead_id},
    )
    await session.commit()
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())

    return {"status": "sent", "lead": updated}


@router.post("/api/leads/{lead_id}/send-sms")
async def send_lead_sms(lead_id: str, body: DirectSendSMSRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    """Send SMS via Twilio. Falls back to graceful error if not configured."""
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    try:
        await assert_outreach_allowed(session, lead_id, "sms", purpose=resolve_outreach_purpose(lead))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    

    status = "sms_bridge_unavailable"
    provider = ""
    try:
        from services.sms_service import sms_service
        result = await sms_service.send_sms(body.recipient, body.message, lead_id)
        if result.get("ok"):
            status = "queued"
            provider = "twilio"
    except Exception:
        pass

    if status != "queued":
        # Hermes bridge fallback
        try:
            if SMS_BRIDGE_URL:
                async with httpx.AsyncClient(timeout=5.0) as client:
                    await client.post(
                        f"{SMS_BRIDGE_URL}/send",
                        json={"to": body.recipient, "message": body.message},
                    )
                status = "queued"
                provider = "sms_bridge"
        except Exception:
            pass

    if status != "queued":
        return {"status": "sms_bridge_unavailable", "note": "Configure TWILIO_* or SMS_BRIDGE_URL."}

    note_history = _append_stage_note(
        lead.get("stage_note_history"),
        body.message,
        lead.get("status") or "captured",
        "sms",
        f"SMS via {provider}",
        body.recipient,
    )
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry("text_sent", body.message, lead.get("status"), "sms", f"SMS via {provider}", body.recipient),
    )
    await session.execute(
        text("UPDATE leads SET stage_note_history = :stage_note_history, activity_log = :activity_log, last_outbound_at = :last_outbound_at, updated_at = :updated_at WHERE id = :id"),
        {"stage_note_history": json.dumps(note_history), "activity_log": json.dumps(activity_log), "last_outbound_at": now_iso(), "updated_at": now_iso(), "id": lead_id},
    )
    await session.commit()
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())

    return {"status": status, "provider": provider, "lead": updated}


@router.get("/api/leads/{lead_id}/cotality/reports")
async def get_lead_cotality_reports(lead_id: str, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    (await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first()
    rows = (await session.execute(
        text("SELECT id, lead_id, report_type, title, created_at, updated_at FROM cotality_reports WHERE lead_id = :lead_id ORDER BY created_at DESC"),
        {"lead_id": lead_id},
    )).mappings().all()

    return [dict(row) for row in rows]


@router.post("/api/leads/{lead_id}/cotality/generate-report")
async def generate_lead_cotality_report(lead_id: str, body: CotalityReportRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    report = await _generate_cotality_report(session, lead, body.report_type)
    
    return report


@router.post("/api/leads/{lead_id}/send-text")
async def send_lead_text(lead_id: str, body: SendTextRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    try:
        await assert_outreach_allowed(session, lead_id, "sms", purpose=resolve_outreach_purpose(lead))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    account = (await session.execute(text("SELECT * FROM communication_accounts WHERE id = :id"), {"id": body.account_id})).mappings().first()
    if not account:
        
        raise HTTPException(status_code=404, detail="Communication account not found")
    account_data = dict(account)
    if account_data.get("provider") == "zoom":
        if body.dry_run or not _bool_db(account_data.get("send_enabled")):
            transport_result = {
                "status": "dry_run",
                "provider": "zoom",
                "would_post_to": f"{(account_data.get('api_base') or 'https://api.zoom.us/v2').rstrip('/')}/{(account_data.get('send_path') or '/phone/sms/messages').lstrip('/')}",
                "from": account_data.get("from_number"),
                "to": body.recipient,
                "message_preview": body.message[:160],
            }
        else:
            transport_result = _zoom_request(
                account_data,
                "POST",
                account_data.get("send_path") or "/phone/sms/messages",
                {"toMembers": [{"phoneNumber": body.recipient}], "message": body.message},
            )
            if not transport_result.get("ok"):
                
                raise HTTPException(status_code=400, detail=f"Zoom text send failed: {transport_result.get('error', 'Unknown error')}")
    else:
        transport_result = await asyncio.to_thread(_send_http_text, account_data, body.recipient, body.message)
    note_history = _append_stage_note(
        lead.get("stage_note_history"),
        body.message,
        lead.get("status") or "captured",
        account_data.get("provider") or "text",
        f"Text via {account_data.get('label')}",
        body.recipient,
    )
    activity_log = _append_activity(
        lead.get("activity_log"),
        _build_activity_entry(
            "text_dry_run" if body.dry_run or not _bool_db(account_data.get("send_enabled")) else "text_sent",
            body.message,
            lead.get("status"),
            account_data.get("provider"),
            f"Text via {account_data.get('label')}",
            body.recipient,
        ),
    )
    await session.execute(
        text("UPDATE leads SET stage_note_history = :stage_note_history, activity_log = :activity_log, last_outbound_at = :last_outbound_at, updated_at = :updated_at WHERE id = :id"),
        {"stage_note_history": json.dumps(note_history), "activity_log": json.dumps(activity_log), "last_outbound_at": now_iso(), "updated_at": now_iso(), "id": lead_id},
    )
    await session.commit()
    updated = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    
    return {"status": "dry_run" if body.dry_run or not _bool_db(account_data.get("send_enabled")) else "sent", "transport": transport_result, "lead": updated}


@router.post("/api/leads/{lead_id}/generate-pdf-report")
async def generate_pdf_report(lead_id: str, body: CotalityReportRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    (await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first()
    report_row = (await session.execute(text("SELECT * FROM cotality_reports WHERE id = :id"), {"id": body.report_id})).mappings().first() if body.report_id else None
    if report_row:
        report_data = dict(report_row)
        report_data["payload"] = json.loads(report_data.pop("json_payload") or "{}")
        html_content = report_data["html_content"]
        report_type = report_data["report_type"]
        report_id = report_data["id"]
    else:
        lead = _decode_row((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
        report_data = await _generate_cotality_report(session, lead, body.report_type)
        html_content = report_data["html_content"]
        report_type = report_data["report_type"]
        report_id = report_data["id"]
    filename = f"Report_{lead_id}_{report_type}_{hashlib.md5(now_iso().encode()).hexdigest()[:6]}.pdf"
    reports_dir = GENERATED_REPORTS_ROOT
    reports_dir.mkdir(parents=True, exist_ok=True)
    file_path = reports_dir / filename
    await html_to_pdf(html_content, str(file_path))
    file_url = build_public_url(f"/api/forms/download/{filename}")
    payload = dict(report_data.get("payload") or {})
    payload["pdf_artifact"] = {
        "file_name": filename,
        "file_url": file_url,
        "file_path": str(file_path),
        "generated_at": now_iso(),
    }
    await session.execute(
        text("UPDATE cotality_reports SET json_payload = :json_payload, updated_at = :updated_at WHERE id = :id"),
        {"json_payload": json.dumps(payload), "updated_at": now_iso(), "id": report_id},
    )
    await session.commit()
    
    return {
        "status": "success",
        "report_id": report_id,
        "file_name": filename,
        "file_url": file_url,
        "file_path": str(file_path)
    }


@router.post("/api/leads/{lead_id}/report-packs")
async def generate_lead_report_pack(lead_id: str, body: ReportPackRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    
    lead = _hydrate_lead((await session.execute(text('SELECT * FROM leads WHERE id = :id'), {'id': lead_id})).mappings().first())
    manifest = await _generate_report_pack_for_lead(
        session,
        lead,
        include_existing_briefs=body.include_existing_briefs,
        output_root=body.output_root,
    )
    
    return {"status": "success", "manifest": manifest}


@router.get("/api/leads/{lead_id}/strategic-reports")
async def get_lead_strategic_reports(
    lead_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Return previously generated CMA/strategic reports for a lead.
    Reads from leads.stage_note (JSON list, newest last).
    """
    row = (
        await session.execute(text("SELECT stage_note FROM leads WHERE id = :id"), {"id": lead_id})
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Lead not found")
    raw = (row.get("stage_note") or "[]")
    try:
        reports = json.loads(raw)
        if not isinstance(reports, list):
            # Legacy: single CMA stored as object
            reports = [reports] if reports else []
    except Exception:
        reports = []
    # Ensure frontend-expected shape: id, title, narrative fields
    shaped = []
    for r in reports:
        shaped.append({
            "id": r.get("id") or r.get("lead_id") or lead_id,
            "title": r.get("headline") or f"Market Analysis — {r.get('address', lead_id)}",
            "narrative": r.get("market_position") or "",
            "why_now": r.get("why_now") or "",
            "value_range": r.get("value_range") or "",
            "call_script_opening": r.get("call_script_opening") or "",
            "sms_draft": r.get("sms_draft") or "",
            "next_step_cta": r.get("next_step_cta") or "",
            "comparables": r.get("comparables") or [],
            "suburb_stats": r.get("suburb_stats") or {},
            "source": r.get("source") or "unknown",
            "generated_at": r.get("generated_at") or "",
            "cma_link": r.get("cma_link") or "",
        })
    return list(reversed(shaped))  # newest first


@router.post("/api/leads/{lead_id}/generate-strategic-report")
async def generate_lead_strategic_report(
    lead_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Generate a full strategic report (CMA) for a lead using Gemini.

    Data pipeline:
      1. Fetch recent comparable sales from Domain API (last 90 days, same suburb)
      2. Read Cotality xlsx suburb stats (local, free)
      3. Build data-rich prompt and call Gemini
      4. Store result in leads.stage_note (JSON list, last 5 kept)

    Returns the generated report. Works even without Domain API — Cotality data alone
    is sufficient for a substantive market position.
    """
    from services.cma_generator import generate_cma_for_lead_id

    cma = await generate_cma_for_lead_id(session, lead_id)
    return {
        "id": cma.get("id"),
        "title": cma.get("headline") or f"Market Analysis — {cma.get('address', lead_id)}",
        "narrative": cma.get("market_position") or "",
        "why_now": cma.get("why_now") or "",
        "value_range": cma.get("value_range") or "",
        "call_script_opening": cma.get("call_script_opening") or "",
        "sms_draft": cma.get("sms_draft") or "",
        "next_step_cta": cma.get("next_step_cta") or "",
        "comparables": cma.get("comparables") or [],
        "suburb_stats": cma.get("suburb_stats") or {},
        "source": cma.get("source") or "unknown",
        "generated_at": cma.get("generated_at") or "",
        "cma_link": cma.get("cma_link") or "",
    }


@router.post("/api/leads/{lead_id}/generate-cma")
async def generate_lead_cma(
    lead_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Generate a Comparative Market Analysis for a lead using Gemini.
    Stores result in leads.stage_note (JSON). Returns the CMA dict.
    """
    from services.cma_generator import generate_cma_for_lead_id

    cma = await generate_cma_for_lead_id(session, lead_id)
    return {"status": "ok", "cma": cma}

@router.post("/api/leads/{lead_id}/log-call")
async def log_lead_call(lead_id: str, body: LogCallRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    from services.call_brief_service import log_call_attempt
    from services.lead_intelligence_service import sync_lead_intelligence_for_lead
    updated = await log_call_attempt(
        session,
        lead_id,
        body.outcome,
        body.note,
        body.duration_seconds,
        body.user_id,
        body.next_action_due,
        body.recording_url,
    )
    await sync_lead_intelligence_for_lead(session, lead_id)
    await session.commit()
    await refresh_hermes_for_lead(session, lead_id, actor="manual_call_log")
    refreshed = (await attach_deterministic_intelligence(session, [updated["lead"]]))[0]
    return {"status": "ok", "lead": refreshed, "call_log_id": updated["call_log_id"]}


class ImportPreviewRequest(BaseModel):
    csv_data: str

class ConfirmLead(BaseModel):
    row_id: Optional[str] = None
    address: str
    suburb: str
    contact_name: str = ""
    owner_type: str = ""
    email: str = ""
    phone: str = ""
    notes: str = ""
    tags: str = ""
    source: str = "rp_data"
    action: str = "CREATE"
    decision: Optional[str] = None
    existing_id: Optional[str] = None
    target_lead_id: Optional[str] = None
    operator_resolution: Optional[str] = None
    requires_confirmation: bool = False
    safe_for_merge_all: bool = False
    match_confidence: Optional[str] = None
    match_reasons: List[str] = []

class ImportConfirmRequest(BaseModel):
    leads: List[ConfirmLead]


@router.post("/api/leads/import/parse-file")
async def parse_leads_import_file(
    file: UploadFile = File(...),
    api_key: str = Depends(get_api_key),
):
    filename = (file.filename or "").strip()
    suffix = Path(filename).suffix.lower()
    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    if suffix == ".csv":
        try:
            csv_data = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            csv_data = raw_bytes.decode("latin-1", errors="ignore")
        return {"csv_data": csv_data}

    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported for lead import.")

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheet = workbook.active
        row_values: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            normalized_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(value for value in normalized_row):
                row_values.append(normalized_row)

        if not row_values:
            raise HTTPException(status_code=400, detail="The uploaded .xlsx file has no readable rows.")

        output = io.StringIO()
        writer = csv.writer(output)
        for row in row_values:
            writer.writerow(row)
        return {"csv_data": output.getvalue()}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to parse .xlsx file: {exc}")


def _format_source_label(source: str) -> str:
    normalized = _normalize_source_tag(source)
    if normalized == "door_knock":
        return "Door-knocking"
    return normalized.replace("_", " ").title()


def _enrich_import_note(note: str, source: str) -> str:
    base = (note or "").strip()
    source_line = f"Source: {_format_source_label(source)}"
    if source_line.lower() in base.lower():
        return base
    return f"{base}\n{source_line}".strip()


_IMPORT_EMAIL_SENT_RE = re.compile(r"\b(?:greeting\s+)?email\s+sent\b", re.IGNORECASE)


def _split_import_notes(note: str) -> tuple[str, list[str]]:
    text_value = str(note or "").strip()
    if not text_value:
        return "", []

    special_notes: list[str] = []
    cleaned = text_value

    if _IMPORT_EMAIL_SENT_RE.search(cleaned):
        special_notes.append("EMAIL SENT")
        cleaned = _IMPORT_EMAIL_SENT_RE.sub("", cleaned)
        cleaned = re.sub(r"\s*[,;]\s*", "; ", cleaned)
        cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" ;,-")

    return cleaned, special_notes


def _extract_import_tags(raw_tags: str, owner_type: str, notes: str) -> list[str]:
    tokens: list[str] = []
    for part in re.split(r"[;,/|]+", str(raw_tags or "")):
        token = _normalize_source_tag(part)
        if token:
            tokens.append(token)
    owner = _normalize_source_tag(owner_type or "")
    if owner:
        tokens.append(owner)
    blob = f"{owner_type or ''} {notes or ''}".lower()
    if re.search(r"\bland\b", blob):
        tokens.append("land")
    if re.search(r"\bbuilder(s)?\b", blob):
        tokens.append("builder")
    return _dedupe_text_list(tokens)


def _derive_import_source_tags(source_tag: str, owner_type: str, raw_tags: str = "", notes: str = "") -> list[str]:
    normalized_source = _normalize_source_tag(source_tag)
    tags = [normalized_source]
    import_tags = _extract_import_tags(raw_tags, owner_type, notes)
    tags.extend(import_tags)
    if normalized_source == "door_knock" and str(owner_type or "").strip().lower() == "builder":
        tags.append("builder")
    if normalized_source == "door_knock" and re.search(r"\bland\b", str(notes or "").lower()):
        tags.append("land")
    return _dedupe_text_list(tags)


def _default_value_for_column(data_type: str, udt_name: str) -> Any:
    dtype = str(data_type or "").lower()
    udt = str(udt_name or "").lower()
    if dtype in {"smallint", "integer", "bigint", "numeric", "real", "double precision"}:
        return 0
    if dtype == "boolean":
        return False
    if dtype in {"json", "jsonb"} or udt in {"json", "jsonb"}:
        return "[]"
    return ""


_CSV_IMPORT_SYNONYMS: dict[str, str] = {
    "type": "source",
    "lead_type": "source",
    "lead type": "source",
    "street": "street",
    "house_no": "house_no",
    "house_number": "house_no",
    "house no": "house_no",
    "address": "address",
    "suburb": "suburb",
    "postcode": "postcode",
    "contact_name": "contact_name",
    "contact": "contact_name",
    "name": "contact_name",
    "full_name": "contact_name",
    "email": "email",
    "e-mail": "email",
    "phone": "phone",
    "mobile": "phone",
    "notes": "notes",
    "note": "notes",
    "comments": "notes",
    "source": "source",
    "owner_type": "owner_type",
    "owner type": "owner_type",
    "tags": "tags",
    "tag": "tags",
    "categories": "tags",
    "category": "tags",
}
_CSV_IMPORT_KEYS = {
    "street",
    "house_no",
    "address",
    "suburb",
    "postcode",
    "contact_name",
    "email",
    "phone",
    "notes",
    "source",
    "owner_type",
    "tags",
}
_ADDRESS_STREET_RE = re.compile(
    r"^(?P<num>\d+[A-Za-z]?)\s+(?P<street>[A-Za-z][A-Za-z .'-]*(?:road|rd|street|st|place|pl|avenue|ave|court|ct|drive|dr|lane|ln|close|cl|crescent|cres|way|wy|boulevard|blvd))\b\s*(?:-\s*(?P<note>.*))?$",
    re.IGNORECASE,
)
_ADDRESS_SPLIT_RE = re.compile(
    r"^(?P<num>\d+[A-Za-z]?)\s*-\s*(?P<street>[A-Za-z][A-Za-z .'-]*(?:road|rd|street|st|place|pl|avenue|ave|court|ct|drive|dr|lane|ln|close|cl|crescent|cres|way|wy|boulevard|blvd))\b\s*(?:-\s*(?P<note>.*))?$",
    re.IGNORECASE,
)
_HOUSE_ONLY_RE = re.compile(
    r"^(?:house\s*number\s*[-:]?\s*)?(?P<num>\d+[A-Za-z]?)\s*(?:\([^)]*\))?\s*(?:[-:]\s*)?(?P<note>.*)$",
    re.IGNORECASE,
)
_STREET_HEADING_RE = re.compile(
    r"^(?P<street>[A-Za-z][A-Za-z .'-]*(?:road|rd|street|st|place|pl|avenue|ave|court|ct|drive|dr|lane|ln|close|cl|crescent|cres|way|wy|boulevard|blvd))\b(?:\s*-\s*(?P<tail>.*))?$",
    re.IGNORECASE,
)
_STREET_HEADING_RELAXED_RE = re.compile(
    r"^(?P<street>[A-Za-z][A-Za-z .'-]{2,})(?:\s*-\s*(?P<tail>.*))?$",
    re.IGNORECASE,
)
_EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
_PHONE_RE = re.compile(r"(?:\+?61|0)\s*\d(?:[\s-]*\d){7,10}")
_BUILDER_TOKENS = (
    "builder",
    "builders",
    "homes",
    "construction",
    "constructions",
    "developments",
    "developer",
    "pty ltd",
)
_NON_SOURCE_TYPE_TOKENS = {"developer", "land", "residential", "commercial", "industrial", "rural"}


def _clean_import_header(value: str) -> str:
    token = re.sub(r"[\s\-]+", "_", str(value or "").strip().lower())
    return token.strip("_")


def _is_probable_suburb(value: str) -> bool:
    token = str(value or "").strip()
    if not token:
        return False
    if any(char.isdigit() for char in token):
        return False
    if len(token.split()) > 4:
        return False
    lowered = token.lower()
    if any(word in lowered for word in ("no answer", "rent", "owner", "card", "dl", "interested")):
        return False
    return True


def _infer_owner_type(contact_name: str, notes: str) -> str:
    blob = f"{contact_name} {notes}".lower()
    if any(token in blob for token in _BUILDER_TOKENS):
        return "builder"
    return ""


def _is_probable_street_heading(value: str) -> bool:
    token = str(value or "").strip()
    if not token:
        return False
    lowered = token.lower()
    if any(char.isdigit() for char in token):
        return False
    if _EMAIL_RE.search(token) or _PHONE_RE.search(token):
        return False
    if len(token.split()) > 4:
        return False
    blocked = (
        "house number",
        "no answer",
        "rent",
        "owner",
        "not interested",
        "drop off",
        "dl",
        "sold",
        "bought",
        "moved",
        "vacant",
        "for sale",
    )
    if any(word in lowered for word in blocked):
        return False
    return True


def _normalize_import_source(raw_source: str, owner_type: str, notes: str, *, default_source: str = "rp_data") -> str:
    raw_token = _normalize_source_tag((raw_source or "").strip())
    if raw_token in _NON_SOURCE_TYPE_TOKENS:
        source = _normalize_source_tag(default_source)
    else:
        source = _normalize_source_tag(raw_token or default_source)
    if source in {"rp_data", "manual"} and owner_type == "builder":
        return "builder"
    if source == "rp_data" and re.search(r"\bdoor[\s_-]?knock(ing)?\b", (notes or "").lower()):
        return "doorknock"
    return source


def _normalize_import_row(raw: dict[str, Any]) -> dict[str, str]:
    normalized: dict[str, str] = {key: "" for key in _CSV_IMPORT_KEYS}
    for key, value in raw.items():
        canonical = _CSV_IMPORT_SYNONYMS.get(_clean_import_header(key))
        if not canonical:
            continue
        normalized[canonical] = str(value or "").strip()
    return normalized


def _parse_structured_import_rows(raw_data: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(raw_data))
    fieldnames = reader.fieldnames or []
    normalized_headers = {_CSV_IMPORT_SYNONYMS.get(_clean_import_header(name), "") for name in fieldnames}
    normalized_headers.discard("")
    if not normalized_headers.intersection(_CSV_IMPORT_KEYS):
        return []
    return [_normalize_import_row(row) for row in reader]


def _append_note(note: str, addition: str) -> str:
    base = (note or "").strip()
    extra = (addition or "").strip()
    if not extra:
        return base
    if not base:
        return extra
    return f"{base}; {extra}"


def _parse_unstructured_import_rows(raw_data: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    current_street = ""
    current_suburb = ""
    current: Optional[dict[str, str]] = None
    pending_builder_mentions: list[str] = []

    def start_row(house_no: str, street: str, note: str) -> dict[str, str]:
        nonlocal current_street
        row = {
            "street": street.strip(),
            "house_no": house_no.strip(),
            "address": "",
            "suburb": current_suburb,
            "postcode": "",
            "contact_name": "",
            "email": "",
            "phone": "",
            "notes": (note or "").strip(),
            "source": "doorknock",
            "owner_type": "",
        }
        if pending_builder_mentions:
            mention_blob = ", ".join(pending_builder_mentions)
            row["notes"] = _append_note(row["notes"], f"Builder roster: {mention_blob}")
            row["owner_type"] = "builder"
            pending_builder_mentions.clear()
        current_street = row["street"] or current_street
        rows.append(row)
        return row

    lines = [line.strip(" \t\r\n-") for line in (raw_data or "").splitlines()]
    for raw_line in lines:
        line = (raw_line or "").strip()
        if not line:
            continue
        lowered = line.lower()

        if current is None and any(token in lowered for token in _BUILDER_TOKENS):
            pending_builder_mentions.append(line)
            continue

        heading_match = _STREET_HEADING_RE.match(line)
        if heading_match and not re.match(r"^\d", line):
            current_street = heading_match.group("street").strip()
            tail = (heading_match.group("tail") or "").strip()
            if _is_probable_suburb(tail):
                current_suburb = tail.title()
            continue

        relaxed_heading = _STREET_HEADING_RELAXED_RE.match(line)
        if relaxed_heading and not re.match(r"^\d", line) and _is_probable_street_heading(relaxed_heading.group("street") or ""):
            current_street = (relaxed_heading.group("street") or "").strip()
            tail = (relaxed_heading.group("tail") or "").strip()
            if _is_probable_suburb(tail):
                current_suburb = tail.title()
            continue

        if _EMAIL_RE.search(line):
            if current is not None:
                current["email"] = _EMAIL_RE.search(line).group(0).strip().lower()
                owner_type = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
                if owner_type:
                    current["owner_type"] = owner_type
            continue

        phone_match = _PHONE_RE.search(line)
        if phone_match:
            if current is not None:
                current["phone"] = phone_match.group(0).strip()
                without_phone = line.replace(phone_match.group(0), "").strip(" -:")
                if without_phone and not current.get("contact_name"):
                    current["contact_name"] = without_phone
                elif without_phone:
                    current["notes"] = _append_note(current.get("notes", ""), without_phone)
                owner_type = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
                if owner_type:
                    current["owner_type"] = owner_type
            continue

        name_match = re.match(r"^name\s*[-:]\s*(.+)$", line, re.IGNORECASE)
        if name_match and current is not None:
            current["contact_name"] = name_match.group(1).strip()
            owner_type = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
            if owner_type:
                current["owner_type"] = owner_type
            continue

        row_match = _ADDRESS_STREET_RE.match(line) or _ADDRESS_SPLIT_RE.match(line)
        if row_match:
            current = start_row(
                row_match.group("num") or "",
                row_match.group("street") or "",
                row_match.group("note") or "",
            )
            current["owner_type"] = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
            continue

        house_only = _HOUSE_ONLY_RE.match(line)
        if house_only and current_street:
            current = start_row(house_only.group("num") or "", current_street, house_only.group("note") or "")
            current["owner_type"] = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
            continue

        if current is not None:
            if any(token in lowered for token in _BUILDER_TOKENS) and not current.get("contact_name"):
                current["contact_name"] = line
                current["owner_type"] = "builder"
                continue
            current["notes"] = _append_note(current.get("notes", ""), line)
            owner_type = _infer_owner_type(current.get("contact_name", ""), current.get("notes", ""))
            if owner_type:
                current["owner_type"] = owner_type

    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        owner_type = row.get("owner_type") or _infer_owner_type(row.get("contact_name", ""), row.get("notes", ""))
        source = _normalize_import_source(row.get("source", ""), owner_type, row.get("notes", ""), default_source="rp_data")
        normalized_rows.append(
            {
                **row,
                "owner_type": owner_type,
                "source": source,
                "suburb": row.get("suburb") or current_suburb,
            }
        )
    return normalized_rows


def _parse_import_rows(raw_data: str) -> list[dict[str, str]]:
    structured = _parse_structured_import_rows(raw_data)
    if structured:
        normalized: list[dict[str, str]] = []
        for row in structured:
            owner_type = row.get("owner_type") or _infer_owner_type(row.get("contact_name", ""), row.get("notes", ""))
            raw_type_token = _normalize_source_tag((row.get("source") or "").strip())
            if not owner_type and raw_type_token == "developer":
                owner_type = "builder"
            source = _normalize_import_source(row.get("source", ""), owner_type, row.get("notes", ""), default_source="rp_data")
            parsed_tags = _extract_import_tags(row.get("tags", ""), owner_type, row.get("notes", ""))
            normalized.append({**row, "owner_type": owner_type, "source": source, "tags": ",".join(parsed_tags)})
        return normalized
    return _parse_unstructured_import_rows(raw_data)

@router.post("/api/leads/import/preview")
async def preview_leads_import(body: ImportPreviewRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    parsed_rows = _parse_import_rows(body.csv_data)
    if not parsed_rows and (body.csv_data or "").strip():
        raise HTTPException(
            status_code=400,
            detail="No importable rows detected. Use CSV columns (street, house_no, suburb, ...) or paste address notes like '3 Wiese place - note'.",
        )
    matches = []
    from core.logic import get_deterministic_id
    for idx, row in enumerate(parsed_rows):
        street = row.get("street", "").strip()
        house_no = row.get("house_no", "").strip()
        suburb = row.get("suburb", "").strip()
        postcode = row.get("postcode", "").strip()
        contact_name = row.get("contact_name", "").strip()
        owner_type = (row.get("owner_type") or "").strip()
        phone = row.get("phone", "").strip()
        email = row.get("email", "").strip()
        raw_notes = row.get("notes", "").strip()
        source = _normalize_source_tag(row.get("source", "rp_data").strip() or "rp_data")

        if house_no and street:
            address = f"{house_no} {street}"
        else:
            address = row.get("address", "").strip()

        if not address:
            continue

        notes = _enrich_import_note(raw_notes, source)
        lead_id = get_deterministic_id(address)
        incoming_identity = build_address_identity(address=address, suburb=suburb, postcode=postcode)
        street_term = incoming_identity.street_name.split(" ")[0] if incoming_identity.street_name else ""
        street_like = f"%{street_term}%"
        house_like = f"%{incoming_identity.house_number}%"
        normalized_suburb = incoming_identity.suburb

        result = await session.execute(
            text(
                """
                SELECT id, address, canonical_address, suburb, postcode, owner_name, contact_emails, contact_phones, source_tags, lat, lng
                FROM leads
                WHERE id = :id
                   OR lower(COALESCE(address, '')) = :normalized_address
                   OR replace(lower(COALESCE(address, '')), ' ', '') = :compact_address
                   OR lower(COALESCE(canonical_address, '')) = :normalized_address
                   OR replace(lower(COALESCE(canonical_address, '')), ' ', '') = :compact_address
                   OR (
                        (:suburb = '' OR lower(COALESCE(suburb, '')) = :suburb)
                        AND (
                            (:street_term != '' AND lower(COALESCE(address, '')) LIKE :street_like)
                            OR (:house_no != '' AND lower(COALESCE(address, '')) LIKE :house_like)
                            OR lower(COALESCE(canonical_address, '')) LIKE :street_like
                        )
                   )
                LIMIT 8
                """
            ),
            {
                "id": lead_id,
                "normalized_address": incoming_identity.normalized,
                "compact_address": incoming_identity.compact,
                "suburb": normalized_suburb,
                "street_term": street_term,
                "street_like": street_like,
                "house_no": incoming_identity.house_number,
                "house_like": house_like,
            },
        )

        ranked_candidates: list[dict[str, Any]] = []
        for candidate in result.mappings().all():
            row_obj = dict(candidate)
            candidate_identity = build_address_identity(
                address=row_obj.get("canonical_address") or row_obj.get("address"),
                suburb=row_obj.get("suburb"),
                postcode=row_obj.get("postcode"),
            )
            verdict = classify_match(
                incoming_identity,
                candidate_identity,
                candidate_lat=float(row_obj.get("lat") or 0),
                candidate_lng=float(row_obj.get("lng") or 0),
            )
            score = 2 if verdict["match_confidence"] == "safe_exact" else 1
            ranked_candidates.append(
                {
                    "id": row_obj.get("id"),
                    "address": row_obj.get("address"),
                    "owner_name": row_obj.get("owner_name"),
                    "contact_emails": _decode_json_list(row_obj.get("contact_emails")),
                    "contact_phones": _decode_json_list(row_obj.get("contact_phones")),
                    "source_tags": _decode_json_list(row_obj.get("source_tags")),
                    "match_confidence": verdict["match_confidence"],
                    "match_reasons": verdict["match_reasons"],
                    "requires_confirmation": verdict["requires_confirmation"],
                    "safe_for_merge_all": verdict["safe_for_merge_all"],
                    "_score": score,
                }
            )

        ranked_candidates.sort(key=lambda item: item["_score"], reverse=True)
        primary = ranked_candidates[0] if ranked_candidates else None
        match_confidence = str(primary.get("match_confidence")) if primary else "new_create"
        requires_confirmation = bool(primary.get("requires_confirmation")) if primary else False
        safe_for_merge_all = bool(primary.get("safe_for_merge_all")) if primary else False
        match_reasons = list(primary.get("match_reasons") or []) if primary else ["no_match_found"]
        existing_id = str(primary.get("id")) if primary else None
        existing_name = str(primary.get("owner_name") or "") if primary else None

        match_type = "create"
        suggested_action = "create"
        if primary and match_confidence == "safe_exact":
            match_type = "exact_match"
            suggested_action = "merge"
        elif primary:
            match_type = "possible_duplicate"
            suggested_action = "review"

        incoming_email = str(email or "").strip().lower()
        incoming_phone = _normalize_phone(phone or "")
        existing_emails = {str(item).strip().lower() for item in (primary.get("contact_emails") or [])} if primary else set()
        existing_phones = {_normalize_phone(str(item or "")) for item in (primary.get("contact_phones") or [])} if primary else set()
        existing_name = str(primary.get("owner_name") or "").strip().lower() if primary else ""
        incoming_name = str(contact_name or "").strip().lower()

        address_match = bool(primary and match_confidence == "safe_exact")
        email_match = bool(incoming_email and incoming_email in existing_emails)
        phone_match = bool(incoming_phone and incoming_phone in existing_phones)
        name_match = bool(incoming_name and existing_name and incoming_name == existing_name)

        identity_hits: list[str] = []
        if address_match:
            identity_hits.append("address_match")
        if email_match:
            identity_hits.append("email_match")
        if phone_match:
            identity_hits.append("phone_match")
        if name_match:
            identity_hits.append("name_match")

        if identity_hits:
            requires_confirmation = True
            safe_for_merge_all = False
            match_type = "possible_duplicate"
            suggested_action = "review"
            match_reasons = _dedupe_text_list([
                *match_reasons,
                *identity_hits,
                "identity_match_requires_confirmation",
            ])

        matches.append({
            "row_id": f"row-{idx}",
            "address": address,
            "suburb": suburb,
            "contact_name": contact_name,
            "owner_type": owner_type,
            "email": email,
            "phone": phone,
            "notes": notes,
            "tags": row.get("tags", ""),
            "source": source,
            "action": "MERGE" if match_type == "exact_match" else "CREATE",
            "decision": "review" if suggested_action == "review" else ("merge" if match_type == "exact_match" else "create"),
            "match_type": match_type,
            "suggested_action": suggested_action,
            "existing_id": existing_id,
            "existing_name": existing_name,
            "existing_options": [{k: v for k, v in c.items() if not k.startswith("_")} for c in ranked_candidates],
            "match_confidence": match_confidence,
            "requires_confirmation": requires_confirmation,
            "safe_for_merge_all": safe_for_merge_all,
            "match_reasons": match_reasons,
            "candidate_count": len(ranked_candidates),
        })

    return {"matches": matches}

@router.post("/api/leads/import/confirm")
async def confirm_leads_import(body: ImportConfirmRequest, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    from core.logic import get_deterministic_id
    from core.utils import now_iso

    now = now_iso()
    imported_ids = []
    skipped = 0
    lead_column_rows = (
        await session.execute(
            text(
                """
                SELECT column_name, data_type, udt_name, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_name = 'leads'
                """
            )
        )
    ).mappings().all()
    lead_column_meta = {
        str(row.get("column_name") or ""): {
            "data_type": str(row.get("data_type") or ""),
            "udt_name": str(row.get("udt_name") or ""),
            "is_nullable": str(row.get("is_nullable") or ""),
            "column_default": row.get("column_default"),
        }
        for row in lead_column_rows
    }

    for index, lead in enumerate(body.leads):
        decision = (lead.operator_resolution or lead.decision or "").strip().lower()
        if not decision:
            decision = "merge" if str(lead.action or "").upper() == "MERGE" else "create"
        if decision in {"skip", "review"}:
            skipped += 1
            continue
        if lead.requires_confirmation and decision == "merge" and not lead.target_lead_id and not lead.existing_id:
            raise HTTPException(status_code=400, detail=f"Row {lead.row_id or '?'} requires explicit merge target confirmation.")

        deterministic_id = get_deterministic_id(lead.address)
        target_id = (lead.target_lead_id or lead.existing_id or deterministic_id).strip()
        existing = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": target_id})).mappings().first()
        if existing is None:
            existing_by_address = (
                await session.execute(
                    text(
                        """
                        SELECT * FROM leads
                        WHERE lower(COALESCE(address, '')) = :address
                          AND (:suburb = '' OR lower(COALESCE(suburb, '')) = :suburb)
                        LIMIT 1
                        """
                    ),
                    {
                        "address": str(lead.address or "").strip().lower(),
                        "suburb": str(lead.suburb or "").strip().lower(),
                    },
                )
            ).mappings().first()
            if existing_by_address is not None:
                existing = existing_by_address
                target_id = str(existing_by_address.get("id") or target_id)
        source_tag = _normalize_source_tag(lead.source)
        owner_type_value = (lead.owner_type or _infer_owner_type(lead.contact_name, lead.notes)).strip()
        normalized_note = _enrich_import_note(lead.notes, source_tag)
        primary_note, special_notes = _split_import_notes(normalized_note)
        normalized_note = primary_note or normalized_note
        source_queue_bucket = _source_queue_bucket(source_tag)
        new_tags = _derive_import_source_tags(source_tag, owner_type_value, getattr(lead, "tags", ""), lead.notes)

        should_merge = existing is not None and decision in {"merge", "create"}

        if should_merge:
            current = dict(existing)
            emails = _dedupe_text_list([*(_decode_json_list(current.get("contact_emails"))), lead.email])
            phones = _dedupe_by_phone([*(_decode_json_list(current.get("contact_phones"))), lead.phone])
            contacts = _merge_contacts(_decode_json_list(current.get("contacts")), lead.contact_name, lead.phone, lead.email)
            contacts = _prioritize_contact(contacts, lead.contact_name, lead.phone, lead.email)
            tags = _dedupe_text_list([*(_decode_json_list(current.get("source_tags"))), *new_tags])
            key_details = _dedupe_text_list([*(_decode_json_list(current.get("key_details"))), normalized_note, *special_notes])
            activity_log = _append_activity(
                _decode_json_list(current.get("activity_log")),
                _build_activity_entry(f"{source_tag}_import", normalized_note or "Import merge", current.get("status") or "captured", source_tag),
            )
            stage_history = _append_stage_note(
                _decode_json_list(current.get("stage_note_history")),
                normalized_note or "Import merge",
                str(current.get("status") or "captured"),
                source_tag,
            )
            for special_note in special_notes:
                stage_history = _append_stage_note(
                    stage_history,
                    special_note,
                    str(current.get("status") or "captured"),
                    "email",
                    "Import note",
                )
                activity_log = _append_activity(
                    activity_log,
                    _build_activity_entry("note", special_note, current.get("status") or "captured", "email", "Import note"),
                )
            merged_notes = _append_notes_text(current.get("notes"), normalized_note or "Import merge")
            for special_note in special_notes:
                merged_notes = _append_notes_text(merged_notes, special_note)

            await session.execute(
                text(
                    """
                    UPDATE leads
                    SET owner_name = :owner_name,
                        owner_type = CASE WHEN :owner_type != '' THEN :owner_type ELSE owner_type END,
                        contact_emails = :emails,
                        contact_phones = :phones,
                        contacts = :contacts,
                        source_tags = :source_tags,
                        key_details = :key_details,
                        stage_note = :stage_note,
                        notes = :notes,
                        stage_note_history = :stage_note_history,
                        activity_log = :activity_log,
                        queue_bucket = CASE
                            WHEN COALESCE(queue_bucket, '') IN ('', 'captured', 'new', 'door_knock') THEN :queue_bucket
                            ELSE queue_bucket
                        END,
                        last_activity_type = :last_activity_type,
                        updated_at = :updated_at
                    WHERE id = :id
                    """
                ),
                {
                    "owner_name": current.get("owner_name") or (lead.contact_name or "").strip() or "Owner record pending",
                    "owner_type": owner_type_value,
                    "emails": json.dumps(emails),
                    "phones": json.dumps(phones),
                    "contacts": json.dumps(contacts),
                    "source_tags": json.dumps(tags),
                    "key_details": json.dumps(key_details),
                    "stage_note": normalized_note or "Import merge",
                    "notes": merged_notes,
                    "stage_note_history": json.dumps(stage_history),
                    "activity_log": json.dumps(activity_log),
                    "queue_bucket": source_queue_bucket,
                    "last_activity_type": f"{source_tag}_import",
                    "updated_at": now,
                    "id": target_id,
                },
            )
            for special_note in special_notes:
                await session.execute(
                    text(
                        """
                        INSERT INTO notes (lead_id, note_type, content, created_at)
                        SELECT :lead_id, :note_type, :content, :created_at
                        WHERE NOT EXISTS (
                            SELECT 1
                            FROM notes
                            WHERE lead_id = :lead_id AND note_type = :note_type AND content = :content
                        )
                        """
                    ),
                    {"lead_id": target_id, "note_type": "email_sent", "content": special_note, "created_at": now},
                )
            imported_ids.append(target_id)
        else:
            lead_id = deterministic_id
            if existing is not None:
                lead_id = hashlib.md5(f"{lead.address}|{lead.email}|{lead.phone}|{now}|{index}".lower().encode("utf-8")).hexdigest()
            stage_history = _append_stage_note([], normalized_note or "Import create", "captured", source_tag)
            for special_note in special_notes:
                stage_history = _append_stage_note(stage_history, special_note, "captured", "email", "Import note")
            activity_log = [
                _build_activity_entry(
                    f"{source_tag}_import",
                    normalized_note or "Import create",
                    "captured",
                    source_tag,
                )
            ]
            for special_note in special_notes:
                activity_log.append(_build_activity_entry("note", special_note, "captured", "email", "Import note"))
            contacts = _merge_contacts([], lead.contact_name, lead.phone, lead.email)
            contacts = _prioritize_contact(contacts, lead.contact_name, lead.phone, lead.email)
            create_notes = normalized_note or "Import create"
            for special_note in special_notes:
                create_notes = _append_notes_text(create_notes, special_note)
            details = [normalized_note] if normalized_note else []
            details.extend(special_notes)
            payload: dict[str, Any] = {
                "id": lead_id,
                "address": lead.address,
                "suburb": lead.suburb,
                "owner_name": lead.contact_name or "Owner record pending",
                "owner_type": owner_type_value or None,
                "trigger_type": _source_trigger_type(source_tag),
                "record_type": "manual_entry",
                "contact_status": "unreviewed",
                "contacts": json.dumps(contacts),
                "contact_emails": json.dumps([lead.email] if lead.email else []),
                "contact_phones": json.dumps([lead.phone] if lead.phone else []),
                "source_tags": json.dumps(new_tags),
                "key_details": json.dumps(_dedupe_text_list(details)),
                "scenario": normalized_note or f"Import from {source_tag}",
                "date_found": now,
                "status": "captured",
                "queue_bucket": source_queue_bucket,
                "route_queue": source_queue_bucket,
                "stage_note": normalized_note or "Import create",
                "notes": create_notes,
                "stage_note_history": json.dumps(stage_history),
                "activity_log": json.dumps(activity_log),
                "last_activity_type": f"{source_tag}_import",
                "created_at": now,
                "updated_at": now,
                "heat_score": 0,
                "confidence_score": 0,
                "lat": 0.0,
                "lng": 0.0,
                "est_value": 0,
                "conversion_score": 0,
                "compliance_score": 0,
                "readiness_score": 0,
                "call_today_score": 0,
                "evidence_score": 0,
                "preferred_contact_method": "",
                "followup_frequency": "none",
                "market_updates_opt_in": False,
                "followup_status": "active",
                "lead_archetype": "",
                "contactability_status": "",
                "owner_verified": False,
                "contact_role": "",
                "cadence_name": "",
                "cadence_step": 0,
                "next_action_type": "",
                "next_action_channel": "",
                "next_action_title": "",
                "next_action_reason": "",
                "next_message_template": "",
                "last_outcome": "",
                "objection_reason": "",
                "preferred_channel": "",
                "strike_zone": "",
                "touches_14d": 0,
                "touches_30d": 0,
            }
            for column_name, meta in lead_column_meta.items():
                if column_name in payload:
                    continue
                if meta.get("is_nullable") == "NO" and meta.get("column_default") is None:
                    payload[column_name] = _default_value_for_column(
                        str(meta.get("data_type") or ""),
                        str(meta.get("udt_name") or ""),
                    )
            insert_columns = [column for column in payload.keys() if column in lead_column_meta]
            insert_sql = f"INSERT INTO leads ({', '.join(insert_columns)}) VALUES ({', '.join(f':{column}' for column in insert_columns)})"
            await session.execute(text(insert_sql), {column: payload[column] for column in insert_columns})
            for special_note in special_notes:
                await session.execute(
                    text(
                        """
                        INSERT INTO notes (lead_id, note_type, content, created_at)
                        SELECT :lead_id, :note_type, :content, :created_at
                        WHERE NOT EXISTS (
                            SELECT 1
                            FROM notes
                            WHERE lead_id = :lead_id AND note_type = :note_type AND content = :content
                        )
                        """
                    ),
                    {"lead_id": lead_id, "note_type": "email_sent", "content": special_note, "created_at": now},
                )
            imported_ids.append(lead_id)

    await session.commit()
    _invalidate_analytics_cache()
    for l_id in set(imported_ids):
        try:
            await refresh_hermes_for_lead(session, l_id, actor="manual_import")
        except Exception:
            pass

    return {"status": "ok", "imported": len(imported_ids), "skipped": skipped}


@router.post("/api/leads/bulk-send-email")
async def bulk_send_email(body: dict, api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    """Send a templated email to multiple leads at once with tracking."""
    lead_ids = body.get("lead_ids", [])
    subject = body.get("subject", "")
    html_body = body.get("body", "")
    if not lead_ids or not subject or not html_body:
        raise HTTPException(status_code=400, detail="lead_ids, subject, and body are required")

    sent = 0
    failed = 0
    results = []

    for lid in lead_ids:
        lead_row = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lid})).mappings().first()
        if not lead_row:
            failed += 1
            results.append({"lead_id": lid, "status": "not_found"})
            continue
        lead = _decode_row(lead_row)
        emails = lead.get("contact_emails") or []
        if isinstance(emails, str):
            try:
                emails = json.loads(emails)
            except Exception:
                emails = [emails] if emails else []
        if not emails:
            failed += 1
            results.append({"lead_id": lid, "status": "no_email"})
            continue

        recipient = emails[0]

        # Fill placeholders — core fields + Cotality-enriched property data
        owner_name = lead.get("owner_name") or "there"
        first_name = owner_name.split()[0] if owner_name and owner_name != "Owner record pending" else "there"
        address = lead.get("address") or ""
        suburb = lead.get("suburb") or ""

        placeholders = {
            "{owner_first_name}": first_name,
            "{owner_name}": owner_name,
            "{address}": address,
            "{suburb}": suburb,
            # Cotality-enriched valuation fields
            "{estimated_value_low}": _format_moneyish(lead.get("estimated_value_low"), "N/A").lstrip("$"),
            "{estimated_value_high}": _format_moneyish(lead.get("estimated_value_high"), "N/A").lstrip("$"),
            "{ownership_duration_years}": str(lead.get("ownership_duration_years") or "N/A"),
            "{yield_estimate}": str(lead.get("yield_estimate") or "N/A"),
            "{rental_estimate_low}": _format_moneyish(lead.get("rental_estimate_low"), "N/A").lstrip("$"),
            "{rental_estimate_high}": _format_moneyish(lead.get("rental_estimate_high"), "N/A").lstrip("$"),
            # Suburb / market stats
            "{suburb_median}": _format_moneyish(lead.get("suburb_median"), "N/A").lstrip("$"),
            "{suburb_growth}": str(lead.get("suburb_growth") or "N/A"),
            "{nearby_sales_count}": str(lead.get("nearby_sales_count") or "N/A"),
            "{suburb_dom}": str(lead.get("suburb_dom") or "N/A"),
        }

        filled_subject = subject
        filled_body = html_body
        for placeholder, value in placeholders.items():
            filled_subject = filled_subject.replace(placeholder, value)
            filled_body = filled_body.replace(placeholder, value)

        # Add tracking
        tracking_id = None
        tracked_body = filled_body
        try:
            from services.email_tracking import generate_tracking_id, wrap_email_with_tracking
            tracking_id = generate_tracking_id()
            backend_url = os.getenv("BASE_URL") or os.getenv("RENDER_EXTERNAL_URL") or "http://localhost:8001"
            tracked_body = wrap_email_with_tracking(filled_body, tracking_id, backend_url)

            # Record send event for tracking resolution
            from models.sql_models import EmailEvent
            send_event = EmailEvent(
                id=str(uuid.uuid4()),
                lead_id=lid,
                tracking_id=tracking_id,
                event_type="send",
                created_at=now_iso(),
            )
            session.add(send_event)
        except Exception:
            pass

        try:
            email_req = SendEmailRequest(
                account_id="",
                recipient=recipient,
                subject=filled_subject,
                body=tracked_body,
            )
            await asyncio.to_thread(send_email_service, None, email_req)

            # Log activity
            activity_log = _append_activity(
                _decode_json_list(lead.get("activity_log")),
                _build_activity_entry("email_sent", f"Bulk email: {filled_subject}", lead.get("status"), "email", filled_subject, recipient),
            )
            await session.execute(
                text("UPDATE leads SET activity_log = :activity_log, last_outbound_at = :ts, updated_at = :ts WHERE id = :id"),
                {"activity_log": json.dumps(activity_log), "ts": now_iso(), "id": lid},
            )
            sent += 1
            results.append({"lead_id": lid, "status": "sent", "recipient": recipient, "tracking_id": tracking_id})
        except Exception as exc:
            failed += 1
            results.append({"lead_id": lid, "status": "failed", "error": str(exc)})

    await session.commit()
    return {"sent": sent, "failed": failed, "results": results}


@router.post("/api/system/door-knock-sync/run")
async def run_door_knock_sync_now(
    session: SessionDep = None,
    api_key: APIKeyDep = "",
):
    if not door_knock_sync_enabled():
        raise HTTPException(status_code=400, detail="DOOR_KNOCK_SYNC_FILE is not configured.")
    result = await run_door_knock_sheet_sync_once(session)
    return {"status": "ok", **result}


@router.get("/api/analytics/email-tracking")
async def get_email_tracking_analytics(api_key: str = Depends(get_api_key), session: AsyncSession = Depends(get_session)):
    """Get email tracking metrics — sends, opens, clicks, rates."""
    try:
        total_sends = (await session.execute(text("SELECT COUNT(*) FROM email_events WHERE event_type = 'send'"))).scalar() or 0
        total_opens = (await session.execute(text("SELECT COUNT(*) FROM email_events WHERE event_type = 'open'"))).scalar() or 0
        total_clicks = (await session.execute(text("SELECT COUNT(*) FROM email_events WHERE event_type = 'click'"))).scalar() or 0
        unique_opens = (await session.execute(text("SELECT COUNT(DISTINCT tracking_id) FROM email_events WHERE event_type = 'open'"))).scalar() or 0
        unique_clicks = (await session.execute(text("SELECT COUNT(DISTINCT tracking_id) FROM email_events WHERE event_type = 'click'"))).scalar() or 0

        open_rate = round(unique_opens / max(total_sends, 1) * 100, 1)
        click_rate = round(unique_clicks / max(total_sends, 1) * 100, 1)

        # Recent events (last 50)
        recent_result = await session.execute(text(
            "SELECT e.tracking_id, e.event_type, e.opened_at, e.link_url, e.lead_id, "
            "l.address, l.owner_name, l.suburb "
            "FROM email_events e LEFT JOIN leads l ON e.lead_id = l.id "
            "ORDER BY e.opened_at DESC LIMIT 50"
        ))
        recent = [dict(r._mapping) for r in recent_result]

        return {
            "total_sends": total_sends,
            "total_opens": total_opens,
            "total_clicks": total_clicks,
            "unique_opens": unique_opens,
            "unique_clicks": unique_clicks,
            "open_rate": open_rate,
            "click_rate": click_rate,
            "recent_events": recent,
        }
    except Exception as exc:
        return {
            "total_sends": 0, "total_opens": 0, "total_clicks": 0,
            "unique_opens": 0, "unique_clicks": 0,
            "open_rate": 0, "click_rate": 0,
            "recent_events": [],
            "error": str(exc),
        }
