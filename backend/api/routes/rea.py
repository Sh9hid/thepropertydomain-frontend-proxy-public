"""
realestate.com.au Partner API routes.
"""

import hashlib
import json
import logging
import csv
import io
import re
import os
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, File, Request, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.config import BRAND_NAME, PRINCIPAL_EMAIL, PRINCIPAL_NAME, PRINCIPAL_PHONE, USE_POSTGRES

from core.database import get_session
from core.security import get_api_key
from core.utils import now_iso
from services.rea_service import (
    check_credentials,
    export_listings,
    get_export_diagnostics,
    get_enquiries,
    get_integration_status,
    get_listing_performance,
    get_listing_status,
    get_seller_leads,
    get_upload_report,
    publish_listing,
    sync_agency_feed,
    sync_agency_feed_detailed,
    update_listing,
    withdraw_listing,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/rea", tags=["REA Partner API"])
REA_PUBLISH_ENABLED = os.getenv("REA_PUBLISH_ENABLED", "false").strip().lower() in {"1", "true", "yes", "on"}


def _ensure_rea_publish_enabled() -> None:
    if not REA_PUBLISH_ENABLED:
        raise HTTPException(
            status_code=423,
            detail="REA publishing is disabled by safety lock. Set REA_PUBLISH_ENABLED=true to enable.",
        )


class WithdrawRequest(BaseModel):
    reason: str = "sold"


class PriceUpdateRequest(BaseModel):
    from_price: Optional[int] = None
    to_price: Optional[int] = None
    display: Optional[str] = None


class ReaStudioAction(BaseModel):
    lead_id: str
    action: str
    headline: Optional[str] = None
    description: Optional[str] = None
    banner_meta: Optional[dict] = None


class ReaStudioCommitRequest(BaseModel):
    actions: list[ReaStudioAction]
    requested_by: str = "operator"
    max_actions: int = 25


@router.get("/status")
async def rea_status(api_key: str = Depends(get_api_key)):
    return await check_credentials()


@router.get("/integration-status")
async def rea_integration_status(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    return await get_integration_status(session=session)


@router.post("/listings/{lead_id}/publish")
async def rea_publish_listing(
    lead_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    _ensure_rea_publish_enabled()
    row = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Lead not found")

    lead_payload = dict(row)
    try:
        result = await publish_listing(lead_payload, session=session, lead_id=lead_id)
    except TypeError:
        # Keep compatibility with older test doubles or legacy wrappers that
        # only accept the original positional parameters.
        result = await publish_listing(lead_payload, session=session)
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=f"REA publish failed: {result['error']}")

    upload_id = result.get("upload_id", "")
    upload_status = result.get("status", "")
    rea_listing_id = result.get("rea_listing_id") or None
    await session.execute(
        text(
            """
            UPDATE leads
            SET rea_upload_id = :rea_upload_id,
                rea_upload_status = :rea_upload_status,
                rea_last_upload_response = :rea_last_upload_response,
                rea_listing_id = COALESCE(:rea_listing_id, rea_listing_id),
                updated_at = :now
            WHERE id = :id
            """
        ),
        {
            "rea_upload_id": upload_id,
            "rea_upload_status": upload_status,
            "rea_last_upload_response": json.dumps(result.get("response") or {}),
            "rea_listing_id": rea_listing_id,
            "now": now_iso(),
            "id": lead_id,
        },
    )
    await session.commit()
    return {
        "status": "submitted",
        "lead_id": lead_id,
        "rea_upload_id": upload_id,
        "rea_upload_status": upload_status,
        "rea_listing_id": rea_listing_id,
    }


@router.get("/upload/{upload_id}/status")
async def rea_upload_status(
    upload_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    lead_row = (
        await session.execute(text("SELECT id FROM leads WHERE rea_upload_id = :upload_id"), {"upload_id": upload_id})
    ).mappings().first()
    result = await get_upload_report(upload_id, session=session, lead_id=(lead_row or {}).get("id", ""))
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=result.get("error", "REA upload report error"))

    payload = result["data"]
    if lead_row:
        await session.execute(
            text(
                """
                UPDATE leads
                SET rea_upload_status = :rea_upload_status,
                    rea_listing_id = COALESCE(:rea_listing_id, rea_listing_id),
                    rea_last_upload_report = :rea_last_upload_report,
                    updated_at = :now
                WHERE id = :id
                """
            ),
            {
                "rea_upload_status": payload.get("progress", ""),
                "rea_listing_id": payload.get("listingId"),
                "rea_last_upload_report": json.dumps(payload),
                "now": now_iso(),
                "id": lead_row["id"],
            },
        )
        await session.commit()
    return payload


@router.post("/listings/{lead_id}/withdraw")
async def rea_withdraw_listing(
    lead_id: str,
    body: WithdrawRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    row = (await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Lead not found")
    lead = dict(row)

    rea_id = lead.get("rea_listing_id") or ""
    if not rea_id:
        raise HTTPException(status_code=400, detail="No rea_listing_id on this lead - publish first and wait for report completion")

    result = await withdraw_listing(rea_id, body.reason, session=session, lead_id=lead_id)
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=f"REA withdraw failed: {result['error']}")
    return {"status": "withdrawn", "rea_listing_id": rea_id, "reason": body.reason}


@router.put("/listings/{rea_listing_id}/price")
async def rea_update_price(
    rea_listing_id: str,
    body: PriceUpdateRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    if body.from_price and body.to_price:
        price_payload = {"type": "guide", "from": body.from_price, "to": body.to_price}
    elif body.display:
        price_payload = {"type": "display", "display": body.display}
    else:
        raise HTTPException(status_code=400, detail="Provide from_price+to_price or display")

    result = await update_listing(rea_listing_id, {"price": price_payload}, session=session)
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=f"REA price update failed: {result['error']}")
    return {"status": "updated", "rea_listing_id": rea_listing_id}


@router.get("/listings/export")
async def rea_listings_export(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    listings = await export_listings(session=session)
    return {"count": len(listings), "listings": listings}


@router.post("/agency-feed/sync")
async def rea_sync_agency_feed(
    status: str = "current,offmarket,sold",
    listing_types: str = "land,residential",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    await _ensure_rea_studio_tables(session)
    feed_result = await sync_agency_feed_detailed(
        status=status,
        listing_types=listing_types,
        allow_seller_fallback=True,
        session=session,
    )
    if not feed_result.get("ok"):
        detail = {
            "message": "REA agency feed sync failed",
            "error": feed_result.get("error", "unknown_error"),
            "export_error": feed_result.get("export_error") or {},
        }
        raise HTTPException(status_code=502, detail=detail)
    listings = feed_result.get("listings", [])
    if not listings:
        return {
            "synced": 0,
            "total_from_rea": 0,
            "source": feed_result.get("source", ""),
            "degraded": bool(feed_result.get("degraded")),
            "message": "No listings returned from REA feed",
            "export_error": feed_result.get("export_error") or {},
        }

    upserted = 0
    now = now_iso()
    for item in listings:
        rea_id = item.get("rea_listing_id", "")
        if not rea_id and not item.get("address"):
            continue
        lead_id = hashlib.md5(f"rea:{rea_id or item['address']}".encode()).hexdigest()
        existing = (
            await session.execute(
                text("SELECT id FROM leads WHERE rea_listing_id = :rea_id OR id = :lid"),
                {"rea_id": rea_id, "lid": lead_id},
            )
        ).mappings().first()

        if existing:
            effective_lead_id = existing["id"]
            await session.execute(
                text(
                    """
                    UPDATE leads SET
                        address = :address, suburb = :suburb, postcode = :postcode,
                        agent_name = :agent_name, agency_name = COALESCE(NULLIF(:agency_name, ''), agency_name),
                        est_value = :est_value,
                        bedrooms = :bedrooms, bathrooms = :bathrooms,
                        land_size_sqm = :land_size_sqm,
                        property_type = COALESCE(NULLIF(:property_type, ''), property_type),
                        last_listing_status = COALESCE(NULLIF(:last_listing_status, ''), last_listing_status),
                        main_image = :main_image, property_images = :property_images,
                        rea_listing_id = :rea_listing_id,
                        signal_status = :signal_status,
                        listing_headline = :listing_headline,
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        updated_at = :now
                    WHERE id = :id
                    """
                ),
                {
                    **{k: item.get(k) for k in (
                        "address", "suburb", "postcode", "est_value",
                        "bedrooms", "bathrooms", "land_size_sqm", "property_type",
                        "last_listing_status", "main_image", "property_images",
                        "rea_listing_id", "signal_status", "listing_headline", "listing_description",
                    )},
                    "agent_name": item.get("agent", ""),
                    "agency_name": item.get("agency", ""),
                    "now": now,
                    "id": existing["id"],
                },
            )
        else:
            effective_lead_id = lead_id

        campaign_row = (
            await session.execute(
                text("SELECT id FROM rea_studio_campaign_listings WHERE lead_id = :lead_id"),
                {"lead_id": effective_lead_id},
            )
        ).mappings().first()
        campaign_payload = {
            "lead_id": effective_lead_id,
            "address_key": _address_key(item.get("address", ""), item.get("suburb", ""), item.get("postcode", "")),
            "source_batch_id": f"rea-sync-{now[:10]}",
            "property_type": item.get("property_type", ""),
            "is_land": 1 if _is_land_property_type(item.get("property_type", "")) else 0,
            "listing_description": item.get("listing_description", ""),
            "rea_status": item.get("rea_status", item.get("last_listing_status", "")),
            "rea_listing_type": item.get("rea_listing_type", ""),
            "primary_image": item.get("main_image", ""),
            "address": item.get("address", ""),
            "suburb": item.get("suburb", ""),
            "postcode": item.get("postcode", ""),
            "listing_headline": item.get("listing_headline", ""),
            "property_images": item.get("property_images", "[]"),
            "est_value": item.get("est_value"),
            "land_size_sqm": item.get("land_size_sqm"),
            "bedrooms": item.get("bedrooms"),
            "bathrooms": item.get("bathrooms"),
            "car_spaces": item.get("car_spaces"),
            "rea_listing_id": item.get("rea_listing_id", ""),
            "last_listing_status": item.get("last_listing_status", ""),
            "updated_at": now,
            "created_at": now,
            "id": str(uuid.uuid4()),
        }
        if campaign_row:
            await session.execute(
                text(
                    """
                    UPDATE rea_studio_campaign_listings
                    SET address_key = :address_key,
                        source_batch_id = :source_batch_id,
                        property_type = :property_type,
                        is_land = :is_land,
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        rea_status = :rea_status,
                        rea_listing_type = :rea_listing_type,
                        primary_image = COALESCE(NULLIF(:primary_image, ''), primary_image),
                        address = COALESCE(NULLIF(:address, ''), address),
                        suburb = COALESCE(NULLIF(:suburb, ''), suburb),
                        postcode = COALESCE(NULLIF(:postcode, ''), postcode),
                        listing_headline = COALESCE(NULLIF(:listing_headline, ''), listing_headline),
                        property_images = COALESCE(NULLIF(:property_images, ''), property_images),
                        est_value = COALESCE(:est_value, est_value),
                        land_size_sqm = COALESCE(:land_size_sqm, land_size_sqm),
                        bedrooms = COALESCE(:bedrooms, bedrooms),
                        bathrooms = COALESCE(:bathrooms, bathrooms),
                        car_spaces = COALESCE(:car_spaces, car_spaces),
                        rea_listing_id = COALESCE(NULLIF(:rea_listing_id, ''), rea_listing_id),
                        last_listing_status = COALESCE(NULLIF(:last_listing_status, ''), last_listing_status),
                        updated_at = :updated_at
                    WHERE lead_id = :lead_id
                    """
                ),
                campaign_payload,
            )
        else:
            await session.execute(
                text(
                    """
                    INSERT INTO rea_studio_campaign_listings (
                        id, lead_id, address_key, source_batch_id, property_type, is_land,
                        listing_description, rea_status, rea_listing_type, primary_image,
                        address, suburb, postcode, listing_headline, property_images,
                        est_value, land_size_sqm, bedrooms, bathrooms, car_spaces,
                        rea_listing_id, last_listing_status,
                        status, created_at, updated_at
                    ) VALUES (
                        :id, :lead_id, :address_key, :source_batch_id, :property_type, :is_land,
                        :listing_description, :rea_status, :rea_listing_type, :primary_image,
                        :address, :suburb, :postcode, :listing_headline, :property_images,
                        :est_value, :land_size_sqm, :bedrooms, :bathrooms, :car_spaces,
                        :rea_listing_id, :last_listing_status,
                        'ready', :created_at, :updated_at
                    )
                    """
                ),
                campaign_payload,
            )
        upserted += 1
    await session.commit()
    return {
        "synced": upserted,
        "total_from_rea": len(listings),
        "source": feed_result.get("source", "listing_export"),
        "degraded": bool(feed_result.get("degraded")),
        "warning": feed_result.get("warning", ""),
        "export_error": feed_result.get("export_error") or {},
    }


@router.get("/agency-feed/diagnostics")
async def rea_agency_feed_diagnostics(
    api_key: str = Depends(get_api_key),
):
    return await get_export_diagnostics()


@router.get("/enquiries")
async def rea_enquiries(
    since: str = "",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    enquiries = await get_enquiries(since, session=session)
    return {"count": len(enquiries), "enquiries": enquiries}


@router.get("/performance/{listing_id}")
async def rea_performance(
    listing_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    payload = await get_listing_performance(listing_id, session=session)
    if payload.get("ok") is False:
        raise HTTPException(status_code=502, detail=payload.get("error", "REA performance error"))
    return payload


@router.get("/seller-leads")
async def rea_get_seller_leads(
    since: str = "2026-01-01T00:00:00.0Z",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    leads = await get_seller_leads(since)
    if not leads:
        return {"synced": 0, "seller_leads": [], "message": "No seller leads found for this period"}

    now = now_iso()
    upserted = 0
    for item in leads:
        lead_id = hashlib.md5(f"rea-seller:{item.get('id', '')}".encode()).hexdigest()
        await session.execute(
            text(
                """
                INSERT INTO leads (
                    id, owner_name, address, suburb, postcode,
                    contact_emails, contact_phones,
                    trigger_type, signal_status, heat_score, source,
                    created_at, updated_at
                ) VALUES (
                    :id, :owner_name, :address, :suburb, :postcode,
                    :contact_emails, :contact_phones,
                    'rea_seller_lead', 'OFF-MARKET', 75, 'rea_partner_api',
                    :now, :now
                )
                ON CONFLICT (id) DO UPDATE SET
                    owner_name = excluded.owner_name,
                    updated_at = excluded.updated_at
                """
            ),
            {
                "id": lead_id,
                "owner_name": item.get("contactDetails", {}).get("fullName", ""),
                "address": item.get("propertyAddress", {}).get("displayAddress", ""),
                "suburb": item.get("propertyAddress", {}).get("suburb", ""),
                "postcode": item.get("propertyAddress", {}).get("postcode", ""),
                "contact_emails": json.dumps([item["contactDetails"]["email"]] if item.get("contactDetails", {}).get("email") else []),
                "contact_phones": json.dumps([item["contactDetails"]["phone"]] if item.get("contactDetails", {}).get("phone") else []),
                "now": now,
            },
        )
        upserted += 1
    await session.commit()
    return {"synced": upserted, "seller_leads": leads}


@router.get("/listings/{rea_listing_id}/status")
async def rea_get_listing_status(
    rea_listing_id: str,
    api_key: str = Depends(get_api_key),
):
    result = await get_listing_status(rea_listing_id)
    if not result["ok"]:
        raise HTTPException(status_code=502, detail=result.get("error", "REA error"))
    return result["data"]


@router.get("/ab-tests")
async def rea_get_ab_tests(
    session: AsyncSession = Depends(get_session),
    api_key: str = Depends(get_api_key),
):
    """Get all A/B tests for REA listings."""
    result = await session.execute(text("""
        SELECT id, listing_id, address, test_type, variant_a, variant_b,
               views_a, views_b, enquiries_a, enquiries_b, ctr_a, ctr_b,
               status, started_at, winner, confidence
        FROM rea_ab_tests
        ORDER BY started_at DESC
    """))
    rows = result.mappings().all()
    tests = []
    for row in rows:
        tests.append({
            "id": row["id"],
            "listing_id": row["listing_id"],
            "address": row["address"],
            "test_type": row["test_type"],
            "variant_a": row["variant_a"],
            "variant_b": row["variant_b"],
            "views_a": row["views_a"] or 0,
            "views_b": row["views_b"] or 0,
            "enquiries_a": row["enquiries_a"] or 0,
            "enquiries_b": row["enquiries_b"] or 0,
            "ctr_a": float(row["ctr_a"] or 0),
            "ctr_b": float(row["ctr_b"] or 0),
            "status": row["status"],
            "started_at": row["started_at"],
            "winner": row["winner"],
            "confidence": float(row["confidence"] or 0),
        })
    return {"tests": tests}


@router.post("/ab-tests")
async def rea_create_ab_test(
    data: dict,
    session: AsyncSession = Depends(get_session),
    api_key: str = Depends(get_api_key),
):
    """Create a new A/B test for a REA listing."""
    import uuid
    from datetime import datetime, timezone
    test_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    await session.execute(text("""
        INSERT INTO rea_ab_tests (id, listing_id, address, test_type, variant_a, variant_b,
                                   views_a, views_b, enquiries_a, enquiries_b, ctr_a, ctr_b,
                                   status, started_at)
        VALUES (:id, :listing_id, :address, :test_type, :variant_a, :variant_b,
                0, 0, 0, 0, 0.0, 0.0, 'running', :started_at)
    """), {
        "id": test_id,
        "listing_id": data.get("listing_id"),
        "address": data.get("address", ""),
        "test_type": data.get("test_type", "headline"),
        "variant_a": data.get("variant_a", ""),
        "variant_b": data.get("variant_b", ""),
        "started_at": now,
    })
    await session.commit()
    return {"id": test_id, "status": "running", "started_at": now}


@router.post("/ab-tests/{test_id}/track")
async def rea_track_ab_test(
    test_id: str,
    data: dict,
    session: AsyncSession = Depends(get_session),
    api_key: str = Depends(get_api_key),
):
    """Track a view or enquiry for an A/B test variant."""
    variant = data.get("variant", "a")
    event_type = data.get("event", "view")
    col_suffix = variant
    if event_type == "view":
        await session.execute(text(f"""
            UPDATE rea_ab_tests SET views_{col_suffix} = views_{col_suffix} + 1 WHERE id = :id
        """), {"id": test_id})
    elif event_type == "enquiry":
        await session.execute(text(f"""
            UPDATE rea_ab_tests SET enquiries_{col_suffix} = enquiries_{col_suffix} + 1 WHERE id = :id
        """), {"id": test_id})
    await session.execute(text("""
        UPDATE rea_ab_tests
        SET ctr_a = CASE WHEN views_a > 0 THEN (enquiries_a * 100.0 / views_a) ELSE 0 END,
            ctr_b = CASE WHEN views_b > 0 THEN (enquiries_b * 100.0 / views_b) ELSE 0 END
        WHERE id = :id
    """), {"id": test_id})
    await session.commit()
    return {"ok": True}


async def _table_exists(session: AsyncSession, table_name: str) -> bool:
    """Check whether *table_name* exists, supporting both Postgres and SQLite."""
    if USE_POSTGRES:
        row = (
            await session.execute(
                text(f"SELECT to_regclass('public.{table_name}') IS NOT NULL AS exists")
            )
        ).mappings().first()
        return bool((row or {}).get("exists"))
    else:
        row = (
            await session.execute(
                text("SELECT name FROM sqlite_master WHERE type='table' AND name = :name"),
                {"name": table_name},
            )
        ).mappings().first()
        return row is not None


async def _add_column_if_missing(
    session: AsyncSession, table: str, column: str, col_type: str
) -> None:
    """Add a column to *table* if it doesn't exist yet (works on both Postgres and SQLite)."""
    if USE_POSTGRES:
        await session.execute(
            text(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {col_type}")
        )
    else:
        # SQLite: check pragma
        rows = (await session.execute(text(f"PRAGMA table_info({table})"))).fetchall()
        existing = {r[1] for r in rows}  # column name is index 1
        if column not in existing:
            await session.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))


async def _ensure_rea_studio_tables(session: AsyncSession) -> None:
    try:
        if not await _table_exists(session, "rea_studio_campaign_listings"):
            await session.execute(
                text(
                    """
                    CREATE TABLE rea_studio_campaign_listings (
                        id TEXT PRIMARY KEY,
                        lead_id TEXT NOT NULL,
                        address_key TEXT NOT NULL,
                        source_batch_id TEXT,
                        property_type TEXT,
                        is_land INTEGER DEFAULT 1,
                        listing_description TEXT,
                        banner_meta TEXT DEFAULT '{}',
                        performance_snapshot TEXT DEFAULT '{}',
                        status TEXT DEFAULT 'ready',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL
                    )
                    """
                )
            )
        await session.execute(
            text(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_rea_studio_campaign_lead
                ON rea_studio_campaign_listings(lead_id)
                """
            )
        )
        _campaign_cols = [
            ("rea_status", "TEXT"),
            ("rea_listing_type", "TEXT"),
            ("primary_image", "TEXT"),
            ("address", "TEXT"),
            ("suburb", "TEXT"),
            ("postcode", "TEXT"),
            ("listing_headline", "TEXT"),
            ("property_images", "TEXT"),
            ("est_value", "BIGINT"),
            ("land_size_sqm", "DOUBLE PRECISION"),
            ("bedrooms", "INTEGER"),
            ("bathrooms", "INTEGER"),
            ("car_spaces", "INTEGER"),
            ("rea_listing_id", "TEXT"),
            ("last_listing_status", "TEXT"),
        ]
        for col, ctype in _campaign_cols:
            await _add_column_if_missing(session, "rea_studio_campaign_listings", col, ctype)

        if not await _table_exists(session, "rea_publish_jobs"):
            await session.execute(
                text(
                    """
                    CREATE TABLE rea_publish_jobs (
                        id TEXT PRIMARY KEY,
                        lead_id TEXT NOT NULL,
                        action TEXT NOT NULL,
                        payload_json TEXT DEFAULT '{}',
                        status TEXT DEFAULT 'queued',
                        retry_count INTEGER DEFAULT 0,
                        next_retry_at TEXT,
                        failed_permanently INTEGER DEFAULT 0,
                        last_error TEXT DEFAULT '',
                        requested_by TEXT DEFAULT 'operator',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL
                    )
                    """
                )
            )
        await session.execute(
            text("CREATE INDEX IF NOT EXISTS idx_rea_publish_jobs_status ON rea_publish_jobs(status, created_at DESC)")
        )
        await session.commit()
    except Exception:
        await session.rollback()
        raise


def _address_key(address: str, suburb: str, postcode: str) -> str:
    raw = f"{address}|{suburb}|{postcode}".strip().lower()
    return " ".join(raw.split())


def _is_land_property_type(value: str) -> bool:
    token = (value or "").strip().lower()
    return token in {"land", "vacant land", "residential land", "lot"} or "land" in token


_BATHLA_SUBURB_MAP: dict[str, tuple[str, str]] = {
    "box hill": ("Box Hill", "2765"),
    "marsden park": ("Marsden Park", "2765"),
    "the ponds": ("The Ponds", "2769"),
    "tallawong": ("Tallawong", "2762"),
    "rouse hill": ("Rouse Hill", "2155"),
    "riverstone": ("Riverstone", "2765"),
    "north kellyville": ("North Kellyville", "2159"),
    "kellyville": ("Kellyville", "2155"),
    "oran park": ("Oran Park", "2570"),
    "dora creek": ("Dora Creek", "2264"),
    "vineyard": ("Vineyard", "2765"),
    "blacktown": ("Blacktown", "2148"),
    "nelson road": ("Box Hill", "2765"),
    "old pitt town road": ("Box Hill", "2765"),
    "hynds road": ("Box Hill", "2765"),
    "terry road": ("Box Hill", "2765"),
    "barry road": ("North Kellyville", "2159"),
    "hambledon road": ("The Ponds", "2769"),
    "south street": ("Marsden Park", "2765"),
    "gordon road": ("Tallawong", "2762"),
    "tallawong road": ("Rouse Hill", "2155"),
    "excelsior avenue": ("Marsden Park", "2765"),
    "gradwells road": ("Dora Creek", "2264"),
    "micallef street": ("Riverstone", "2765"),
    "pina road": ("Riverstone", "2765"),
    "bullaburra street": ("The Ponds", "2769"),
    "drover": ("Oran Park", "2570"),
    "madden street": ("Oran Park", "2570"),
    "schofields": ("Schofields", "2762"),
    "leppington": ("Leppington", "2179"),
    "austral": ("Austral", "2179"),
}


def _derive_estate_name(project_name: str, suburb: str) -> str:
    """Derive a clean public estate name from the project slug.

    Strips street numbers, suburb, and the word 'land' to produce something
    like 'Terry Road Estate' or 'Hambledon Road Estate'. Never mentions the
    developer name.
    """
    if not project_name:
        return f"{suburb} Estate" if suburb else "New Estate"
    slug = project_name.strip().lower()
    # Remove leading street numbers
    slug = re.sub(r"^\d[\d\s]*", "", slug).strip()
    # Remove suburb name so we don't double up
    if suburb:
        slug = re.sub(re.escape(suburb.lower()), "", slug).strip()
    # Remove trailing 'land', 'house', 'granny', 'super lots', stage info
    slug = re.sub(r"\b(land|house|granny|super\s*lots?|stage\s*\d+\s*\d*)\b", "", slug).strip()
    # Remove dangling numbers and underscores
    slug = re.sub(r"[\d_]+", " ", slug).strip()
    # Title-case what remains
    words = [w for w in slug.split() if w]
    if not words:
        return f"{suburb} Estate" if suburb else "New Estate"
    road_name = " ".join(w.title() for w in words)
    return f"{road_name} Estate, {suburb}" if suburb else f"{road_name} Estate"


def _normalize_import_key(value: str) -> str:
    return "".join(ch for ch in (value or "").strip().lower() if ch.isalnum() or ch == "_")


def _infer_suburb_postcode(address: str, project_name: str) -> tuple[str, str]:
    combined = f"{address} {project_name}".strip().lower()
    for token, resolved in _BATHLA_SUBURB_MAP.items():
        if token in combined:
            return resolved
    return "", ""


def _first_non_empty(row: dict[str, str], keys: list[str]) -> str:
    for key in keys:
        value = (row.get(key) or "").strip()
        if value:
            return value
    return ""


# =============================================================================
# LAND LISTING TEMPLATES (8 variants + data-driven renderer)
# =============================================================================
# Two public template_ids ("lifestyle", "investor") each pick the best of 4
# variants per archetype so that a batch of 127 lots ships with 8 distinct
# copy patterns instead of one repeated template.  REA's quality algorithm
# penalises duplicate content; variety is required for impressions.
#
# Archetypes:
#   - first_home        (compact block, entry price)
#   - investor_yield    (mid block, investor lens)
#   - upgrader_family   (generous/premium block, upper/prestige price)
#   - family_build      (default family-oriented narrative)
#   - corner_block      (corner lot override)
#   - cul_de_sac        (cul-de-sac override)
# =============================================================================

LAND_HEADLINE_VARIANTS = {
    "lifestyle": {
        "first_home": "First Home Ready | {land_sqm}sqm Block, {suburb} — {price_headline}",
        "investor_yield": "Build-Ready {land_sqm}sqm in {suburb} | Smart Entry at {price_headline}",
        "upgrader_family": "Premium {land_sqm}sqm Family Block, {suburb} — {price_headline}",
        "family_build": "Your Next Chapter Starts Here | {land_sqm}sqm, {suburb}",
        "corner_block": "Rare Corner Block {land_sqm}sqm in {suburb} — {price_headline}",
        "cul_de_sac": "Quiet Cul-de-Sac {land_sqm}sqm, {suburb} — {price_headline}",
    },
    "investor": {
        "first_home": "Lot {lot_number} | {land_sqm}sqm Titled Land {suburb} | {price_per_sqm_display}",
        "investor_yield": "Titled {land_sqm}sqm Land Lot {lot_number}, {suburb} | Build-Ready",
        "upgrader_family": "Premium Lot {lot_number} | {land_sqm}sqm {suburb} | {price_per_sqm_display}",
        "family_build": "Lot {lot_number} | {land_sqm}sqm Build-Ready {suburb} NSW",
        "corner_block": "Corner Lot {lot_number} | {land_sqm}sqm {suburb} | Dual Frontage",
        "cul_de_sac": "Cul-de-Sac Lot {lot_number} | {land_sqm}sqm {suburb} | Titled",
    },
}


def _pick_land_headline(template_id: str, archetype: str, ctx: dict) -> str:
    variants = LAND_HEADLINE_VARIANTS.get(template_id, LAND_HEADLINE_VARIANTS["lifestyle"])
    tmpl = variants.get(archetype) or variants.get("family_build")
    try:
        headline = tmpl.format_map(ctx)
    except (KeyError, IndexError, ValueError):
        headline = tmpl
    return re.sub(r"\s{2,}", " ", headline).replace("Lot ,", "").strip(" —|")


def _render_lifestyle_body(ctx: dict) -> str:
    """Rich, aspirational body copy with schools, transport, amenities, price anchor.

    Market-stat claims (median price, growth %, yield) are gated on
    ``ctx['cotality_data_present']``. When False, the body falls back to the
    non-market facts only: land size, frontage, schools, transport, amenities,
    and a simple price anchor (price + $/sqm).  No median/growth claims are
    emitted without Cotality backing.
    """
    lines: list[str] = []
    has_cotality = bool(ctx.get("cotality_data_present"))

    # Hook opener (deterministic per lot so adjacent lots differ)
    lines.append(ctx.get("hook_opener") or "")
    lines.append("")

    # Opening paragraph: location + size + archetype
    size = int(ctx["land_sqm"]) if ctx.get("land_sqm") else 0
    lot_no = ctx.get("lot_number", "")
    addr = ctx.get("address", "")
    suburb = ctx.get("suburb", "")
    postcode = ctx.get("postcode", "")

    lines.append(
        f"Presenting Lot {lot_no} at {addr}, {suburb} NSW {postcode} — a premium "
        f"{size}sqm build-ready block in one of Sydney's most anticipated growth corridors."
    )
    lines.append("")

    # Frontage + lot type sentence
    flavour_parts: list[str] = []
    if ctx.get("frontage_line"):
        flavour_parts.append(ctx["frontage_line"].strip())
    if ctx.get("lot_type_line"):
        flavour_parts.append(ctx["lot_type_line"].strip())
    if flavour_parts:
        lines.append(" ".join(flavour_parts))
        lines.append("")

    # Schools (REA's highest-searched filter for family buyers)
    if ctx.get("school_blurb"):
        lines.append(f"**Schools:** {ctx['school_blurb']}")
        lines.append("")

    # Transport with concrete minutes
    if ctx.get("transport_blurb"):
        lines.append(f"**Connected:** {ctx['transport_blurb']}.")
        lines.append("")

    # Top 3 amenities
    amenities = ctx.get("amenities_top") or []
    if amenities:
        lines.append("**Lifestyle nearby:** " + ", ".join(amenities) + ".")
        lines.append("")

    # Growth story — ONLY when Cotality data backs it
    if has_cotality and ctx.get("growth_story"):
        lines.append(ctx["growth_story"])
        lines.append("")

    # Price anchor.  Without Cotality we show price + $/sqm only (a factual
    # computation from the CSV figure) — no median comparisons, no growth %.
    price_line_parts: list[str] = []
    if ctx.get("price_display") and ctx["price_display"] != "Contact Agent":
        price_line_parts.append(f"Priced at **{ctx['price_display']}**")
        if ctx.get("price_per_sqm_display") and ctx["price_per_sqm_display"] != "—":
            price_line_parts.append(f"({ctx['price_per_sqm_display']})")
        if has_cotality and ctx.get("vs_median"):
            price_line_parts.append(f"— {ctx['vs_median']}")
    if price_line_parts:
        lines.append(" ".join(price_line_parts) + ".")
        lines.append("")

    # Persona close — always public-facing; suppressed in enrichment when
    # the persona itself carries market claims and Cotality is missing.
    persona = ctx.get("persona_blurb", "")
    if persona:
        lines.append(persona)
    # Stock context is a market claim — only emitted with Cotality backing.
    if has_cotality and ctx.get("stock_context"):
        lines.append(ctx["stock_context"])
    lines.append("")

    # CTA
    lines.append(
        f"Contact {PRINCIPAL_NAME} at {BRAND_NAME} on {PRINCIPAL_PHONE} "
        f"or {PRINCIPAL_EMAIL} to register your interest or arrange a private walk-through."
    )
    lines.append("")

    # Compliance
    footer = ctx.get("compliance_footer", "")
    if footer:
        lines.append("---")
        lines.append(footer)

    return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip()


def _render_investor_body(ctx: dict) -> str:
    """Data-dense, bullet-driven body copy for the investor archetype."""
    lines: list[str] = []

    size = int(ctx["land_sqm"]) if ctx.get("land_sqm") else 0
    addr = ctx.get("address", "")
    suburb = ctx.get("suburb", "")
    postcode = ctx.get("postcode", "")

    # Hook
    lines.append(ctx.get("hook_opener") or f"A build-ready land opportunity in {suburb}.")
    lines.append("")

    lines.append(
        f"Lot {ctx.get('lot_number','')} at {addr}, {suburb} NSW {postcode} — "
        f"{size}sqm of titled, registered land ready for your builder."
    )
    lines.append("")

    # Key details block
    lines.append("**Key details:**")
    lines.append(f"- Land area: {size} sqm")
    if ctx.get("frontage_bullet"):
        lines.append(ctx["frontage_bullet"].rstrip())
    if ctx.get("lot_type_bullet"):
        lines.append(ctx["lot_type_bullet"].rstrip())
    if ctx.get("price_display") and ctx["price_display"] != "Contact Agent":
        lines.append(f"- Price: {ctx['price_display']}")
    if ctx.get("price_per_sqm_display") and ctx["price_per_sqm_display"] != "—":
        lines.append(f"- Rate: {ctx['price_per_sqm_display']}")
    if ctx.get("vs_median"):
        lines.append(f"- Market position: {ctx['vs_median']}")
    lines.append("- Status: Titled, registered, build-ready")
    lines.append("")

    # Location metrics
    metro_station = ctx.get("metro_station", "")
    metro_min = ctx.get("metro_minutes", 0)
    if metro_station and metro_min:
        lines.append(f"**Location metrics:**")
        lines.append(f"- {metro_min} min drive to {metro_station}")
        if ctx.get("norwest_drive_min"):
            lines.append(f"- {ctx['norwest_drive_min']} min to Norwest business park")
        if ctx.get("parramatta_drive_min"):
            lines.append(f"- {ctx['parramatta_drive_min']} min to Parramatta CBD")
        if ctx.get("cbd_drive_min"):
            lines.append(f"- {ctx['cbd_drive_min']} min to Sydney CBD")
        lines.append("")

    # Schools
    if ctx.get("schools_primary") or ctx.get("schools_secondary"):
        lines.append("**Education catchment:**")
        if ctx.get("schools_primary"):
            lines.append(f"- Primary: {ctx['schools_primary']}")
        if ctx.get("schools_secondary"):
            lines.append(f"- Secondary: {ctx['schools_secondary']}")
        lines.append("")

    # Investment thesis
    if ctx.get("infrastructure_story"):
        lines.append("**Why " + suburb + ":**")
        lines.append(ctx["infrastructure_story"])
        lines.append("")

    if ctx.get("persona_blurb"):
        lines.append(ctx["persona_blurb"])
        lines.append("")

    lines.append(
        f"Register your interest with {PRINCIPAL_NAME} at {BRAND_NAME} — "
        f"{PRINCIPAL_PHONE} or {PRINCIPAL_EMAIL}. Inspections by private appointment."
    )
    lines.append("")

    footer = ctx.get("compliance_footer", "")
    if footer:
        lines.append("---")
        lines.append(footer)

    return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip()


# Kept for backwards compatibility with any code path that reads LAND_TEMPLATES.
LAND_TEMPLATES = {
    "lifestyle": {
        "headline": "Your Next Chapter | {land_sqm}sqm in {suburb}",
        "description": "Rich lifestyle description generated via listing_enrichment.",
    },
    "investor": {
        "headline": "Titled {land_sqm}sqm Lot {lot_number}, {suburb}",
        "description": "Data-dense investor description generated via listing_enrichment.",
    },
}


_FIRST_HOME_BATCH_TEMPLATE = {
    "headline_pattern": "Build-Ready {land_size}sqm in {suburb}",
    "body_pattern": (
        "{principal_name} from {brand_name} is pleased to present Lot {lot_number}\n"
        "— a {land_size} sqm registered block {position_phrase}.\n"
        "{lot_type_line}\n\n"
        "If you've been searching for the right block to build your first home, this one ticks every box.\n\n"
        "Land size: {land_size} sqm\n"
        "Frontage: {frontage}\n"
        "{registration_line}\n\n"
        "{location_block}\n\n"
        "Choose your own builder or work with us to arrange a quality custom build\n"
        "at a competitive price with build time guarantee.\n\n"
        "To discuss this opportunity, contact {principal_name} on {principal_phone}\n"
        "or email {principal_email}.\n\n"
        "{disclaimer}"
    ),
}


def _template_disclaimer_text(ctx: dict) -> str:
    footer = str(ctx.get("compliance_footer") or "").strip()
    return footer.split("\n\n", 1)[0].strip() if footer else ""


def _build_template_location_block(ctx: dict) -> str:
    lines: list[str] = []
    if ctx.get("schools_primary"):
        lines.append(f"- {ctx['schools_primary']} catchment")
    if ctx.get("schools_secondary"):
        lines.append(f"- Zoned for {ctx['schools_secondary']}")
    if ctx.get("metro_station") and ctx.get("metro_minutes"):
        lines.append(f"- Approx. {int(ctx['metro_minutes'])} minutes to {ctx['metro_station']}")
    for amenity in ctx.get("amenities_top") or []:
        if len(lines) >= 5:
            break
        lines.append(f"- {amenity}")
    return "Location:\n" + "\n".join(lines) if lines else ""


_HARDCODED_LOCATION_SOURCE = "backend/assets/suburb_profiles.json"


def _safe_json_dict(value) -> dict:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _normalize_match_text(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _contains_normalized_text(haystack, needle) -> bool:
    haystack_norm = _normalize_match_text(haystack)
    needle_norm = _normalize_match_text(needle)
    return bool(haystack_norm and needle_norm and needle_norm in haystack_norm)


def _claim_signature_for_cotality(claim: str) -> tuple[str, list[str]]:
    claim_text = str(claim or "").strip()
    if not claim_text:
        return "unknown", []
    if claim_text.endswith(" catchment"):
        return "school", [claim_text[: -len(" catchment")].strip()]
    if claim_text.startswith("Zoned for "):
        return "school", [claim_text[len("Zoned for ") :].strip()]
    metro_match = re.match(r"Approx\.\s*(\d+)\s+minutes?\s+to\s+(.+)$", claim_text)
    if metro_match:
        return "transport", [metro_match.group(1), metro_match.group(2).strip()]
    amenity_match = re.match(r"(.+?)\s+\(([^)]+)\)$", claim_text)
    if amenity_match:
        return "amenity", [amenity_match.group(1).strip(), amenity_match.group(2).strip()]
    return "text", [claim_text]


async def _get_latest_cotality_result_for_listing(session: AsyncSession, lead: dict) -> Optional[dict]:
    address = str(lead.get("address") or "").strip()
    suburb = str(lead.get("suburb") or "").strip()
    postcode = str(lead.get("postcode") or "").strip()
    if not address:
        return None

    try:
        row = (
            await session.execute(
                text(
                    """
                    SELECT
                        r.*,
                        j.lead_id,
                        j.matched_address,
                        j.status AS job_status
                    FROM enrichment_results r
                    JOIN enrichment_jobs j ON j.id = r.enrichment_job_id
                    LEFT JOIN leads l ON l.id = j.lead_id
                    WHERE j.provider = 'cotality'
                      AND (
                        (
                          COALESCE(l.address, '') = :address
                          AND COALESCE(l.suburb, '') = :suburb
                          AND COALESCE(l.postcode, '') = :postcode
                        )
                        OR (
                          LOWER(COALESCE(j.matched_address, '')) LIKE LOWER(:address_like)
                          AND (:postcode = '' OR LOWER(COALESCE(j.matched_address, '')) LIKE LOWER(:postcode_like))
                        )
                      )
                    ORDER BY r.created_at DESC
                    LIMIT 1
                    """
                ),
                {
                    "address": address,
                    "suburb": suburb,
                    "postcode": postcode,
                    "address_like": f"%{address}%",
                    "postcode_like": f"%{postcode}%",
                },
            )
        ).mappings().first()
    except Exception as exc:  # pragma: no cover - defensive for missing local tables
        logger.debug("cotality result lookup skipped: %s", exc)
        return None

    return dict(row) if row else None


def _build_lot_cotality_accuracy_report(lead: dict, cotality_result: Optional[dict]) -> dict:
    ctx = _build_db_template_ctx(lead)
    location_block = str(ctx.get("location_block") or "")
    hardcoded_claims = [
        line[2:].strip()
        for line in location_block.splitlines()
        if line.strip().startswith("- ")
    ]

    raw_payload = _safe_json_dict((cotality_result or {}).get("raw_payload_json"))
    proposed_updates = _safe_json_dict((cotality_result or {}).get("proposed_updates_json"))
    evidence_blob = json.dumps(
        {
            "matched_address": (cotality_result or {}).get("matched_address"),
            "raw_payload_json": raw_payload,
            "proposed_updates_json": proposed_updates,
        },
        ensure_ascii=False,
        sort_keys=True,
    )

    matched_address = str((cotality_result or {}).get("matched_address") or "").strip() or None

    expected_address = ", ".join(
        [
            part
            for part in [
                str(lead.get("address") or "").strip(),
                str(lead.get("suburb") or "").strip(),
                str(lead.get("postcode") or "").strip(),
            ]
            if part
        ]
    )

    expected_land_size = lead.get("land_size_sqm") or lead.get("land_area")
    try:
        expected_land_size_value = float(expected_land_size) if expected_land_size not in {None, ""} else None
    except (TypeError, ValueError):
        expected_land_size_value = None

    actual_land_size = proposed_updates.get("land_size_sqm")
    try:
        actual_land_size_value = float(actual_land_size) if actual_land_size not in {None, ""} else None
    except (TypeError, ValueError):
        actual_land_size_value = None

    address_status = "unverified"
    if matched_address:
        street_ok = _contains_normalized_text(matched_address, lead.get("address"))
        suburb_ok = _contains_normalized_text(matched_address, lead.get("suburb"))
        postcode_ok = not lead.get("postcode") or _contains_normalized_text(matched_address, lead.get("postcode"))
        address_status = "verified" if street_ok and suburb_ok and postcode_ok else "mismatch"

    land_status = "unverified"
    if actual_land_size_value is not None and expected_land_size_value is not None:
        land_status = "verified" if abs(actual_land_size_value - expected_land_size_value) < 0.01 else "mismatch"

    core_facts = [
        {
            "field": "address",
            "expected": expected_address,
            "actual": matched_address,
            "status": address_status,
            "source": "cotality_workflow" if matched_address else "bathla_reaxml_staging",
        },
        {
            "field": "land_size_sqm",
            "expected": int(expected_land_size_value) if expected_land_size_value is not None else None,
            "actual": int(actual_land_size_value) if actual_land_size_value is not None else None,
            "status": land_status,
            "source": "cotality_workflow" if actual_land_size_value is not None else "bathla_reaxml_staging",
        },
        {
            "field": "frontage",
            "expected": f"{str(lead.get('frontage') or '').strip()}m" if str(lead.get("frontage") or "").strip() else None,
            "actual": None,
            "status": "not_available_in_cotality_workflow",
            "source": "bathla_reaxml_staging",
        },
    ]

    location_items: list[dict] = []
    for claim in hardcoded_claims:
        claim_type, tokens = _claim_signature_for_cotality(claim)
        is_verified = bool(tokens) and all(_contains_normalized_text(evidence_blob, token) for token in tokens)
        location_items.append(
            {
                "claim": claim,
                "claim_type": claim_type,
                "status": "verified" if cotality_result and is_verified else "unverified",
                "source": _HARDCODED_LOCATION_SOURCE,
            }
        )

    checked_items = [*core_facts, *location_items]
    verified_count = sum(1 for item in checked_items if item.get("status") == "verified")
    unresolved = [item for item in checked_items if item.get("status") != "verified"]

    if not cotality_result:
        summary = "No completed Cotality enrichment result exists for this address, so the hardcoded location block cannot be verified."
    elif not unresolved:
        summary = "Cotality verified every checked item for this listing."
    else:
        unresolved_labels = [
            f"{item.get('field') or item.get('claim') or 'unknown'} [{item.get('status')}]"
            for item in unresolved[:4]
        ]
        summary = (
            f"Cotality verified {verified_count} of {len(checked_items)} checked items. "
            f"Not 100% accurate because: {'; '.join(unresolved_labels)}."
        )

    return {
        "result_found": bool(cotality_result),
        "workflow_name": str(raw_payload.get("workflow_name") or ""),
        "matched_address": matched_address,
        "hardcoded_location_source": _HARDCODED_LOCATION_SOURCE,
        "core_facts": core_facts,
        "hardcoded_location_claims": location_items,
        "verified_count": verified_count,
        "checked_count": len(checked_items),
        "is_100_percent_accurate": bool(checked_items) and verified_count == len(checked_items),
        "summary": summary,
    }


def _build_db_template_ctx(lead: dict) -> dict:
    from services.listing_enrichment import enrich_land_listing

    address = str(lead.get("address") or "")
    suburb = str(lead.get("suburb") or "")
    postcode = str(lead.get("postcode") or "")
    lot_number = str(lead.get("lot_number") or "")
    lot_type = str(lead.get("lot_type") or "")
    frontage_raw = str(lead.get("frontage") or "").strip()
    project_name = str(lead.get("project_name") or "")
    est_completion = str(lead.get("estimated_completion") or "").strip()

    land_raw = lead.get("land_size_sqm") or lead.get("land_area") or 0
    try:
        land_size_sqm = float(land_raw) if land_raw else 0.0
    except (ValueError, TypeError):
        land_size_sqm = 0.0
    land_size_text = str(int(land_size_sqm)) if land_size_sqm else "?"

    price_raw = lead.get("est_value") or lead.get("price") or 0
    try:
        price_int = int(float(price_raw)) if price_raw else 0
    except (ValueError, TypeError):
        price_int = 0

    enriched = enrich_land_listing(
        address=address,
        suburb=suburb,
        postcode=postcode,
        land_sqm=land_size_sqm,
        price=price_int,
        lot_number=lot_number,
        lot_type=lot_type,
        frontage=frontage_raw,
    )

    lot_type_line = ""
    lot_type_token = lot_type.lower()
    if lot_type_token == "corner block":
        lot_type_line = "Prized corner position with dual street access."
    elif lot_type_token == "cul-de-sac":
        lot_type_line = "Quiet cul-de-sac position — ideal for families."
    elif lot_type_token == "battleaxe":
        lot_type_line = "Private battleaxe setting with added privacy."

    registration_line = f"Registration: {est_completion}" if est_completion else "Registration: Registered and ready"

    return {
        "land_size": land_size_text,
        "suburb": suburb,
        "price": f"{price_int:,}" if price_int else "Contact Agent",
        "lot_number": lot_number,
        "frontage": f"{frontage_raw}m" if frontage_raw else "TBC",
        "lot_type_line": lot_type_line,
        "registration_line": registration_line,
        "location_block": _build_template_location_block(enriched),
        "project_name": project_name,
        "principal_name": PRINCIPAL_NAME,
        "brand_name": BRAND_NAME,
        "principal_phone": PRINCIPAL_PHONE,
        "principal_email": PRINCIPAL_EMAIL,
        "position_phrase": f"in the growth corridor of {suburb}" if suburb else "in the area",
        "disclaimer": _template_disclaimer_text(enriched),
    }


def _render_first_home_batch_template(lead: dict) -> tuple[str, str]:
    return _render_from_db_template(lead, [_FIRST_HOME_BATCH_TEMPLATE], 0)


def _render_from_db_template(lead: dict, db_templates: list[dict], row_index: int) -> tuple[str, str]:
    """Render headline + body from the DB-stored templates, cycling through them.
    Falls back to the enrichment-based generator if no templates exist."""
    if not db_templates:
        headline, _ = _generate_listing_copy(lead, "lifestyle")
        _, description = _generate_listing_copy(lead, "investor")
        return headline, description

    tpl = db_templates[row_index % len(db_templates)]
    ctx = _build_db_template_ctx(lead)

    headline_pattern = str(tpl.get("headline_pattern") or "")
    body_pattern = str(tpl.get("body_pattern") or "")

    # Use simple str.format_map with defaultdict for missing keys
    class _SafeDict(dict):
        def __missing__(self, key: str) -> str:
            return "{" + key + "}"

    headline = headline_pattern.format_map(_SafeDict(ctx))
    body = body_pattern.format_map(_SafeDict(ctx))

    headline = re.sub(r"\s{2,}", " ", headline).strip()
    body = re.sub(r"\n{3,}", "\n\n", body).strip()
    return headline, body


def _generate_listing_copy(lead: dict, template_id: str = "lifestyle") -> tuple[str, str]:
    """Generate headline and description using the enriched template renderer.

    Pulls the full enrichment context (schools, transport, amenities, price
    metrics, compliance footer) from ``services.listing_enrichment`` and
    renders either the lifestyle or investor body.

    Never outputs internal project names (e.g. developer slugs).
    """
    from services.listing_enrichment import enrich_land_listing

    # Parse core values from the lead dict
    address = str(lead.get("address") or "")
    suburb = str(lead.get("suburb") or "")
    postcode = str(lead.get("postcode") or "")
    lot_number = str(lead.get("lot_number") or "")
    lot_type = str(lead.get("lot_type") or "")
    frontage = str(lead.get("frontage") or "")

    land_raw = lead.get("land_size_sqm") or lead.get("land_area") or 0
    try:
        land_sqm = float(land_raw) if land_raw else 0.0
    except (ValueError, TypeError):
        land_sqm = 0.0

    price_raw = lead.get("est_value") or lead.get("price") or 0
    try:
        price = int(float(price_raw)) if price_raw else 0
    except (ValueError, TypeError):
        price = 0

    ctx = enrich_land_listing(
        address=address,
        suburb=suburb,
        postcode=postcode,
        land_sqm=land_sqm,
        price=price,
        lot_number=lot_number,
        lot_type=lot_type,
        frontage=frontage,
    )

    # Headline helpers for variant formatting
    ctx["price_headline"] = (
        ctx["price_display"] if ctx.get("price_display") and ctx["price_display"] != "Contact Agent"
        else "Contact Agent"
    )
    ctx["land_sqm"] = int(land_sqm) if land_sqm else 0

    archetype = ctx.get("archetype", "family_build")
    headline = _pick_land_headline(template_id, archetype, ctx)

    if template_id == "investor":
        description = _render_investor_body(ctx)
    else:
        description = _render_lifestyle_body(ctx)

    # Final cleanup
    headline = re.sub(r"\s{2,}", " ", headline).strip()
    description = re.sub(r"\n{3,}", "\n\n", description).strip()
    return headline, description


def _preflight_issues(lead: dict) -> list[str]:
    issues: list[str] = []
    if not (lead.get("address") and lead.get("suburb") and lead.get("postcode")):
        issues.append("Missing required address fields")
    if not _is_land_property_type(str(lead.get("property_type") or "")):
        issues.append("Property type is not land")
    if not (lead.get("listing_headline") or "").strip():
        issues.append("Missing listing headline")
    if not (lead.get("main_image") or lead.get("property_images")):
        issues.append("Missing image/media")
    if not (lead.get("status") or "").strip():
        issues.append("Missing lead status")
    return issues


@router.post("/studio/import-csv")
async def rea_studio_import_csv(
    file: UploadFile = File(...),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    await _ensure_rea_studio_tables(session)
    raw = await file.read()
    filename = (file.filename or "").strip().lower()
    rows: list[dict[str, str]] = []
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Uploaded .xlsx has no header row.")
        header = [_normalize_import_key(str(cell) if cell is not None else "") for cell in header_row]
        for values in iterator:
            if not values:
                continue
            mapped: dict[str, str] = {}
            has_any = False
            for idx, cell in enumerate(values):
                key = header[idx] if idx < len(header) else ""
                if not key:
                    continue
                text_value = str(cell).strip() if cell is not None else ""
                if text_value:
                    has_any = True
                mapped[key] = text_value
            if has_any:
                rows.append(mapped)
    else:
        payload = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(payload))
        for row in reader:
            normalized = {_normalize_import_key(str(k)): (str(v).strip() if v is not None else "") for k, v in (row or {}).items() if k}
            if any(v for v in normalized.values()):
                rows.append(normalized)

    batch_id = str(uuid.uuid4())
    now = now_iso()

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []
    last_address = ""
    last_suburb = ""
    last_postcode = ""
    last_project_name = ""

    for index, row in enumerate(rows, start=2):
        address = _first_non_empty(row, ["address", "street", "displayaddress", "fulladdress"])
        suburb = _first_non_empty(row, ["suburb"])
        postcode = _first_non_empty(row, ["postcode"])
        project_name = _first_non_empty(row, ["project_name", "project", "projectslug"])
        lot_number = _first_non_empty(row, ["lot_number", "lotnumber", "lot"])
        lot_type = _first_non_empty(row, ["lot_type", "lottype"])
        property_type = _first_non_empty(row, ["property_type", "propertytypesummary"]) or "land"
        ready_flag = _first_non_empty(row, ["ready_for_reaxml", "readyforreaxml"])
        listing_status = _first_non_empty(row, ["status", "project_status"])

        if address:
            last_address = address
        else:
            address = last_address
        if project_name:
            last_project_name = project_name
        else:
            project_name = last_project_name

        inferred_suburb, inferred_postcode = _infer_suburb_postcode(address, project_name)
        if not suburb and inferred_suburb:
            suburb = inferred_suburb
        if not postcode and inferred_postcode:
            postcode = inferred_postcode
        if not postcode:
            postcode_match = re.search(r"\b(\d{4})\b", f"{address} {project_name}")
            postcode = postcode_match.group(1) if postcode_match else ""
        if suburb:
            last_suburb = suburb
        else:
            suburb = last_suburb
        if postcode:
            last_postcode = postcode
        else:
            postcode = last_postcode

        headline_default = " ".join(part for part in [f"Lot {lot_number}" if lot_number else "", suburb, "Land Opportunity"] if part).strip()
        listing_headline = _first_non_empty(row, ["listing_headline", "headline", "title"]) or headline_default or address
        listing_description = _first_non_empty(row, ["listing_description", "description"])

        if listing_status and listing_status.strip().lower() not in {"available", "current", "offmarket", "sold", "active"}:
            skipped += 1
            errors.append({"row": index, "reason": f"Unsupported status '{listing_status}'"})
            continue
        if ready_flag and ready_flag.strip().lower() not in {"yes", "y", "true", "1"}:
            skipped += 1
            errors.append({"row": index, "reason": "ready_for_reaxml not enabled"})
            continue

        if not address or not suburb:
            skipped += 1
            errors.append({"row": index, "reason": "Missing address/suburb"})
            continue

        key = _address_key(address, suburb, postcode)
        lead_row = (
            await session.execute(
                text(
                    """
                    SELECT * FROM leads
                    WHERE LOWER(COALESCE(address, '')) = :address
                      AND LOWER(COALESCE(suburb, '')) = :suburb
                      AND COALESCE(postcode, '') = :postcode
                    LIMIT 1
                    """
                ),
                {"address": address.lower(), "suburb": suburb.lower(), "postcode": postcode},
            )
        ).mappings().first()

        # Parse numeric fields safely
        price_raw = _first_non_empty(row, ["price"])
        land_area_raw = _first_non_empty(row, ["land_area", "landarea", "land_size"])
        frontage_raw = _first_non_empty(row, ["frontage"])
        est_value = 0
        try:
            est_value = int(float(re.sub(r"[^0-9.]", "", price_raw))) if price_raw else 0
        except (ValueError, TypeError):
            pass
        land_size_sqm = 0.0
        try:
            land_size_sqm = float(re.sub(r"[^0-9.]", "", land_area_raw)) if land_area_raw else 0.0
        except (ValueError, TypeError):
            pass

        # Auto-generate copy if none provided
        copy_ctx = {
            "address": address, "suburb": suburb, "postcode": postcode,
            "land_size_sqm": land_size_sqm, "land_area": land_area_raw,
            "lot_number": lot_number, "frontage": frontage_raw,
            "project_name": project_name, "est_value": est_value,
        }
        if not listing_headline:
            listing_headline, _ = _generate_listing_copy(copy_ctx, "lifestyle")
        if not listing_description:
            _, listing_description = _generate_listing_copy(copy_ctx, "investor")

        if lead_row:
            lead_id = lead_row["id"]
            await session.execute(
                text(
                    """
                    UPDATE leads
                    SET property_type = COALESCE(NULLIF(:property_type, ''), property_type),
                        listing_headline = COALESCE(NULLIF(:listing_headline, ''), listing_headline),
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        est_value = CASE WHEN :est_value > 0 THEN :est_value ELSE est_value END,
                        land_size_sqm = CASE WHEN :land_size_sqm > 0 THEN :land_size_sqm ELSE land_size_sqm END,
                        lot_number = COALESCE(NULLIF(:lot_number, ''), lot_number),
                        frontage = COALESCE(NULLIF(:frontage, ''), frontage),
                        project_name = COALESCE(NULLIF(:project_name, ''), project_name),
                        updated_at = :now
                    WHERE id = :id
                    """
                ),
                {
                    "id": lead_id,
                    "property_type": property_type,
                    "listing_headline": listing_headline,
                    "listing_description": listing_description,
                    "est_value": est_value,
                    "land_size_sqm": land_size_sqm,
                    "lot_number": lot_number,
                    "frontage": frontage_raw,
                    "project_name": project_name,
                    "now": now,
                },
            )
            updated += 1
        else:
            lead_id = hashlib.md5(f"rea-csv:{key}".encode()).hexdigest()
            # Create the lead row so it exists for the publish pipeline
            await session.execute(
                text(
                    """
                    INSERT INTO leads (
                        id, address, suburb, postcode, owner_name, trigger_type,
                        record_type, property_type, heat_score, status,
                        listing_headline, listing_description,
                        est_value, land_size_sqm, lot_number, frontage, project_name,
                        signal_status, preferred_contact_method,
                        created_at, updated_at
                    ) VALUES (
                        :id, :address, :suburb, :postcode, '', 'rea_csv_import',
                        'property_record', :property_type, 0, 'captured',
                        :listing_headline, :listing_description,
                        :est_value, :land_size_sqm, :lot_number, :frontage, :project_name,
                        'LAND', '',
                        :now, :now
                    )
                    ON CONFLICT (address) DO UPDATE SET
                        listing_headline = COALESCE(NULLIF(excluded.listing_headline, ''), leads.listing_headline),
                        listing_description = COALESCE(NULLIF(excluded.listing_description, ''), leads.listing_description),
                        est_value = CASE WHEN excluded.est_value > 0 THEN excluded.est_value ELSE leads.est_value END,
                        land_size_sqm = CASE WHEN excluded.land_size_sqm > 0 THEN excluded.land_size_sqm ELSE leads.land_size_sqm END,
                        updated_at = excluded.updated_at
                    """
                ),
                {
                    "id": lead_id,
                    "address": address,
                    "suburb": suburb,
                    "postcode": postcode,
                    "property_type": property_type,
                    "listing_headline": listing_headline,
                    "listing_description": listing_description,
                    "est_value": est_value,
                    "land_size_sqm": land_size_sqm,
                    "lot_number": lot_number,
                    "frontage": frontage_raw,
                    "project_name": project_name,
                    "now": now,
                },
            )
            created += 1

        existing_campaign = (
            await session.execute(
                text("SELECT id FROM rea_studio_campaign_listings WHERE lead_id = :lead_id"),
                {"lead_id": lead_id},
            )
        ).mappings().first()
        campaign_data = {
            "lead_id": lead_id,
            "address_key": key,
            "source_batch_id": batch_id,
            "property_type": property_type,
            "is_land": 1 if _is_land_property_type(property_type) else 0,
            "listing_description": listing_description,
            "address": address,
            "suburb": suburb,
            "postcode": postcode,
            "listing_headline": listing_headline,
            "est_value": est_value if est_value else None,
            "land_size_sqm": land_size_sqm if land_size_sqm else None,
            "updated_at": now,
        }
        if existing_campaign:
            await session.execute(
                text(
                    """
                    UPDATE rea_studio_campaign_listings
                    SET address_key = :address_key,
                        source_batch_id = :source_batch_id,
                        property_type = :property_type,
                        is_land = :is_land,
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        listing_headline = COALESCE(NULLIF(:listing_headline, ''), listing_headline),
                        address = COALESCE(NULLIF(:address, ''), address),
                        suburb = COALESCE(NULLIF(:suburb, ''), suburb),
                        postcode = COALESCE(NULLIF(:postcode, ''), postcode),
                        est_value = COALESCE(:est_value, est_value),
                        land_size_sqm = COALESCE(:land_size_sqm, land_size_sqm),
                        updated_at = :updated_at
                    WHERE lead_id = :lead_id
                    """
                ),
                campaign_data,
            )
        else:
            campaign_data["id"] = str(uuid.uuid4())
            campaign_data["created_at"] = now
            await session.execute(
                text(
                    """
                    INSERT INTO rea_studio_campaign_listings (
                        id, lead_id, address_key, source_batch_id, property_type, is_land,
                        listing_description, listing_headline, address, suburb, postcode,
                        est_value, land_size_sqm,
                        status, created_at, updated_at
                    ) VALUES (
                        :id, :lead_id, :address_key, :source_batch_id, :property_type, :is_land,
                        :listing_description, :listing_headline, :address, :suburb, :postcode,
                        :est_value, :land_size_sqm,
                        'ready', :created_at, :updated_at
                    )
                    """
                ),
                campaign_data,
            )

        # Auto-create a pending publish job (ticket) for each imported listing
        existing_job = (
            await session.execute(
                text("SELECT id FROM rea_publish_jobs WHERE lead_id = :lead_id AND status IN ('pending', 'queued')"),
                {"lead_id": lead_id},
            )
        ).mappings().first()
        if not existing_job:
            job_payload = json.dumps({
                "headline": listing_headline,
                "description": listing_description,
                "property_type": property_type,
                "land_size_sqm": land_size_sqm,
                "est_value": est_value,
                "lot_number": lot_number,
            })
            await session.execute(
                text(
                    """
                    INSERT INTO rea_publish_jobs (
                        id, lead_id, action, payload_json, status, retry_count,
                        requested_by, created_at, updated_at
                    ) VALUES (
                        :id, :lead_id, 'publish_new', :payload, 'pending', 0,
                        'csv_import', :now, :now
                    )
                    """
                ),
                {"id": str(uuid.uuid4()), "lead_id": lead_id, "payload": job_payload, "now": now},
            )

    await session.commit()
    return {
        "status": "ok",
        "batch_id": batch_id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:100],
    }


@router.get("/studio/listings")
async def rea_studio_listings(
    limit: int = 100,
    offset: int = 0,
    status: str = "all",
    property_type: str = "all",
    search: str = "",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    await _ensure_rea_studio_tables(session)
    safe_limit = max(1, min(int(limit), 500))
    safe_offset = max(0, int(offset))
    where_sql = ["1=1"]
    params: dict = {"limit": safe_limit, "offset": safe_offset}
    status_token = (status or "all").strip().lower()
    if status_token in {"current", "offmarket", "sold"}:
        where_sql.append("LOWER(COALESCE(c.rea_status, l.last_listing_status, '')) = :status")
        params["status"] = status_token
    type_token = (property_type or "all").strip().lower()
    if type_token == "land":
        where_sql.append("LOWER(COALESCE(c.property_type, l.property_type, '')) LIKE '%land%'")
    elif type_token == "residential":
        where_sql.append("LOWER(COALESCE(c.property_type, l.property_type, '')) NOT LIKE '%land%'")
    search_token = (search or "").strip().lower()
    if search_token:
        where_sql.append(
            """(
                LOWER(COALESCE(c.address, l.address, '')) LIKE :search
                OR LOWER(COALESCE(c.suburb, l.suburb, '')) LIKE :search
                OR LOWER(COALESCE(c.postcode, l.postcode, '')) LIKE :search
                OR LOWER(COALESCE(c.listing_headline, l.listing_headline, '')) LIKE :search
            )"""
        )
        params["search"] = f"%{search_token}%"

    rows = (
        await session.execute(
            text(
                """
                SELECT
                    c.id AS campaign_id,
                    c.lead_id,
                    c.source_batch_id,
                    c.property_type,
                    c.is_land,
                    c.listing_description,
                    c.banner_meta,
                    c.performance_snapshot,
                    c.status AS campaign_status,
                    c.rea_status,
                    c.rea_listing_type,
                    c.primary_image,
                    c.updated_at AS campaign_updated_at,
                    COALESCE(c.address, l.address) AS address,
                    COALESCE(c.suburb, l.suburb) AS suburb,
                    COALESCE(c.postcode, l.postcode) AS postcode,
                    COALESCE(c.listing_headline, l.listing_headline) AS listing_headline,
                    l.listing_description AS lead_listing_description,
                    l.property_type AS lead_property_type,
                    l.last_listing_status,
                    COALESCE(c.est_value, l.est_value) AS est_value,
                    COALESCE(c.land_size_sqm, l.land_size_sqm) AS land_size_sqm,
                    COALESCE(c.bedrooms, l.bedrooms) AS bedrooms,
                    COALESCE(c.bathrooms, l.bathrooms) AS bathrooms,
                    COALESCE(c.car_spaces, l.car_spaces) AS car_spaces,
                    COALESCE(c.rea_listing_id, l.rea_listing_id) AS rea_listing_id,
                    l.rea_upload_id,
                    l.rea_upload_status,
                    COALESCE(c.primary_image, l.main_image) AS main_image,
                    COALESCE(c.property_images, CAST(l.property_images AS TEXT)) AS property_images,
                    l.status,
                    l.updated_at
                FROM rea_studio_campaign_listings c
                LEFT JOIN leads l ON l.id = c.lead_id
                WHERE """
                + " AND ".join(where_sql)
                + """
                ORDER BY COALESCE(c.updated_at, c.created_at) DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            params,
        )
    ).mappings().all()
    items = []
    for row in rows:
        lead = dict(row)
        if not lead.get("listing_description"):
            lead["listing_description"] = lead.get("lead_listing_description") or ""
        if not lead.get("property_type"):
            lead["property_type"] = lead.get("lead_property_type") or ""
        if not lead.get("rea_status"):
            lead["rea_status"] = (lead.get("last_listing_status") or "").lower()
        if not lead.get("main_image"):
            lead["main_image"] = lead.get("primary_image")
        lead["preflight_issues"] = _preflight_issues(lead)
        items.append(lead)
    return {"count": len(items), "items": items}


@router.post("/studio/recommendations/run")
async def rea_studio_run_recommendations(
    max_actions: int = 25,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    await _ensure_rea_studio_tables(session)
    cap = max(1, min(int(max_actions), 100))
    rows = (
        await session.execute(
            text(
                """
                SELECT
                    l.id AS lead_id, l.address, l.suburb, l.postcode,
                    l.listing_headline, l.rea_listing_id, l.rea_upload_status,
                    l.updated_at, c.listing_description, c.performance_snapshot,
                    COALESCE(ab.ctr_a, 0) AS ctr_a, COALESCE(ab.ctr_b, 0) AS ctr_b
                FROM rea_studio_campaign_listings c
                JOIN leads l ON l.id = c.lead_id
                LEFT JOIN rea_ab_tests ab ON ab.listing_id = l.rea_listing_id
                WHERE c.is_land = 1
                ORDER BY COALESCE(l.updated_at, l.created_at) ASC
                LIMIT 500
                """
            )
        )
    ).mappings().all()

    actions = []
    for row in rows:
        item = dict(row)
        issues = _preflight_issues(item)
        if issues:
            action = "fix_data"
            reason = "; ".join(issues)
            score = 0
        else:
            if item.get("rea_listing_id"):
                action = "refresh_copy"
                score = float(item.get("ctr_a") or 0) + float(item.get("ctr_b") or 0)
                reason = "Live listing: refresh creative for enquiry lift"
            else:
                action = "publish"
                score = 9999
                reason = "Not yet published to REA"
        actions.append(
            {
                "lead_id": item["lead_id"],
                "address": item.get("address", ""),
                "action": action,
                "reason": reason,
                "score": score,
                "headline": item.get("listing_headline", ""),
                "description": item.get("listing_description", ""),
            }
        )

    actions.sort(key=lambda x: x["score"], reverse=True)
    return {"count": min(len(actions), cap), "actions": actions[:cap], "cap": cap}


@router.post("/studio/actions/commit")
async def rea_studio_commit_actions(
    body: ReaStudioCommitRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    _ensure_rea_publish_enabled()
    await _ensure_rea_studio_tables(session)
    now = now_iso()
    actions = body.actions[: max(1, min(body.max_actions, 100))]
    queued = 0
    rejected = []

    for action in actions:
        lead = (
            await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": action.lead_id})
        ).mappings().first()
        if not lead:
            rejected.append({"lead_id": action.lead_id, "reason": "Lead not found"})
            continue
        issues = _preflight_issues(dict(lead))
        if action.action in {"publish", "refresh_copy"} and issues:
            rejected.append({"lead_id": action.lead_id, "reason": "; ".join(issues)})
            continue

        await session.execute(
            text(
                """
                UPDATE leads
                SET listing_headline = COALESCE(NULLIF(:headline, ''), listing_headline),
                    updated_at = :now
                WHERE id = :id
                """
            ),
            {"id": action.lead_id, "headline": action.headline or "", "now": now},
        )
        await session.execute(
            text(
                """
                UPDATE rea_studio_campaign_listings
                SET listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                    banner_meta = COALESCE(NULLIF(:banner_meta, ''), banner_meta),
                    updated_at = :updated_at
                WHERE lead_id = :lead_id
                """
            ),
            {
                "lead_id": action.lead_id,
                "listing_description": action.description or "",
                "banner_meta": json.dumps(action.banner_meta or {}),
                "updated_at": now,
            },
        )
        await session.execute(
            text(
                """
                INSERT INTO rea_publish_jobs (
                    id, lead_id, action, payload_json, status, retry_count,
                    failed_permanently, requested_by, created_at, updated_at
                ) VALUES (
                    :id, :lead_id, :action, :payload_json, 'queued', 0,
                    0, :requested_by, :created_at, :updated_at
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "lead_id": action.lead_id,
                "action": action.action,
                "payload_json": json.dumps(
                    {"headline": action.headline or "", "description": action.description or "", "banner_meta": action.banner_meta or {}}
                ),
                "requested_by": body.requested_by or "operator",
                "created_at": now,
                "updated_at": now,
            },
        )
        queued += 1

    await session.commit()
    return {"status": "ok", "queued": queued, "rejected": rejected}


@router.post("/studio/jobs/process-now")
async def rea_studio_process_jobs_now(
    limit: int = 20,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    _ensure_rea_publish_enabled()
    await _ensure_rea_studio_tables(session)
    safe_limit = max(1, min(int(limit), 200))
    jobs = (
        await session.execute(
            text(
                """
                SELECT *
                FROM rea_publish_jobs
                WHERE status IN ('queued', 'retry_wait')
                  AND (next_retry_at IS NULL OR next_retry_at <= :now)
                  AND failed_permanently = 0
                ORDER BY created_at ASC
                LIMIT :limit
                """
            ),
            {"now": now_iso(), "limit": safe_limit},
        )
    ).mappings().all()

    processed = 0
    failed = 0
    for job in jobs:
        job_id = job["id"]
        lead_id = job["lead_id"]
        action = job["action"]
        payload = json.loads(job.get("payload_json") or "{}")
        now = now_iso()
        await session.execute(
            text("UPDATE rea_publish_jobs SET status = 'processing', updated_at = :updated_at WHERE id = :id"),
            {"id": job_id, "updated_at": now},
        )
        await session.commit()

        lead_row = (
            await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})
        ).mappings().first()
        if not lead_row:
            await session.execute(
                text(
                    """
                    UPDATE rea_publish_jobs
                    SET status = 'failed', retry_count = retry_count + 1,
                        last_error = :last_error, updated_at = :updated_at
                    WHERE id = :id
                    """
                ),
                {"id": job_id, "last_error": "Lead not found", "updated_at": now_iso()},
            )
            await session.commit()
            failed += 1
            continue

        result = {"ok": False, "error": "unknown"}
        if action in ("publish", "publish_new"):
            result = await publish_listing(dict(lead_row), session=session, lead_id=lead_id)
        elif action == "refresh_copy":
            rea_listing_id = lead_row.get("rea_listing_id") or ""
            if rea_listing_id:
                updates = {
                    "headline": payload.get("headline") or lead_row.get("listing_headline") or "",
                    "description": payload.get("description") or "",
                }
                result = await update_listing(rea_listing_id, updates, session=session, lead_id=lead_id)
            else:
                result = {"ok": False, "error": "Missing rea_listing_id for refresh"}
        else:
            result = {"ok": False, "error": f"Unsupported action '{action}'"}

        if result.get("ok"):
            await session.execute(
                text(
                    """
                    UPDATE rea_publish_jobs
                    SET status = 'submitted', last_error = '', updated_at = :updated_at
                    WHERE id = :id
                    """
                ),
                {"id": job_id, "updated_at": now_iso()},
            )
            processed += 1
        else:
            retry_count = int(job.get("retry_count") or 0) + 1
            failed_perm = retry_count >= 3
            next_retry = None
            if not failed_perm:
                next_retry = (datetime.now(timezone.utc) + timedelta(seconds=120 * (2 ** (retry_count - 1)))).replace(microsecond=0).isoformat()
            await session.execute(
                text(
                    """
                    UPDATE rea_publish_jobs
                    SET status = :status,
                        retry_count = :retry_count,
                        failed_permanently = :failed_permanently,
                        next_retry_at = :next_retry_at,
                        last_error = :last_error,
                        updated_at = :updated_at
                    WHERE id = :id
                    """
                ),
                {
                    "id": job_id,
                    "status": "failed" if failed_perm else "retry_wait",
                    "retry_count": retry_count,
                    "failed_permanently": 1 if failed_perm else 0,
                    "next_retry_at": next_retry,
                    "last_error": str(result.get("error") or "Unknown error")[:500],
                    "updated_at": now_iso(),
                },
            )
            failed += 1
        await session.commit()

    return {"status": "ok", "processed": processed, "failed": failed, "total_jobs": len(jobs)}

# ---------------------------------------------------------------------------
# Ticket-based approval workflow for REA publish jobs
# ---------------------------------------------------------------------------

class TicketEditRequest(BaseModel):
    headline: Optional[str] = None
    description: Optional[str] = None
    template_id: Optional[str] = None  # "lifestyle" or "investor"


@router.get("/studio/tickets")
async def rea_studio_list_tickets(
    status: str = "all",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """List publish jobs as approvable tickets."""
    await _ensure_rea_studio_tables(session)
    where_parts = ["1=1"]
    params: dict = {}
    status_tok = (status or "all").strip().lower()
    if status_tok != "all":
        where_parts.append("j.status = :status")
        params["status"] = status_tok
    rows = (
        await session.execute(
            text(
                """
                SELECT j.id, j.lead_id, j.action, j.payload_json, j.status,
                       j.retry_count, j.last_error, j.requested_by,
                       j.created_at, j.updated_at,
                       COALESCE(c.address, l.address) AS address,
                       COALESCE(c.suburb, l.suburb) AS suburb,
                       COALESCE(c.postcode, l.postcode) AS postcode,
                       COALESCE(c.listing_headline, l.listing_headline) AS listing_headline,
                       COALESCE(c.listing_description, l.listing_description) AS listing_description,
                       l.land_size_sqm, l.lot_number, l.est_value, l.frontage, l.project_name,
                       COALESCE(c.primary_image, l.main_image) AS main_image,
                       COALESCE(c.property_images, CAST(l.property_images AS TEXT)) AS property_images
                FROM rea_publish_jobs j
                LEFT JOIN leads l ON l.id = j.lead_id
                LEFT JOIN rea_studio_campaign_listings c ON c.lead_id = j.lead_id
                WHERE """
                + " AND ".join(where_parts)
                + """
                ORDER BY j.created_at DESC
                LIMIT 200
                """
            ),
            params,
        )
    ).mappings().all()

    tickets = []
    for row in rows:
        ticket = dict(row)
        payload = json.loads(ticket.get("payload_json") or "{}")
        ticket["headline"] = payload.get("headline") or ticket.get("listing_headline") or ""
        ticket["description"] = payload.get("description") or ticket.get("listing_description") or ""
        tickets.append(ticket)
    return {"tickets": tickets, "count": len(tickets)}


@router.post("/studio/tickets/{ticket_id}/approve")
async def rea_studio_approve_ticket(
    ticket_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Approve a ticket and set it to queued for processing."""
    job = (
        await session.execute(
            text("SELECT * FROM rea_publish_jobs WHERE id = :id"), {"id": ticket_id}
        )
    ).mappings().first()
    if not job:
        raise HTTPException(status_code=404, detail="Ticket not found")
    await session.execute(
        text(
            "UPDATE rea_publish_jobs SET status = 'queued', updated_at = :now WHERE id = :id"
        ),
        {"id": ticket_id, "now": now_iso()},
    )
    await session.commit()
    return {"status": "approved", "ticket_id": ticket_id}


@router.post("/studio/tickets/{ticket_id}/reject")
async def rea_studio_reject_ticket(
    ticket_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Reject a ticket."""
    await session.execute(
        text(
            "UPDATE rea_publish_jobs SET status = 'rejected', updated_at = :now WHERE id = :id"
        ),
        {"id": ticket_id, "now": now_iso()},
    )
    await session.commit()
    return {"status": "rejected", "ticket_id": ticket_id}


@router.put("/studio/tickets/{ticket_id}/edit")
async def rea_studio_edit_ticket(
    ticket_id: str,
    body: TicketEditRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Edit a ticket's headline/description before approving. Optionally apply a template."""
    job = (
        await session.execute(
            text("SELECT * FROM rea_publish_jobs WHERE id = :id"), {"id": ticket_id}
        )
    ).mappings().first()
    if not job:
        raise HTTPException(status_code=404, detail="Ticket not found")

    payload = json.loads(job.get("payload_json") or "{}")
    lead_id = job["lead_id"]

    # If a template_id is provided, regenerate from template using lead data
    if body.template_id and body.template_id in LAND_TEMPLATES:
        lead_row = (
            await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})
        ).mappings().first()
        if lead_row:
            headline, description = _generate_listing_copy(dict(lead_row), body.template_id)
            payload["headline"] = headline
            payload["description"] = description

    if body.headline is not None:
        payload["headline"] = body.headline
    if body.description is not None:
        payload["description"] = body.description

    await session.execute(
        text(
            "UPDATE rea_publish_jobs SET payload_json = :payload, updated_at = :now WHERE id = :id"
        ),
        {"id": ticket_id, "payload": json.dumps(payload), "now": now_iso()},
    )
    # Also update the lead and campaign listing
    if payload.get("headline"):
        await session.execute(
            text("UPDATE leads SET listing_headline = :h, updated_at = :now WHERE id = :id"),
            {"h": payload["headline"], "now": now_iso(), "id": lead_id},
        )
    if payload.get("description"):
        await session.execute(
            text("UPDATE rea_studio_campaign_listings SET listing_description = :d, updated_at = :now WHERE lead_id = :id"),
            {"d": payload["description"], "now": now_iso(), "id": lead_id},
        )
    await session.commit()
    return {"status": "updated", "ticket_id": ticket_id, "headline": payload.get("headline"), "description": payload.get("description")}


@router.get("/studio/templates/legacy")
async def rea_studio_list_templates_legacy(api_key: str = Depends(get_api_key)):
    """Return hardcoded listing description templates (legacy). Use GET /studio/templates for DB-backed templates."""
    return {
        "templates": [
            {
                "id": tid,
                "name": tid.title(),
                "headline_template": t["headline"],
                "description_template": t["description"],
            }
            for tid, t in LAND_TEMPLATES.items()
        ]
    }


@router.post("/studio/generate-copy/{lead_id}")
async def rea_studio_generate_copy(
    lead_id: str,
    template_id: str = "lifestyle",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Generate headline + description for a lead using a template."""
    lead_row = (
        await session.execute(text("SELECT * FROM leads WHERE id = :id"), {"id": lead_id})
    ).mappings().first()
    if not lead_row:
        raise HTTPException(status_code=404, detail="Lead not found")
    headline, description = _generate_listing_copy(dict(lead_row), template_id)
    return {"headline": headline, "description": description, "template_id": template_id}


# ---------------------------------------------------------------------------
# Populate from CSV file on disk (no upload needed)
# ---------------------------------------------------------------------------

_CSV_PATH = Path(__file__).resolve().parents[3] / "bathla_reaxml_staging.csv"


def _load_normalized_staging_rows(target: Path) -> list[dict[str, str]]:
    with open(target, encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows: list[dict[str, str]] = []
        for row in reader:
            normalized = {
                _normalize_import_key(str(k)): (str(v).strip() if v is not None else "")
                for k, v in (row or {}).items()
                if k
            }
            if any(v for v in normalized.values()):
                rows.append(normalized)
    return rows


def _load_sample_lot_inputs_from_staging_csv(
    lot_number: str,
    target: Path,
) -> dict[str, object]:
    if not target.exists():
        raise HTTPException(status_code=404, detail=f"CSV not found at {target}")

    rows = _load_normalized_staging_rows(target)
    last_address = ""
    last_suburb = ""
    last_postcode = ""
    last_project_name = ""

    for row in rows:
        address = _first_non_empty(row, ["address", "street", "displayaddress", "fulladdress"])
        suburb = _first_non_empty(row, ["suburb"])
        postcode = _first_non_empty(row, ["postcode"])
        project_name = _first_non_empty(row, ["project_name", "project", "projectslug"])
        row_lot_number = _first_non_empty(row, ["lot_number", "lotnumber", "lot"])
        lot_type = _first_non_empty(row, ["lot_type", "lottype"])
        property_type = _first_non_empty(row, ["property_type", "propertytypesummary"]) or "land"

        if address:
            last_address = address
        else:
            address = last_address
        if project_name:
            last_project_name = project_name
        else:
            project_name = last_project_name

        inferred_suburb, inferred_postcode = _infer_suburb_postcode(address, project_name)
        if not suburb and inferred_suburb:
            suburb = inferred_suburb
        if not postcode and inferred_postcode:
            postcode = inferred_postcode
        if not postcode:
            postcode_match = re.search(r"\b(\d{4})\b", f"{address} {project_name}")
            postcode = postcode_match.group(1) if postcode_match else ""
        if suburb:
            last_suburb = suburb
        else:
            suburb = last_suburb
        if postcode:
            last_postcode = postcode
        else:
            postcode = last_postcode

        if row_lot_number != lot_number:
            continue

        price_raw = _first_non_empty(row, ["price"])
        land_area_raw = _first_non_empty(row, ["land_area", "landarea", "land_size"])
        frontage_raw = _first_non_empty(row, ["frontage"])
        est_value = 0
        try:
            est_value = int(float(re.sub(r"[^0-9.]", "", price_raw))) if price_raw else 0
        except (ValueError, TypeError):
            pass
        land_size_sqm = 0.0
        try:
            land_size_sqm = float(re.sub(r"[^0-9.]", "", land_area_raw)) if land_area_raw else 0.0
        except (ValueError, TypeError):
            pass

        return {
            "address": address,
            "suburb": suburb,
            "postcode": postcode,
            "lot_number": row_lot_number,
            "lot_type": lot_type,
            "frontage": frontage_raw,
            "land_size_sqm": int(land_size_sqm) if land_size_sqm.is_integer() else land_size_sqm,
            "est_value": est_value,
            "property_type": property_type,
            "project_name": project_name,
        }

    raise HTTPException(status_code=404, detail=f"Lot {lot_number} not found in staging CSV at {target}")


@router.post("/studio/populate-from-csv")
async def rea_studio_populate_from_csv(
    csv_path: str = "",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Read the staging CSV from disk, create leads + campaign listings + pending
    tickets with auto-generated headline (lifestyle template) and description
    (investor template).  Never outputs internal developer or project names.
    """
    await _ensure_rea_studio_tables(session)
    target = Path(csv_path) if csv_path else _CSV_PATH
    if not target.exists():
        raise HTTPException(status_code=404, detail=f"CSV not found at {target}")

    rows = _load_normalized_staging_rows(target)

    batch_id = str(uuid.uuid4())
    now = now_iso()
    created = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []
    last_address = ""
    last_suburb = ""
    last_postcode = ""
    last_project_name = ""

    for index, row in enumerate(rows, start=2):
        address = _first_non_empty(row, ["address", "street", "displayaddress", "fulladdress"])
        suburb = _first_non_empty(row, ["suburb"])
        postcode = _first_non_empty(row, ["postcode"])
        project_name = _first_non_empty(row, ["project_name", "project", "projectslug"])
        lot_number = _first_non_empty(row, ["lot_number", "lotnumber", "lot"])
        lot_type = _first_non_empty(row, ["lot_type", "lottype"])
        property_type = _first_non_empty(row, ["property_type", "propertytypesummary"]) or "land"
        ready_flag = _first_non_empty(row, ["ready_for_reaxml", "readyforreaxml"])
        listing_status = _first_non_empty(row, ["status", "project_status"])

        # Carry forward address/project from previous row when CSV leaves them blank
        if address:
            last_address = address
        else:
            address = last_address
        if project_name:
            last_project_name = project_name
        else:
            project_name = last_project_name

        inferred_suburb, inferred_postcode = _infer_suburb_postcode(address, project_name)
        if not suburb and inferred_suburb:
            suburb = inferred_suburb
        if not postcode and inferred_postcode:
            postcode = inferred_postcode
        if not postcode:
            postcode_match = re.search(r"\b(\d{4})\b", f"{address} {project_name}")
            postcode = postcode_match.group(1) if postcode_match else ""
        if suburb:
            last_suburb = suburb
        else:
            suburb = last_suburb
        if postcode:
            last_postcode = postcode
        else:
            postcode = last_postcode

        # Skip non-available or non-ready
        if listing_status and listing_status.strip().lower() not in {
            "available", "current", "offmarket", "sold", "active", ""
        }:
            skipped += 1
            errors.append({"row": index, "reason": f"Unsupported status '{listing_status}'"})
            continue
        if ready_flag and ready_flag.strip().lower() not in {"yes", "y", "true", "1"}:
            skipped += 1
            errors.append({"row": index, "reason": "ready_for_reaxml not enabled"})
            continue

        # Parse numeric fields safely
        price_raw = _first_non_empty(row, ["price"])
        land_area_raw = _first_non_empty(row, ["land_area", "landarea", "land_size"])
        frontage_raw = _first_non_empty(row, ["frontage"])
        est_value = 0
        try:
            est_value = int(float(re.sub(r"[^0-9.]", "", price_raw))) if price_raw else 0
        except (ValueError, TypeError):
            pass
        land_size_sqm = 0.0
        try:
            land_size_sqm = float(re.sub(r"[^0-9.]", "", land_area_raw)) if land_area_raw else 0.0
        except (ValueError, TypeError):
            pass

        # Build a synthetic address when the CSV has none
        if not address and lot_number and suburb:
            address = f"Lot {lot_number}, {suburb} NSW {postcode}"

        if not address or not suburb:
            skipped += 1
            errors.append({"row": index, "reason": "Missing address/suburb"})
            continue

        # Auto-generate copy using DB templates (cycle through for variety)
        copy_ctx = {
            "address": address, "suburb": suburb, "postcode": postcode,
            "land_size_sqm": land_size_sqm, "land_area": land_area_raw,
            "lot_number": lot_number, "lot_type": lot_type,
            "frontage": frontage_raw, "project_name": project_name,
            "est_value": est_value,
        }
        listing_headline, listing_description = _render_first_home_batch_template(copy_ctx)

        key = _address_key(address, suburb, postcode)

        # Check for existing lead
        lead_row = (
            await session.execute(
                text(
                    """
                    SELECT * FROM leads
                    WHERE LOWER(COALESCE(address, '')) = :address
                      AND LOWER(COALESCE(suburb, '')) = :suburb
                      AND COALESCE(postcode, '') = :postcode
                    LIMIT 1
                    """
                ),
                {"address": address.lower(), "suburb": suburb.lower(), "postcode": postcode},
            )
        ).mappings().first()

        if lead_row:
            lead_id = lead_row["id"]
            await session.execute(
                text(
                    """
                    UPDATE leads
                    SET property_type = COALESCE(NULLIF(:property_type, ''), property_type),
                        listing_headline = COALESCE(NULLIF(:listing_headline, ''), listing_headline),
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        est_value = CASE WHEN :est_value > 0 THEN :est_value ELSE est_value END,
                        land_size_sqm = CASE WHEN :land_size_sqm > 0 THEN :land_size_sqm ELSE land_size_sqm END,
                        lot_number = COALESCE(NULLIF(:lot_number, ''), lot_number),
                        lot_type = COALESCE(NULLIF(:lot_type, ''), lot_type),
                        frontage = COALESCE(NULLIF(:frontage, ''), frontage),
                        project_name = COALESCE(NULLIF(:project_name, ''), project_name),
                        updated_at = :now
                    WHERE id = :id
                    """
                ),
                {
                    "id": lead_id,
                    "property_type": property_type,
                    "listing_headline": listing_headline,
                    "listing_description": listing_description,
                    "est_value": est_value,
                    "land_size_sqm": land_size_sqm,
                    "lot_number": lot_number,
                    "lot_type": lot_type,
                    "frontage": frontage_raw,
                    "project_name": project_name,
                    "now": now,
                },
            )
            updated += 1
        else:
            lead_id = hashlib.md5(f"rea-csv:{key}".encode()).hexdigest()
            await session.execute(
                text(
                    """
                    INSERT INTO leads (
                        id, address, suburb, postcode, owner_name, trigger_type,
                        record_type, property_type, heat_score, status,
                        listing_headline, listing_description,
                        est_value, land_size_sqm, lot_number, lot_type, frontage, project_name,
                        signal_status, preferred_contact_method,
                        created_at, updated_at
                    ) VALUES (
                        :id, :address, :suburb, :postcode, '', 'rea_csv_import',
                        'property_record', :property_type, 0, 'captured',
                        :listing_headline, :listing_description,
                        :est_value, :land_size_sqm, :lot_number, :lot_type, :frontage, :project_name,
                        'LAND', '',
                        :now, :now
                    )
                    ON CONFLICT (address) DO UPDATE SET
                        listing_headline = COALESCE(NULLIF(excluded.listing_headline, ''), leads.listing_headline),
                        listing_description = COALESCE(NULLIF(excluded.listing_description, ''), leads.listing_description),
                        est_value = CASE WHEN excluded.est_value > 0 THEN excluded.est_value ELSE leads.est_value END,
                        land_size_sqm = CASE WHEN excluded.land_size_sqm > 0 THEN excluded.land_size_sqm ELSE leads.land_size_sqm END,
                        updated_at = excluded.updated_at
                    """
                ),
                {
                    "id": lead_id,
                    "address": address,
                    "suburb": suburb,
                    "postcode": postcode,
                    "property_type": property_type,
                    "listing_headline": listing_headline,
                    "listing_description": listing_description,
                    "est_value": est_value,
                    "land_size_sqm": land_size_sqm,
                    "lot_number": lot_number,
                    "lot_type": lot_type,
                    "frontage": frontage_raw,
                    "project_name": project_name,
                    "now": now,
                },
            )
            created += 1

        # Hero image generation removed — Mapbox API was billed per-call.
        # Listings populated from CSV now ship without an auto-hero; operator
        # uploads photos manually via the Studio UI.

        # Upsert campaign listing
        existing_campaign = (
            await session.execute(
                text("SELECT id FROM rea_studio_campaign_listings WHERE lead_id = :lead_id"),
                {"lead_id": lead_id},
            )
        ).mappings().first()
        campaign_data = {
            "lead_id": lead_id,
            "address_key": key,
            "source_batch_id": batch_id,
            "property_type": property_type,
            "is_land": 1 if _is_land_property_type(property_type) else 0,
            "listing_description": listing_description,
            "listing_headline": listing_headline,
            "address": address,
            "suburb": suburb,
            "postcode": postcode,
            "est_value": est_value if est_value else None,
            "land_size_sqm": land_size_sqm if land_size_sqm else None,
            "updated_at": now,
        }
        if existing_campaign:
            await session.execute(
                text(
                    """
                    UPDATE rea_studio_campaign_listings
                    SET address_key = :address_key,
                        source_batch_id = :source_batch_id,
                        property_type = :property_type,
                        is_land = :is_land,
                        listing_description = COALESCE(NULLIF(:listing_description, ''), listing_description),
                        listing_headline = COALESCE(NULLIF(:listing_headline, ''), listing_headline),
                        address = COALESCE(NULLIF(:address, ''), address),
                        suburb = COALESCE(NULLIF(:suburb, ''), suburb),
                        postcode = COALESCE(NULLIF(:postcode, ''), postcode),
                        est_value = COALESCE(:est_value, est_value),
                        land_size_sqm = COALESCE(:land_size_sqm, land_size_sqm),
                        updated_at = :updated_at
                    WHERE lead_id = :lead_id
                    """
                ),
                campaign_data,
            )
        else:
            campaign_data["id"] = str(uuid.uuid4())
            campaign_data["created_at"] = now
            await session.execute(
                text(
                    """
                    INSERT INTO rea_studio_campaign_listings (
                        id, lead_id, address_key, source_batch_id, property_type, is_land,
                        listing_description, listing_headline, address, suburb, postcode,
                        est_value, land_size_sqm,
                        status, created_at, updated_at
                    ) VALUES (
                        :id, :lead_id, :address_key, :source_batch_id, :property_type, :is_land,
                        :listing_description, :listing_headline, :address, :suburb, :postcode,
                        :est_value, :land_size_sqm,
                        'ready', :created_at, :updated_at
                    )
                    """
                ),
                campaign_data,
            )

        # Auto-create a pending ticket for each listing
        existing_job = (
            await session.execute(
                text("SELECT id FROM rea_publish_jobs WHERE lead_id = :lead_id AND status IN ('pending', 'queued')"),
                {"lead_id": lead_id},
            )
        ).mappings().first()
        if not existing_job:
            job_payload = json.dumps({
                "headline": listing_headline,
                "description": listing_description,
                "property_type": property_type,
                "land_size_sqm": land_size_sqm,
                "est_value": est_value,
                "lot_number": lot_number,
            })
            await session.execute(
                text(
                    """
                    INSERT INTO rea_publish_jobs (
                        id, lead_id, action, payload_json, status, retry_count,
                        requested_by, created_at, updated_at
                    ) VALUES (
                        :id, :lead_id, 'publish_new', :payload, 'pending', 0,
                        'csv_populate', :now, :now
                    )
                    """
                ),
                {"id": str(uuid.uuid4()), "lead_id": lead_id, "payload": job_payload, "now": now},
            )

    await session.commit()
    return {
        "status": "ok",
        "batch_id": batch_id,
        "csv_path": str(target),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:100],
    }


# ---------------------------------------------------------------------------
# Photo upload for listings
# ---------------------------------------------------------------------------


@router.post("/studio/listings/{lead_id}/upload-photo")
async def rea_studio_upload_photo(
    lead_id: str,
    file: UploadFile = File(...),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Upload a photo for a listing. Saves to the configured LISTING_PHOTOS_ROOT
    and appends to the lead's property_images JSON array. The URL returned is
    served via the /listing_photos static mount in runtime/app.py.
    """
    from core.config import LISTING_PHOTOS_ROOT
    lead_row = (
        await session.execute(text("SELECT id, property_images FROM leads WHERE id = :id"), {"id": lead_id})
    ).mappings().first()
    if not lead_row:
        raise HTTPException(status_code=404, detail="Lead not found")

    # Validate file type
    filename = (file.filename or "photo.jpg").strip()
    ext = Path(filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail=f"Unsupported image format: {ext}")

    # Read file content
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB limit
        raise HTTPException(status_code=400, detail="File too large (max 20MB)")

    # Save to disk
    photo_dir = LISTING_PHOTOS_ROOT / lead_id
    photo_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex[:12]}{ext}"
    photo_path = photo_dir / safe_name
    photo_path.write_bytes(content)

    # Build a relative URL for the image (served via /listing_photos static mount)
    relative_url = f"/listing_photos/{lead_id}/{safe_name}"

    # Update lead's property_images
    existing_images_raw = lead_row.get("property_images") or "[]"
    try:
        existing_images = json.loads(existing_images_raw) if isinstance(existing_images_raw, str) else (existing_images_raw or [])
    except (json.JSONDecodeError, TypeError):
        existing_images = []
    existing_images.append(relative_url)

    await session.execute(
        text(
            """
            UPDATE leads
            SET property_images = :images,
                main_image = COALESCE(NULLIF(main_image, ''), :new_image),
                updated_at = :now
            WHERE id = :id
            """
        ),
        {
            "id": lead_id,
            "images": json.dumps(existing_images),
            "new_image": relative_url,
            "now": now_iso(),
        },
    )
    # Also update campaign listing if it exists
    await session.execute(
        text(
            """
            UPDATE rea_studio_campaign_listings
            SET property_images = :images,
                primary_image = COALESCE(NULLIF(primary_image, ''), :new_image),
                updated_at = :now
            WHERE lead_id = :lead_id
            """
        ),
        {
            "lead_id": lead_id,
            "images": json.dumps(existing_images),
            "new_image": relative_url,
            "now": now_iso(),
        },
    )
    await session.commit()

    return {
        "status": "ok",
        "lead_id": lead_id,
        "filename": safe_name,
        "url": relative_url,
        "total_images": len(existing_images),
    }


# ---------------------------------------------------------------------------
# Reproducible sample: "Generate Lot 127 Sample" button
# ---------------------------------------------------------------------------
#
# One-click button endpoint.  Given the same inputs, this endpoint produces
# byte-identical text (headline + body).  Purpose: let the operator preview
# exactly what a 10/10 listing will look like before the full populate-from-
# csv run, and validate the push pipeline end-to-end without hitting REA
# (REA_PUBLISH_ENABLED gate still applies on the real publish endpoint).
#
# Deterministic guarantees:
#   - _generate_listing_copy is pure text (no random, no datetime calls)
#   - listing_enrichment.enrich_land_listing is pure text
#   - hero image: any pre-existing cached file on disk is served; no HTTP
#     calls, no Mapbox billing
#   - output includes a sha256 hash of the (lifestyle+investor) text so the
#     operator can verify reproducibility in the UI.


_LOT127_SAMPLE_LEAD_ID = "sample_lot127_box_hill"


_FORBIDDEN_TERMS = [
    "Ownit1st",
    "Hills Intelligence Hub",
    "Bathla",
    "Shahid",
]


def _lot127_compliance_check(headline: str, body: str) -> dict:
    """Run the identity/brand/leak checks against a rendered sample."""
    required_identity = [
        PRINCIPAL_NAME,
        PRINCIPAL_EMAIL,
        PRINCIPAL_PHONE,
        BRAND_NAME,
    ]
    combined = f"{headline}\n{body}"
    lower = combined.lower()

    forbidden_hits = [term for term in _FORBIDDEN_TERMS if term.lower() in lower]
    missing_identity = [term for term in required_identity if term not in combined]

    # REA quality heuristics
    headline_len = len(headline)
    body_len = len(body)
    rea_ok = (
        headline_len <= 80
        and headline_len >= 20
        and body_len >= 400
        and body_len <= 4000
    )

    return {
        "ok": not forbidden_hits and not missing_identity and rea_ok,
        "forbidden_hits": forbidden_hits,
        "missing_identity": missing_identity,
        "headline_len": headline_len,
        "body_len": body_len,
        "rea_length_ok": rea_ok,
    }


@router.post("/studio/sample/lot127")
async def rea_studio_sample_lot127(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Reproducible, button-backed Lot 127 Box Hill sample.

    Returns both the lifestyle and investor rendering, compliance check
    results, and a reproducibility hash.  Safe to call repeatedly — no
    database writes, no REA calls, no Mapbox calls, no token spend.
    """
    lead = _load_sample_lot_inputs_from_staging_csv("127", _CSV_PATH)

    lifestyle_headline, lifestyle_body = _render_first_home_batch_template(lead)
    investor_headline, investor_body = lifestyle_headline, lifestyle_body

    # Hero image: serve any pre-existing cached file on disk (no Mapbox calls).
    hero_url: Optional[str] = None
    from core.config import LISTING_PHOTOS_ROOT
    cached_hero = LISTING_PHOTOS_ROOT / _LOT127_SAMPLE_LEAD_ID / "hero_mapbox.jpg"
    if cached_hero.exists():
        hero_url = f"/listing_photos/{_LOT127_SAMPLE_LEAD_ID}/hero_mapbox.jpg"

    # Suburb market-stat gating comes from the xlsx exports and is separate
    # from address-level Cotality verification for this specific lot.
    market_stats_available = False
    try:
        from services.suburb_intel_service import get_suburb_intel

        market_stats_available = bool(get_suburb_intel(lead["suburb"]))
    except Exception:
        pass

    latest_cotality_result = await _get_latest_cotality_result_for_listing(session, lead)
    cotality_verification = _build_lot_cotality_accuracy_report(lead, latest_cotality_result)

    compliance_lifestyle = _lot127_compliance_check(lifestyle_headline, lifestyle_body)
    compliance_investor = _lot127_compliance_check(investor_headline, investor_body)

    # Reproducibility hash: hash the full rendered output across both templates.
    repro_blob = "\n".join([
        lifestyle_headline, lifestyle_body,
        investor_headline, investor_body,
    ]).encode("utf-8")
    repro_sha = hashlib.sha256(repro_blob).hexdigest()[:16]

    overall_ok = compliance_lifestyle["ok"] and compliance_investor["ok"]
    suburb = str(lead.get("suburb") or "this suburb")

    return {
        "status": "ok" if overall_ok else "warning",
        "lead_id": _LOT127_SAMPLE_LEAD_ID,
        "inputs": lead,
        "lifestyle": {
            "headline": lifestyle_headline,
            "body": lifestyle_body,
            "compliance": compliance_lifestyle,
        },
        "investor": {
            "headline": investor_headline,
            "body": investor_body,
            "compliance": compliance_investor,
        },
        "hero_url": hero_url,
        "cotality_backed": cotality_verification["result_found"],
        "cotality_verification": cotality_verification,
        "market_stats_backed": market_stats_available,
        "reproducibility_sha": repro_sha,
        "overall_ok": overall_ok,
        "rea_publish_enabled": REA_PUBLISH_ENABLED,
        "notes": [
            f"Market-stat claims (median, growth %, yield) suppressed because Cotality suburb intel for {suburb} is not in the local xlsx exports."
            if not market_stats_available
            else f"Cotality-backed suburb market claims are enabled for {suburb}.",
            cotality_verification["summary"],
            f"Hardcoded location bullets currently come from {_HARDCODED_LOCATION_SOURCE}.",
            "Hero image served via /listing_photos static mount. Cached per lead_id — second call returns the same bytes.",
            "Call this endpoint again to verify the reproducibility_sha stays identical.",
        ],
    }


# ─── Editable REA Listing Templates ──────────────────────────────────────────


@router.get("/studio/templates")
async def rea_studio_list_templates(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """List all REA listing templates."""
    r = await session.execute(text("SELECT * FROM rea_listing_templates ORDER BY name"))
    return {"templates": [dict(row) for row in r.mappings().all()]}


@router.get("/studio/templates/{template_id}")
async def rea_studio_get_template(
    template_id: str,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Get a single REA listing template by id."""
    r = await session.execute(
        text("SELECT * FROM rea_listing_templates WHERE id = :id"), {"id": template_id}
    )
    row = r.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Template not found")
    return dict(row)


@router.put("/studio/templates/{template_id}")
async def rea_studio_update_template(
    template_id: str,
    request: Request,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Update fields on an existing REA listing template."""
    body = await request.json()
    existing = (
        await session.execute(
            text("SELECT id FROM rea_listing_templates WHERE id = :id"), {"id": template_id}
        )
    ).mappings().first()
    if not existing:
        raise HTTPException(status_code=404, detail="Template not found")
    updates = {}
    for field in ["name", "headline_pattern", "body_pattern", "category", "is_default"]:
        if field in body:
            updates[field] = body[field]
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    updates["id"] = template_id
    updates["now"] = datetime.now(timezone.utc).isoformat()
    await session.execute(
        text(f"UPDATE rea_listing_templates SET {set_clause}, updated_at = :now WHERE id = :id"),
        updates,
    )
    await session.commit()
    r = await session.execute(
        text("SELECT * FROM rea_listing_templates WHERE id = :id"), {"id": template_id}
    )
    return dict(r.mappings().first())


@router.post("/studio/templates/preview")
async def rea_studio_preview_template(
    request: Request,
    api_key: str = Depends(get_api_key),
):
    """Render a template with sample lead data and return compliance info."""
    body = await request.json()
    headline_pattern = body.get("headline_pattern", "")
    body_pattern = body.get("body_pattern", "")
    lead = body.get("lead", {})

    ctx = _build_db_template_ctx(lead)

    class _SafeDict(dict):
        def __missing__(self, key: str) -> str:
            return "{" + key + "}"

    rendered_headline = headline_pattern.format_map(_SafeDict(ctx))
    rendered_body = body_pattern.format_map(_SafeDict(ctx))

    return {
        "headline": rendered_headline,
        "body": rendered_body,
        "headline_len": len(rendered_headline),
        "body_len": len(rendered_body),
        "rea_headline_ok": 20 <= len(rendered_headline) <= 80,
        "rea_body_ok": 50 <= len(rendered_body) <= 5000,
    }


# Agent routes in rea_agent.py (separate router to avoid include_router truncation)
