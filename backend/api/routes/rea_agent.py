"""
REA Listing Agent API routes - ATLAS agent for self-improving listing management.
"""

import hashlib
import io
import logging
import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_session
from core.security import get_api_key
from core.settings import get_settings

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/rea/agent", tags=["REA Listing Agent - ATLAS"])
_LAND_FILTER_SQL = "(LOWER(COALESCE(property_type, '')) = 'land' OR LOWER(COALESCE(trigger_type, '')) IN ('bathla_land', 'rea_feed'))"
LISTING_COPY_DISCLAIMER = (
    "Disclaimer: Information is indicative only, subject to change without notice, "
    "and should not be relied on as legal or financial advice. Buyers must make their own enquiries."
)


class EditListingRequest(BaseModel):
    headline: str | None = None
    description: str | None = None


class PushPlanRequest(BaseModel):
    daily_limit: int = 15


class RefreshPlanRequest(BaseModel):
    limit: int = 10


class ExecuteRequest(BaseModel):
    lead_ids: list[str]
    confirm_publish: bool = False


class BulkCopyRequest(BaseModel):
    limit: int = 151
    include_live: bool = True
    enforce_disclaimer: bool = True


@router.get("/analyze")
async def agent_analyze(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import analyze_portfolio

    return await analyze_portfolio(session)


@router.post("/push-plan")
async def agent_push_plan(
    body: PushPlanRequest = PushPlanRequest(),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import generate_push_plan

    return await generate_push_plan(session, daily_limit=body.daily_limit)


@router.post("/refresh-plan")
async def agent_refresh_plan(
    body: RefreshPlanRequest = RefreshPlanRequest(),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import generate_refresh_plan

    return await generate_refresh_plan(session, limit=body.limit)


@router.post("/execute-push")
async def agent_execute_push(
    body: ExecuteRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import execute_push

    settings = get_settings()
    if not body.lead_ids:
        raise HTTPException(status_code=400, detail="No lead_ids provided")
    if not settings.rea_auto_publish_enabled:
        raise HTTPException(
            status_code=403,
            detail="REA auto-publish is disabled by policy. Set REA_AUTO_PUBLISH_ENABLED=true to enable.",
        )
    if settings.rea_require_explicit_push_confirmation and not body.confirm_publish:
        raise HTTPException(
            status_code=400,
            detail="Explicit publish confirmation is required for REA push operations.",
        )
    return await execute_push(session, body.lead_ids)


@router.post("/execute-refresh")
async def agent_execute_refresh(
    body: ExecuteRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import execute_refresh

    if not body.lead_ids:
        raise HTTPException(status_code=400, detail="No lead_ids provided")
    return await execute_refresh(session, body.lead_ids)


@router.post("/pull-performance")
async def agent_pull_performance(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import pull_performance

    return await pull_performance(session)


@router.post("/self-improve")
async def agent_self_improve(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from hermes.workers.rea_listing_worker import self_improve

    return await self_improve(session)


@router.get("/listings")
async def agent_listings(
    status: str = "all",
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    # Keep REA Studio cards usable even when upstream imports omitted copy.
    await session.execute(
        text(
            f"""
            UPDATE leads
            SET
                listing_headline = CASE
                    WHEN COALESCE(TRIM(listing_headline), '') = ''
                        THEN COALESCE(NULLIF(TRIM(address), ''), 'Land Opportunity')
                    ELSE listing_headline
                END,
                listing_description = CASE
                    WHEN COALESCE(TRIM(listing_description), '') = '' THEN
                        CASE
                            WHEN COALESCE(TRIM(suburb), '') <> '' THEN 'Land opportunity in ' || suburb || '. Contact agent for full details. ' || :disclaimer
                            ELSE 'Land opportunity. Contact agent for full details. ' || :disclaimer
                        END
                    ELSE listing_description
                END
            WHERE {_LAND_FILTER_SQL}
              AND (
                COALESCE(TRIM(listing_headline), '') = ''
                OR COALESCE(TRIM(listing_description), '') = ''
              )
            """
        )
        ,
        {"disclaimer": LISTING_COPY_DISCLAIMER}
    )
    await session.commit()

    conditions = [_LAND_FILTER_SQL]
    if status == "live":
        conditions.append("COALESCE(rea_listing_id, '') <> ''")
    elif status == "pending":
        conditions.append("COALESCE(rea_upload_id, '') <> '' AND COALESCE(rea_listing_id, '') = ''")
    elif status == "ready":
        conditions.append("COALESCE(rea_listing_id, '') = '' AND COALESCE(rea_upload_id, '') = ''")

    where = " AND ".join(conditions)
    rows = (
        await session.execute(
            text(
                f"""
        SELECT id, address, suburb, postcode, land_size_sqm, estimated_value_mid,
               listing_headline, listing_description, main_image, property_type,
               rea_listing_id, rea_upload_id, rea_upload_status,
               rea_title_variant, rea_desc_variant, rea_last_edit_at,
               rea_views, rea_enquiries, lot_number, lot_type, frontage,
               project_name, updated_at, created_at
        FROM leads
        WHERE {where}
        ORDER BY suburb, address
    """
            )
        )
    ).mappings().all()

    listings = []
    for row in rows:
        item = dict(row)
        views = int(item.get("rea_views") or 0)
        enquiries = int(item.get("rea_enquiries") or 0)
        item["ctr"] = round(enquiries / max(views, 1) * 100, 2) if views > 0 else 0
        if item.get("rea_listing_id"):
            item["status"] = "live"
        elif item.get("rea_upload_id"):
            item["status"] = "pending"
        else:
            item["status"] = "ready"
        listings.append(item)

    return {
        "count": len(listings),
        "listings": listings,
        "summary": {
            "total": len(listings),
            "live": sum(1 for i in listings if i["status"] == "live"),
            "pending": sum(1 for i in listings if i["status"] == "pending"),
            "ready": sum(1 for i in listings if i["status"] == "ready"),
            "total_views": sum(int(i.get("rea_views") or 0) for i in listings),
            "total_enquiries": sum(int(i.get("rea_enquiries") or 0) for i in listings),
        },
    }


@router.patch("/listings/{lead_id}")
async def agent_edit_listing(
    lead_id: str,
    body: EditListingRequest,
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    from core.utils import now_iso

    row = (
        await session.execute(
            text("SELECT rea_listing_id, rea_last_edit_at FROM leads WHERE id = :id"),
            {"id": lead_id},
        )
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Lead not found")

    updates_sql = []
    params: dict = {"id": lead_id, "now": now_iso()}

    if body.headline is not None:
        updates_sql.append("listing_headline = :headline")
        params["headline"] = body.headline
    if body.description is not None:
        updates_sql.append("listing_description = :description")
        params["description"] = body.description
    if not updates_sql:
        raise HTTPException(status_code=400, detail="Nothing to update")

    updates_sql.append("rea_last_edit_at = :now")
    updates_sql.append("updated_at = :now")
    await session.execute(text(f"UPDATE leads SET {', '.join(updates_sql)} WHERE id = :id"), params)
    await session.commit()

    rea_listing_id = row.get("rea_listing_id") or row.get("rea_listing_id")
    rea_result = None
    if rea_listing_id:
        try:
            from services.rea_service import update_listing as rea_update

            rea_updates = {}
            if body.headline is not None:
                rea_updates["headline"] = body.headline
            if body.description is not None:
                rea_updates["description"] = body.description
            rea_result = await rea_update(rea_listing_id, rea_updates, session=session, lead_id=lead_id)
        except Exception as exc:
            logger.warning("REA update failed for %s: %s", lead_id, exc)
            rea_result = {"ok": False, "error": str(exc)}

    return {
        "ok": True,
        "lead_id": lead_id,
        "rea_synced": bool(rea_result and rea_result.get("ok")),
        "rea_result": rea_result,
    }


_SUBURB_MAP = {
    "box hill": ("Box Hill", "2765"),
    "marsden park": ("Marsden Park", "2765"),
    "the ponds": ("The Ponds", "2769"),
    "tallawong": ("Tallawong", "2762"),
    "rouse hill": ("Rouse Hill", "2155"),
    "riverstone": ("Riverstone", "2765"),
    "north kellyville": ("North Kellyville", "2155"),
    "kellyville": ("Kellyville", "2155"),
    "oran park": ("Oran Park", "2570"),
    "dora creek": ("Dora Creek", "2264"),
    "vineyard": ("Vineyard", "2765"),
    "blacktown": ("Blacktown", "2148"),
    "nelson road": ("Box Hill", "2765"),
    "old pitt town road": ("Box Hill", "2765"),
    "schofields": ("Schofields", "2762"),
    "leppington": ("Leppington", "2179"),
    "austral": ("Austral", "2179"),
}


def _extract_suburb(project_name: str) -> tuple[str, str]:
    pn = (project_name or "").lower()
    for key, (suburb, postcode) in _SUBURB_MAP.items():
        if key in pn:
            return suburb, postcode
    return "", ""


@router.post("/import-staging")
async def agent_import_staging(
    file: UploadFile = File(...),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Import Bathla staging XLSX/CSV into leads table with all land fields."""
    raw = await file.read()
    filename = (file.filename or "").strip().lower()
    rows: list[dict] = []

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row_vals))
            if any(v for v in d.values() if v):
                rows.append(d)
        wb.close()
    else:
        import csv as csv_mod

        payload = raw.decode("utf-8-sig", errors="ignore")
        reader = csv_mod.DictReader(io.StringIO(payload))
        for row in reader:
            if any(v for v in (row or {}).values() if v and str(v).strip()):
                rows.append(row)

    now = datetime.now(timezone.utc).isoformat()
    created = 0
    updated = 0
    skipped = 0
    last_project_name = ""

    for d in rows:
        address = str(d.get("address") or "").strip()
        lot_number = str(d.get("lot_number") or d.get("lot") or "").strip()
        project_name = str(d.get("project_name") or d.get("project") or "").strip()
        if project_name:
            last_project_name = project_name
        elif not project_name:
            project_name = last_project_name

        if not address and project_name and lot_number:
            pn_clean = re.sub(r"\s+(land|lot|estate)$", "", project_name.strip(), flags=re.IGNORECASE).strip()
            for key in _SUBURB_MAP:
                if pn_clean.lower().endswith(key):
                    pn_clean = pn_clean[: -(len(key))].strip()
                    break
            address = f"Lot {lot_number}/{pn_clean}".strip()

        if not address:
            skipped += 1
            continue

        suburb, postcode = _extract_suburb(project_name)
        land_area = float(d.get("land_area") or d.get("land_size_sqm") or d.get("land_size") or 0)
        price = float(d.get("price") or d.get("estimated_value_mid") or 0)
        lot_type = str(d.get("lot_type") or "")
        frontage = str(d.get("frontage") or "")
        status = str(d.get("status") or "Available")
        ready = str(d.get("ready_for_reaxml") or "").strip().upper() in {"YES", "Y", "TRUE", "1"}

        if status != "Available" or not ready:
            skipped += 1
            continue

        lead_id = hashlib.md5(f"bathla:{address}:{suburb}".encode()).hexdigest()
        existing = (
            await session.execute(text("SELECT id FROM leads WHERE id = :id"), {"id": lead_id})
        ).mappings().first()

        if existing:
            await session.execute(
                text(
                    """
                    UPDATE leads SET
                        land_size_sqm = :land_size, property_type = 'Land',
                        estimated_value_mid = :price, lot_number = :lot_number,
                        lot_type = :lot_type, frontage = :frontage,
                        project_name = :project_name, trigger_type = 'bathla_land',
                        updated_at = :now
                    WHERE id = :id
                    """
                ),
                {
                    "id": lead_id,
                    "land_size": land_area,
                    "price": price,
                    "lot_number": lot_number,
                    "lot_type": lot_type,
                    "frontage": frontage,
                    "project_name": project_name,
                    "now": now,
                },
            )
            updated += 1
        else:
            await session.execute(
                text(
                    """
                    INSERT INTO leads (
                        id, address, suburb, postcode, property_type,
                        record_type, land_size_sqm, estimated_value_mid,
                        lot_number, lot_type, frontage, project_name,
                        trigger_type, signal_status, heat_score,
                        status, created_at, updated_at
                    ) VALUES (
                        :id, :address, :suburb, :postcode, 'Land',
                        'property_record', :land_size, :price,
                        :lot_number, :lot_type, :frontage, :project_name,
                        'bathla_land', 'AVAILABLE', 50,
                        'captured', :now, :now
                    )
                    ON CONFLICT (id) DO UPDATE SET
                        land_size_sqm = excluded.land_size_sqm,
                        estimated_value_mid = excluded.estimated_value_mid,
                        property_type = 'Land',
                        updated_at = excluded.updated_at
                    """
                ),
                {
                    "id": lead_id,
                    "address": address,
                    "suburb": suburb,
                    "postcode": postcode,
                    "land_size": land_area,
                    "price": price,
                    "lot_number": lot_number,
                    "lot_type": lot_type,
                    "frontage": frontage,
                    "project_name": project_name,
                    "now": now,
                },
            )
            created += 1

    await session.commit()
    return {
        "status": "ok",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated,
    }


@router.post("/sync-rea")
async def agent_sync_rea(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """Pull live listings from REA Partner API and upsert into leads table."""
    from services.rea_service import sync_agency_feed_detailed

    feed = await sync_agency_feed_detailed(
        status="current",
        listing_types="land,residential",
        allow_seller_fallback=False,
        session=session,
    )
    if not feed.get("ok"):
        raise HTTPException(status_code=502, detail=feed.get("error", "REA sync failed"))

    listings = feed.get("listings", [])
    now = datetime.now(timezone.utc).isoformat()
    created = 0
    updated = 0

    for item in listings:
        rea_listing_id = item.get("rea_listing_id", "")
        address = item.get("address", "")
        if not address:
            continue

        lead_id = hashlib.md5(f"rea:{rea_listing_id or address}".encode()).hexdigest()
        existing = (
            await session.execute(
                text("SELECT id FROM leads WHERE rea_listing_id = :rea_id OR id = :lid"),
                {"rea_id": rea_listing_id, "lid": lead_id},
            )
        ).mappings().first()

        is_land = item.get("rea_listing_type", "").lower() in ("land",) or "land" in (item.get("property_type") or "").lower()
        property_type = "Land" if is_land else item.get("property_type", "residential")

        if existing:
            await session.execute(
                text("""
                    UPDATE leads SET
                        address = COALESCE(NULLIF(:address, ''), address),
                        suburb = COALESCE(NULLIF(:suburb, ''), suburb),
                        postcode = COALESCE(NULLIF(:postcode, ''), postcode),
                        property_type = :property_type,
                        land_size_sqm = COALESCE(:land_size, land_size_sqm),
                        estimated_value_mid = COALESCE(:est_value, estimated_value_mid),
                        listing_headline = COALESCE(NULLIF(:headline, ''), listing_headline),
                        listing_description = COALESCE(NULLIF(:description, ''), listing_description),
                        main_image = COALESCE(NULLIF(:main_image, ''), main_image),
                        property_images = COALESCE(NULLIF(:property_images, '[]'), property_images),
                        rea_listing_id = COALESCE(NULLIF(:rea_listing_id, ''), rea_listing_id),
                        signal_status = :signal_status,
                        agent_name = COALESCE(NULLIF(:agent_name, ''), agent_name),
                        updated_at = :now
                    WHERE id = :id
                """),
                {
                    "id": existing["id"],
                    "address": address,
                    "suburb": item.get("suburb", ""),
                    "postcode": item.get("postcode", ""),
                    "property_type": property_type,
                    "land_size": item.get("land_size_sqm"),
                    "est_value": item.get("est_value"),
                    "headline": item.get("listing_headline", ""),
                    "description": item.get("listing_description", ""),
                    "main_image": item.get("main_image", ""),
                    "property_images": item.get("property_images", "[]"),
                    "rea_listing_id": rea_listing_id,
                    "signal_status": item.get("signal_status", "LIVE"),
                    "agent_name": item.get("agent", ""),
                    "now": now,
                },
            )
            updated += 1
        else:
            await session.execute(
                text("""
                    INSERT INTO leads (
                        id, address, suburb, postcode, property_type,
                        record_type, land_size_sqm, estimated_value_mid,
                        listing_headline, listing_description,
                        main_image, property_images,
                        rea_listing_id, trigger_type, signal_status,
                        heat_score, status, agent_name,
                        created_at, updated_at
                    ) VALUES (
                        :id, :address, :suburb, :postcode, :property_type,
                        'property_record', :land_size, :est_value,
                        :headline, :description,
                        :main_image, :property_images,
                        :rea_listing_id, 'rea_feed', :signal_status,
                        50, 'captured', :agent_name,
                        :now, :now
                    )
                    ON CONFLICT (id) DO UPDATE SET
                        rea_listing_id = excluded.rea_listing_id,
                        signal_status = excluded.signal_status,
                        updated_at = excluded.updated_at
                """),
                {
                    "id": lead_id,
                    "address": address,
                    "suburb": item.get("suburb", ""),
                    "postcode": item.get("postcode", ""),
                    "property_type": property_type,
                    "land_size": item.get("land_size_sqm"),
                    "est_value": item.get("est_value"),
                    "headline": item.get("listing_headline", ""),
                    "description": item.get("listing_description", ""),
                    "main_image": item.get("main_image", ""),
                    "property_images": item.get("property_images", "[]"),
                    "rea_listing_id": rea_listing_id,
                    "signal_status": item.get("signal_status", "LIVE"),
                    "agent_name": item.get("agent", ""),
                    "now": now,
                },
            )
            created += 1

    await session.commit()
    return {
        "status": "ok",
        "created": created,
        "updated": updated,
        "total_from_rea": len(listings),
        "source": feed.get("source", "listing_export"),
    }


LAND_LISTING_TEMPLATES = [
    {
        "id": "location_value",
        "name": "Location + Value",
        "headline": "{suburb} Land - {size}sqm, Lot {lot_number}, Ready to Build",
        "description": (
            "Premium {size}sqm block in {suburb}'s newest release. "
            "Lot {lot_number} offers a {frontage}m frontage on a quiet, family-friendly street "
            "with all services connected and ready to build.\n\n"
            "Minutes to {suburb} Town Centre, schools, parks, and the new metro line. "
            "This is one of the last opportunities to secure land in a growth corridor "
            "that's seen 15%+ capital gains over the past 12 months.\n\n"
            "Registered and titled - no waiting. Build your dream home or invest "
            "in one of Sydney's strongest-performing land corridors.\n\n"
            "Contact Nitin Puri - Laing+Simmons Oakville | Windsor\n"
            "0485 857 881 | oakville@lsre.com.au"
        ),
    },
    {
        "id": "urgency_scarcity",
        "name": "Urgency + Scarcity",
        "headline": "Last Lots in {suburb} - {size}sqm From ${price_display}",
        "description": (
            "Don't miss out - only a handful of lots remain in {suburb}'s most "
            "sought-after release. This {size}sqm registered block (Lot {lot_number}) "
            "is ready to build on today.\n\n"
            "With Sydney's land supply tightening and prices rising month-on-month, "
            "this is your window to lock in at current pricing. Similar lots in "
            "this estate have already sold.\n\n"
            "Key details:\n"
            "- {size}sqm registered land\n"
            "- {frontage}m frontage\n"
            "- All services connected\n"
            "- Walk to schools, shops, and transport\n"
            "- No developer restrictions - build what you want\n\n"
            "Nitin Puri - Laing+Simmons Oakville | Windsor\n"
            "0485 857 881 | oakville@lsre.com.au"
        ),
    },
    {
        "id": "dream_home",
        "name": "Dream Home Builder",
        "headline": "Build Your Dream Home - {size}sqm in {suburb}",
        "description": (
            "Picture this: your own home, designed exactly the way you want it, "
            "on a {size}sqm canvas in one of Sydney's fastest-growing suburbs.\n\n"
            "Lot {lot_number} in {suburb} is registered, serviced, and waiting "
            "for your vision. Whether it's a single-storey family home or a "
            "contemporary two-storey design, this block gives you the space "
            "and flexibility to build without compromise.\n\n"
            "{suburb} offers everything a growing family needs - top-rated schools, "
            "new shopping centres, parklands, and easy access to the motorway "
            "and upcoming metro stations.\n\n"
            "First home buyers: you may be eligible for the $10,000 First Home "
            "Owner Grant plus stamp duty savings on land.\n\n"
            "Nitin Puri - Laing+Simmons Oakville | Windsor\n"
            "0485 857 881 | oakville@lsre.com.au"
        ),
    },
    {
        "id": "investor_yield",
        "name": "Investor Focus",
        "headline": "{suburb} Growth Corridor - {size}sqm Land, High Demand Area",
        "description": (
            "Smart investors are targeting {suburb} for good reason. This "
            "{size}sqm registered lot sits in the heart of Sydney's north-west "
            "growth corridor - where infrastructure spend is driving sustained "
            "capital growth.\n\n"
            "The numbers speak for themselves:\n"
            "- Median land prices up 12-18% year-on-year\n"
            "- New metro station opening within 2km\n"
            "- Major town centre development approved\n"
            "- Strong rental demand from young professionals and families\n\n"
            "Build a dual-income property or a quality family home that tenants "
            "will compete for. Either way, this is land banking in a corridor "
            "that rewards early movers.\n\n"
            "Nitin Puri - Laing+Simmons Oakville | Windsor\n"
            "0485 857 881 | oakville@lsre.com.au"
        ),
    },
]


@router.get("/templates")
async def agent_listing_templates(api_key: str = Depends(get_api_key)):
    return {"templates": LAND_LISTING_TEMPLATES}


# ---------------------------------------------------------------------------
# CAMPAIGN ENGINE — 151-listing rollout with A/B testing
# ---------------------------------------------------------------------------

# Suburb priority tiers — push highest-demand suburbs first to maximise
# early visibility and gather performance data in the best markets.
SUBURB_PRIORITY = {
    "Box Hill": 1,       # Highest volume, strongest demand
    "North Kellyville": 1,
    "Tallawong": 2,
    "Rouse Hill": 2,
    "Vineyard": 2,
    "The Ponds": 3,
    "Marsden Park": 3,
    "Oran Park": 3,
    "Riverstone": 4,
    "Dora Creek": 4,
    "Blacktown": 5,
    "Schofields": 3,
    "Leppington": 4,
    "Austral": 4,
}

# 8 headline templates — each listing gets ONE variant. Within each suburb,
# variants are evenly distributed so we get a clean A/B comparison.
HEADLINE_VARIANTS = [
    # Location-first (high intent searchers)
    "{suburb} Land — {size}sqm, Lot {lot}, Ready to Build",
    "Lot {lot} {suburb} — {size}sqm Flat Block, {lot_type_tag}",
    "{size}sqm in {suburb} — Lot {lot}, From {price_display}",
    # Benefit-first (emotional buyers)
    "Build Your Dream Home — {size}sqm in {suburb}",
    "Premium {size}sqm Land — Lot {lot}, {suburb} Estate",
    # Urgency (fear of missing out)
    "Last Lots in {suburb} — {size}sqm From {price_display}",
    # Investor angle
    "{suburb} Growth Corridor — {size}sqm Land, High Demand",
    # Lifestyle
    "Family Land in {suburb} — {size}sqm, Walk to Schools & Parks",
]

DESC_VARIANTS = [
    # A: Feature-focused
    """{address}, {suburb} NSW {postcode} — Lot {lot}

A premium {size}sqm block in one of Sydney's fastest-growing corridors.

What you get:
• {size} square metres of flat, build-ready land
{lot_type_line}
{frontage_line}
• All services available
• Walk to future shops, schools, and parklands
• Strong capital growth area — {suburb} is booming

{price_line}. Land in {suburb} is moving fast — secure your block before it's gone.

Nitin Puri | Laing+Simmons Oakville | Windsor
0430 042 041 | oakville@lsre.com.au""",

    # B: Story-driven
    """Lot {lot} at {address}, {suburb}

If you've been looking for the right block to build on, this might be it.

{size} square metres of flat land in {suburb} — a suburb that's transformed over the past few years from farmland into a thriving community with new schools, parks, and retail.

The block:
• {size}sqm, level and cleared
{lot_type_line}
{frontage_line}
• Ready for your builder — no clearing or levelling needed

{suburb} is one of those suburbs where early buyers have done very well. This is your chance to get in while blocks are still available.

{price_line}.

Call Nitin on 0430 042 041 to discuss.
Laing+Simmons Oakville | Windsor""",

    # C: Data-driven
    """{address}, {suburb} NSW {postcode}

Lot {lot} | {size}sqm | {suburb}

Block details:
• Land area: {size} square metres
{lot_type_line}
{frontage_line}
• Services: Available
• Status: Ready to build

Location highlights:
• {suburb} median land price trending upward
• New infrastructure and amenities being delivered
• Easy access to M7 and future metro connections

{price_line}. Genuine enquiries welcome.

Nitin Puri
Laing+Simmons Oakville | Windsor
0430 042 041""",
]


def _fill_template(tpl: str, lot: dict) -> str:
    """Fill a headline or description template with lot data."""
    size = int(float(lot.get("land_size_sqm") or 0))
    price = int(float(lot.get("estimated_value_mid") or lot.get("price") or 0))
    lot_type = lot.get("lot_type") or ""
    frontage = lot.get("frontage") or ""
    return tpl.format(
        suburb=lot.get("suburb", ""),
        size=size,
        lot=lot.get("lot_number", ""),
        address=lot.get("address", ""),
        postcode=lot.get("postcode", ""),
        price_display=f"${price:,}" if price else "Contact Agent",
        lot_type_tag=lot_type if lot_type else "Level Block",
        lot_type_line=f"• {lot_type}" if lot_type else "• Level, regular-shaped block",
        frontage_line=f"• {frontage}m frontage" if frontage else "",
        price_line=f"Priced at ${price:,}" if price else "Contact agent for pricing",
    )


def _with_disclaimer(description: str) -> str:
    base = (description or "").strip()
    if not base:
        return LISTING_COPY_DISCLAIMER
    if LISTING_COPY_DISCLAIMER in base:
        return base
    return f"{base}\n\n{LISTING_COPY_DISCLAIMER}"


def _build_hero_image_prompt(lot: dict, headline: str, description: str) -> str:
    address = str(lot.get("address") or "").strip()
    suburb = str(lot.get("suburb") or "").strip()
    postcode = str(lot.get("postcode") or "").strip()
    size = int(float(lot.get("land_size_sqm") or 0)) if lot.get("land_size_sqm") else 0
    price = int(float(lot.get("estimated_value_mid") or lot.get("price") or 0)) if (lot.get("estimated_value_mid") or lot.get("price")) else 0
    price_display = f"${price:,}" if price else "Contact Agent"
    size_display = f"{size} sqm" if size > 0 else "Land"
    return (
        "Create a polished real-estate hero image (JPG, 1600x900) for a land listing.\n"
        "Use the supplied satellite reference image as the base.\n"
        "Style direction: high-attention thumbnail energy (MrBeast-style composition), but professional real-estate branding and compliant tone.\n"
        "Keep the full background monochrome/desaturated except the target parcel zone.\n"
        "Target parcel treatment: premium golden border + subtle glow, high clarity, no misleading map edits.\n"
        "Typography hierarchy (clean, legible, ad-safe):\n"
        f"1) Primary headline: \"{headline}\"\n"
        f"2) Secondary line: \"{address}, {suburb} {postcode}. Land listing.\"\n"
        f"3) Price badge: \"{price_display}\"\n"
        f"4) Size badge: \"{size_display}\"\n"
        "Use strong contrast, safe margins, and minimal clutter. No fake claims, no prohibited urgency language, no guarantee wording.\n"
        "Output: clean ad-ready JPG only.\n"
        f"Context description for alignment: {description[:220]}"
    )


class CampaignPlanRequest(BaseModel):
    daily_limit: int = 15
    prioritize_corner_blocks: bool = True


@router.post("/campaign-plan")
async def agent_campaign_plan(
    body: CampaignPlanRequest = CampaignPlanRequest(),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Generate a full rollout plan for 151 listings.

    Strategy:
    - Suburb-priority ordering (highest demand first)
    - Corner blocks get pushed earlier within each suburb
    - 8 headline variants evenly distributed within each suburb
    - 3 description variants rotated
    - Daily batches of 15 (REA-compliant)
    - Each listing gets a deterministic variant assignment
    """
    rows = (
        await session.execute(
            text("""
                SELECT id, address, suburb, postcode, land_size_sqm,
                       estimated_value_mid, lot_number, lot_type, frontage,
                       project_name, rea_listing_id, rea_upload_id,
                       rea_title_variant, listing_headline, listing_description
                FROM leads
                WHERE (LOWER(COALESCE(property_type, '')) = 'land'
                       OR LOWER(COALESCE(trigger_type, '')) IN ('bathla_land', 'rea_feed'))
                ORDER BY suburb, address
            """)
        )
    ).mappings().all()

    # Split into already-live and unpushed
    already_live = [dict(r) for r in rows if r.get("rea_listing_id")]
    unpushed = [dict(r) for r in rows if not r.get("rea_listing_id") and not r.get("rea_upload_id")]

    # Sort unpushed by: suburb priority → corner blocks first → lot number
    def sort_key(lot):
        suburb = (lot.get("suburb") or "").strip()
        priority = SUBURB_PRIORITY.get(suburb, 5)
        is_corner = 1 if (lot.get("lot_type") or "").lower() == "corner block" else 2
        if not body.prioritize_corner_blocks:
            is_corner = 1
        lot_num = lot.get("lot_number") or "999"
        return (priority, is_corner, suburb, lot_num)

    unpushed.sort(key=sort_key)

    # Assign variants — even distribution within each suburb
    suburb_counters: dict[str, int] = {}
    num_h_variants = len(HEADLINE_VARIANTS)
    num_d_variants = len(DESC_VARIANTS)

    for lot in unpushed:
        suburb = (lot.get("suburb") or "").strip()
        idx = suburb_counters.get(suburb, 0)
        suburb_counters[suburb] = idx + 1
        lot["_h_variant"] = idx % num_h_variants
        lot["_d_variant"] = idx % num_d_variants

    # Build daily batches
    daily_limit = max(1, min(body.daily_limit, 15))  # Cap at 15
    days = []
    for batch_start in range(0, len(unpushed), daily_limit):
        batch = unpushed[batch_start : batch_start + daily_limit]
        day_num = len(days) + 1
        items = []
        for lot in batch:
            h_variant = lot["_h_variant"]
            d_variant = lot["_d_variant"]
            headline = _fill_template(HEADLINE_VARIANTS[h_variant], lot)
            description = _with_disclaimer(_fill_template(DESC_VARIANTS[d_variant], lot))
            items.append({
                "lead_id": lot["id"],
                "address": lot.get("address", ""),
                "suburb": lot.get("suburb", ""),
                "postcode": lot.get("postcode", ""),
                "land_size": lot.get("land_size_sqm"),
                "price": lot.get("estimated_value_mid"),
                "lot_number": lot.get("lot_number", ""),
                "lot_type": lot.get("lot_type", ""),
                "headline_variant": h_variant,
                "desc_variant": d_variant,
                "headline": headline,
                "description_preview": description[:200] + "..." if len(description) > 200 else description,
                "hero_image_prompt": _build_hero_image_prompt(lot, headline, description),
            })
        days.append({
            "day": day_num,
            "count": len(items),
            "suburbs": list({i["suburb"] for i in items if i["suburb"]}),
            "listings": items,
        })

    # Variant distribution summary
    variant_dist: dict[int, dict] = {}
    for lot in unpushed:
        v = lot["_h_variant"]
        if v not in variant_dist:
            variant_dist[v] = {"variant": v, "template": HEADLINE_VARIANTS[v][:60], "count": 0, "suburbs": set()}
        variant_dist[v]["count"] += 1
        variant_dist[v]["suburbs"].add(lot.get("suburb", ""))
    for v in variant_dist.values():
        v["suburbs"] = sorted(v["suburbs"])

    # Suburb breakdown
    suburb_summary = {}
    for lot in unpushed:
        s = lot.get("suburb", "Unknown")
        if s not in suburb_summary:
            suburb_summary[s] = {"suburb": s, "count": 0, "priority": SUBURB_PRIORITY.get(s, 5), "corner_blocks": 0}
        suburb_summary[s]["count"] += 1
        if (lot.get("lot_type") or "").lower() == "corner block":
            suburb_summary[s]["corner_blocks"] += 1

    return {
        "status": "ok",
        "summary": {
            "total_to_push": len(unpushed),
            "already_live": len(already_live),
            "days_needed": len(days),
            "daily_limit": daily_limit,
            "headline_variants": num_h_variants,
            "desc_variants": num_d_variants,
        },
        "suburb_breakdown": sorted(suburb_summary.values(), key=lambda s: (s["priority"], s["suburb"])),
        "variant_distribution": sorted(variant_dist.values(), key=lambda v: v["variant"]),
        "schedule": days,
        "strategy_notes": [
            f"Push {daily_limit} listings per day over {len(days)} days",
            "Highest-demand suburbs (Box Hill, North Kellyville) go first",
            "Corner blocks prioritised within each suburb",
            f"{num_h_variants} headline variants evenly distributed per suburb for A/B testing",
            f"{num_d_variants} description variants rotated for variety",
            "After 7 days live, underperformers get refreshed with winning variant style",
            "Max 1 edit per listing per 24 hours (REA compliant)",
            "All land listings are FREE on REA",
        ],
    }


@router.get("/campaign-status")
async def agent_campaign_status(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Show campaign rollout progress + variant A/B performance + refresh recommendations.
    """
    rows = (
        await session.execute(
            text("""
                SELECT id, address, suburb, postcode, land_size_sqm,
                       estimated_value_mid, lot_number, lot_type, frontage,
                       project_name, rea_listing_id, rea_upload_id, rea_upload_status,
                       rea_title_variant, rea_views, rea_enquiries,
                       rea_last_edit_at, listing_headline, created_at, updated_at
                FROM leads
                WHERE (LOWER(COALESCE(property_type, '')) = 'land'
                       OR LOWER(COALESCE(trigger_type, '')) IN ('bathla_land', 'rea_feed'))
                ORDER BY suburb, address
            """)
        )
    ).mappings().all()

    live = []
    pending = []
    unpushed = []
    for r in rows:
        d = dict(r)
        if d.get("rea_listing_id"):
            d["status"] = "live"
            live.append(d)
        elif d.get("rea_upload_id"):
            d["status"] = "pending"
            pending.append(d)
        else:
            d["status"] = "ready"
            unpushed.append(d)

    # Variant performance (for live listings)
    variant_perf: dict[int, dict] = {}
    for l in live:
        v = int(l.get("rea_title_variant") or 0)
        views = int(l.get("rea_views") or 0)
        enq = int(l.get("rea_enquiries") or 0)
        if v not in variant_perf:
            variant_perf[v] = {
                "variant": v,
                "template": HEADLINE_VARIANTS[v % len(HEADLINE_VARIANTS)][:60] if v < len(HEADLINE_VARIANTS) else "Unknown",
                "listings": 0,
                "total_views": 0,
                "total_enquiries": 0,
            }
        variant_perf[v]["listings"] += 1
        variant_perf[v]["total_views"] += views
        variant_perf[v]["total_enquiries"] += enq

    for v in variant_perf.values():
        v["avg_views"] = round(v["total_views"] / max(v["listings"], 1), 1)
        v["avg_enquiries"] = round(v["total_enquiries"] / max(v["listings"], 1), 1)
        v["ctr"] = round(v["total_enquiries"] / max(v["total_views"], 1) * 100, 2) if v["total_views"] > 0 else 0

    # Suburb performance
    suburb_perf: dict[str, dict] = {}
    for l in live:
        s = l.get("suburb", "Unknown")
        views = int(l.get("rea_views") or 0)
        enq = int(l.get("rea_enquiries") or 0)
        if s not in suburb_perf:
            suburb_perf[s] = {"suburb": s, "live": 0, "total_views": 0, "total_enquiries": 0}
        suburb_perf[s]["live"] += 1
        suburb_perf[s]["total_views"] += views
        suburb_perf[s]["total_enquiries"] += enq
    for s in suburb_perf.values():
        s["avg_views"] = round(s["total_views"] / max(s["live"], 1), 1)
        s["ctr"] = round(s["total_enquiries"] / max(s["total_views"], 1) * 100, 2) if s["total_views"] > 0 else 0

    # Refresh recommendations — listings live 7+ days with below-average views
    now = datetime.now(timezone.utc)
    avg_views = sum(int(l.get("rea_views") or 0) for l in live) / max(len(live), 1)
    refresh_recs = []
    for l in live:
        views = int(l.get("rea_views") or 0)
        last_edit = l.get("rea_last_edit_at") or l.get("updated_at") or ""
        if last_edit:
            try:
                edit_dt = datetime.fromisoformat(last_edit.replace("Z", "+00:00"))
                days_since = (now - edit_dt).days
            except Exception:
                days_since = 0
        else:
            days_since = 0

        if days_since >= 7 and views < avg_views * 0.7:
            current_variant = int(l.get("rea_title_variant") or 0)
            # Find best performing variant
            best_variant = max(variant_perf.values(), key=lambda v: v["ctr"]) if variant_perf else None
            new_variant = best_variant["variant"] if best_variant and best_variant["variant"] != current_variant else (current_variant + 1) % len(HEADLINE_VARIANTS)
            refresh_recs.append({
                "lead_id": l["id"],
                "address": l.get("address", ""),
                "suburb": l.get("suburb", ""),
                "current_views": views,
                "avg_views": round(avg_views, 1),
                "days_since_edit": days_since,
                "current_variant": current_variant,
                "recommended_variant": new_variant,
                "recommended_headline": _fill_template(HEADLINE_VARIANTS[new_variant % len(HEADLINE_VARIANTS)], l),
                "reason": f"Below average views ({views} vs {avg_views:.0f} avg) after {days_since} days",
            })

    return {
        "progress": {
            "total": len(rows),
            "live": len(live),
            "pending": len(pending),
            "ready": len(unpushed),
            "pct_complete": round(len(live) / max(len(rows), 1) * 100, 1),
        },
        "variant_performance": sorted(variant_perf.values(), key=lambda v: -v["ctr"]),
        "suburb_performance": sorted(suburb_perf.values(), key=lambda s: -s["ctr"]),
        "refresh_recommendations": refresh_recs,
        "next_actions": _campaign_next_actions(live, pending, unpushed),
    }


def _campaign_next_actions(live: list, pending: list, unpushed: list) -> list[str]:
    """Generate contextual next-action suggestions."""
    actions = []
    if unpushed:
        actions.append(f"Push next batch — {len(unpushed)} listings waiting, push up to 15 today")
    if pending:
        actions.append(f"Check upload status — {len(pending)} listings pending REA processing")
    if live and all(int(l.get("rea_views") or 0) == 0 for l in live):
        actions.append("Pull performance data — live listings have no view counts yet")
    if len(live) >= 10:
        actions.append("Review A/B results — enough data to compare variant performance")
    if not unpushed and not pending:
        actions.append("All listings deployed — focus on refresh cycle and performance monitoring")
    return actions


@router.post("/campaign-execute-day")
async def agent_campaign_execute_day(
    body: CampaignPlanRequest = CampaignPlanRequest(),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Execute one day's worth of the campaign — assign variants, save headlines,
    and mark listings as ready for push. Does NOT publish to REA (that requires
    REA_AUTO_PUBLISH_ENABLED=true and a separate execute-push call).
    """
    # Get unpushed listings sorted by priority
    rows = (
        await session.execute(
            text("""
                SELECT id, address, suburb, postcode, land_size_sqm,
                       estimated_value_mid, lot_number, lot_type, frontage,
                       project_name
                FROM leads
                WHERE (LOWER(COALESCE(property_type, '')) = 'land'
                       OR LOWER(COALESCE(trigger_type, '')) IN ('bathla_land', 'rea_feed'))
                  AND COALESCE(rea_listing_id, '') = ''
                  AND COALESCE(rea_upload_id, '') = ''
                ORDER BY suburb, address
            """)
        )
    ).mappings().all()

    if not rows:
        return {"status": "ok", "prepared": 0, "message": "No unpushed listings remaining"}

    lots = [dict(r) for r in rows]

    # Sort by priority
    def sort_key(lot):
        suburb = (lot.get("suburb") or "").strip()
        priority = SUBURB_PRIORITY.get(suburb, 5)
        is_corner = 0 if (lot.get("lot_type") or "").lower() == "corner block" else 1
        return (priority, is_corner, suburb, lot.get("lot_number") or "999")

    lots.sort(key=sort_key)

    # Take today's batch
    daily_limit = max(1, min(body.daily_limit, 15))
    batch = lots[:daily_limit]

    # Assign variants (check what variants already exist in each suburb to keep balanced)
    suburb_existing = {}
    existing_rows = (
        await session.execute(
            text("""
                SELECT suburb, rea_title_variant FROM leads
                WHERE COALESCE(rea_title_variant, -1) >= 0
                  AND COALESCE(rea_listing_id, '') <> ''
            """)
        )
    ).mappings().all()
    for r in existing_rows:
        s = (r.get("suburb") or "").strip()
        suburb_existing.setdefault(s, []).append(int(r["rea_title_variant"]))

    now = datetime.now(timezone.utc).isoformat()
    prepared = []
    for lot in batch:
        suburb = (lot.get("suburb") or "").strip()
        existing_variants = suburb_existing.get(suburb, [])
        # Pick the least-used variant for this suburb
        variant_counts = {i: existing_variants.count(i) for i in range(len(HEADLINE_VARIANTS))}
        h_variant = min(variant_counts, key=variant_counts.get)
        d_variant = h_variant % len(DESC_VARIANTS)

        headline = _fill_template(HEADLINE_VARIANTS[h_variant], lot)
        description = _with_disclaimer(_fill_template(DESC_VARIANTS[d_variant], lot))

        await session.execute(
            text("""
                UPDATE leads SET
                    listing_headline = :headline,
                    listing_description = :description,
                    rea_title_variant = :h_variant,
                    rea_desc_variant = :d_variant,
                    updated_at = :now
                WHERE id = :id
            """),
            {
                "id": lot["id"],
                "headline": headline,
                "description": description,
                "h_variant": h_variant,
                "d_variant": d_variant,
                "now": now,
            },
        )
        suburb_existing.setdefault(suburb, []).append(h_variant)
        prepared.append({
            "lead_id": lot["id"],
            "address": lot.get("address", ""),
            "suburb": suburb,
            "headline": headline,
            "headline_variant": h_variant,
            "desc_variant": d_variant,
            "hero_image_prompt": _build_hero_image_prompt(lot, headline, description),
        })

    await session.commit()
    return {
        "status": "ok",
        "prepared": len(prepared),
        "remaining": len(lots) - len(prepared),
        "listings": prepared,
        "next_step": "Review headlines above, then call execute-push with these lead_ids to publish to REA",
    }


@router.post("/bulk-generate-copy")
async def agent_bulk_generate_copy(
    body: BulkCopyRequest = BulkCopyRequest(),
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Generate/refresh conversion-focused listing headline + description in bulk.
    Default target is 151 records for campaign rollout.
    """
    limit = max(1, min(int(body.limit or 151), 500))
    live_clause = "" if body.include_live else "AND COALESCE(rea_listing_id, '') = ''"

    rows = (
        await session.execute(
            text(
                f"""
                SELECT id, address, suburb, postcode, land_size_sqm, estimated_value_mid,
                       lot_number, lot_type, frontage, project_name, rea_listing_id
                FROM leads
                WHERE {_LAND_FILTER_SQL}
                  {live_clause}
                ORDER BY suburb, address
                LIMIT :limit
                """
            ),
            {"limit": limit},
        )
    ).mappings().all()

    if not rows:
        return {"status": "ok", "updated": 0, "message": "No land listings matched the bulk copy criteria"}

    now = datetime.now(timezone.utc).isoformat()
    updated = 0
    previews = []
    for row in rows:
        lot = dict(row)
        # Deterministic variant choice by listing id for stable copy.
        digest = int(hashlib.md5(str(lot["id"]).encode()).hexdigest()[:8], 16)
        h_variant = digest % len(HEADLINE_VARIANTS)
        d_variant = digest % len(DESC_VARIANTS)
        headline = _fill_template(HEADLINE_VARIANTS[h_variant], lot)
        description = _fill_template(DESC_VARIANTS[d_variant], lot)
        if body.enforce_disclaimer:
            description = _with_disclaimer(description)

        await session.execute(
            text(
                """
                UPDATE leads
                SET listing_headline = :headline,
                    listing_description = :description,
                    rea_title_variant = :h_variant,
                    rea_desc_variant = :d_variant,
                    updated_at = :now
                WHERE id = :id
                """
            ),
            {
                "id": lot["id"],
                "headline": headline,
                "description": description,
                "h_variant": h_variant,
                "d_variant": d_variant,
                "now": now,
            },
        )
        updated += 1
        if len(previews) < 5:
            previews.append(
                {
                    "lead_id": lot["id"],
                    "address": lot.get("address", ""),
                    "headline": headline,
                    "description_preview": description[:220],
                    "hero_image_prompt": _build_hero_image_prompt(lot, headline, description),
                }
            )

    await session.commit()
    return {
        "status": "ok",
        "updated": updated,
        "requested": limit,
        "disclaimer_enforced": body.enforce_disclaimer,
        "previews": previews,
    }


@router.get("/creative-template")
async def agent_creative_template(
    api_key: str = Depends(get_api_key),
    session: AsyncSession = Depends(get_session),
):
    """
    Return a reusable image prompt template and one sample built from a random land listing.
    """
    row = (
        await session.execute(
            text(
                f"""
                SELECT id, address, suburb, postcode, land_size_sqm, estimated_value_mid,
                       lot_number, lot_type, frontage, listing_headline, listing_description
                FROM leads
                WHERE {_LAND_FILTER_SQL}
                ORDER BY RANDOM()
                LIMIT 1
                """
            )
        )
    ).mappings().first()

    if not row:
        return {
            "status": "ok",
            "message": "No land listing found for sample generation",
            "prompt_template": "Use listing address/suburb/price/size and apply monochrome background + highlighted parcel zone + premium golden border.",
        }

    lot = dict(row)
    headline = (lot.get("listing_headline") or _fill_template(HEADLINE_VARIANTS[0], lot)).strip()
    description = (lot.get("listing_description") or _with_disclaimer(_fill_template(DESC_VARIANTS[0], lot))).strip()

    return {
        "status": "ok",
        "sample": {
            "lead_id": lot["id"],
            "address": lot.get("address", ""),
            "suburb": lot.get("suburb", ""),
            "postcode": lot.get("postcode", ""),
            "headline": headline,
            "description_preview": description[:240],
            "hero_image_prompt": _build_hero_image_prompt(lot, headline, description),
        },
        "prompt_template_notes": [
            "Monochrome background except parcel zone",
            "Gold parcel border + premium glow",
            "High attention composition but professional real-estate style",
            "Strictly factual text overlays; no misleading claims",
        ],
    }
